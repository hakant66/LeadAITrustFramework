# app/routers/admin.py
from __future__ import annotations

import io
import os
import httpx
import hashlib
import re
import logging
from pathlib import Path
from uuid import uuid4
from datetime import date, datetime, timezone
from typing import Any, List, Optional
import json

import asyncpg
from fastapi import APIRouter, BackgroundTasks, HTTPException, File, UploadFile, Request, Depends, Form, Query
from fastapi.responses import StreamingResponse, Response, FileResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from uuid import UUID
from app.db import get_db
from app.services.scorecard_read import fetch_project_pillars 
from app.db_async import get_pool
from app.scorecard import ensure_schema, get_project_id_by_slug
from app.dependencies import (
    get_entity_id_with_auth_viewer,
    get_entity_id_with_auth_editor,
    get_entity_id_or_first_for_viewer,
    get_entity_id_optional,
)
from app.services.s3_client import (
    presign_put,
    presign_get,
    object_uri,
    delete_object,
    upload_stream_to_s3,
    s3_ready,
    get_object,
)
from app.services.provenance import compute_s3_sha256
from app.services.audit_log import append_audit_event
from app.services.data_governance import (
    compute_data_governance_warnings as compute_governance_warnings,
)
from app.services.policy_alerts import compute_policy_alerts
from app.services.knowledge_vault_ingest import (
    ingest_report_source,
    search_knowledge_vault,
)
from app.services.email_service import EmailConfigError, send_control_assignment_email, send_policy_review_email
from app.services.sms_service import SMSConfigError, send_policy_review_sms
from app.services.langfuse_adapter import LangfuseAdapter, LangfuseMetric



router = APIRouter(prefix="/admin", tags=["admin"])
email_logger = logging.getLogger("leadai.email")

# Where to drop uploaded binaries locally (dev-only stub storage)
EVIDENCE_FILE_ROOT = os.getenv("EVIDENCE_FILE_ROOT") or os.getenv(
    "EVIDENCE_LOCAL_ROOT", "/tmp/leadai-evidence"
)
EVIDENCE_FILE_PREFIX = os.getenv("EVIDENCE_FILE_PREFIX") or ""


async def _log_audit(event_type: str, **kwargs) -> None:
    try:
        await append_audit_event(event_type=event_type, **kwargs)
    except Exception:
        # Never break user flows on audit failures
        pass


def _normalize_locale(locale: Optional[str]) -> Optional[str]:
    if not locale:
        return None
    normalized = locale.strip().lower()
    if not normalized:
        return None
    return normalized.split(",")[0].split("-")[0]


async def _get_policy_review_recipients(
    conn: asyncpg.Connection, entity_id: UUID
) -> tuple[dict, set[str], set[str]]:
    entity = await conn.fetchrow(
        """
        SELECT full_legal_name, slug,
               authorized_representative_email,
               authorized_representative_phone,
               ai_compliance_officer_email,
               executive_sponsor_email
        FROM entity
        WHERE id = $1
        """,
        entity_id,
    )
    entity_name = entity["full_legal_name"] if entity else None
    entity_slug = entity["slug"] if entity else None

    emails: set[str] = set()
    phones: set[str] = set()

    for field in [
        entity.get("authorized_representative_email") if entity else None,
        entity.get("ai_compliance_officer_email") if entity else None,
        entity.get("executive_sponsor_email") if entity else None,
    ]:
        if field and str(field).strip():
            emails.add(str(field).strip())

    phone = entity.get("authorized_representative_phone") if entity else None
    if phone and str(phone).strip():
        phones.add(str(phone).strip())

    user_rows = await conn.fetch(
        """
        SELECT DISTINCT u.email
        FROM user_entity_access uea
        JOIN user_mapping um ON um.backend_user_id = uea.user_id
        JOIN auth."User" u ON u.id = um.nextauth_user_id
        WHERE uea.entity_id = $1
          AND uea.role IN ('admin', 'editor')
          AND u.email IS NOT NULL
        """,
        entity_id,
    )
    for row in user_rows:
        if row.get("email") and str(row["email"]).strip():
            emails.add(str(row["email"]).strip())

    meta = {"entity_name": entity_name, "entity_slug": entity_slug}
    return meta, emails, phones


async def _get_entity_id_from_project_slug(
    conn: asyncpg.Connection,
    project_slug: str,
    entity_id: UUID | None = None,
) -> UUID:
    """
    Get entity_id from project_slug, or use provided entity_id.
    Raises HTTPException if project not found or entity_id mismatch.
    """
    if entity_id:
        # Verify project belongs to this entity
        proj_entity_id = await conn.fetchval(
            "SELECT entity_id FROM entity_projects WHERE slug = $1",
            project_slug,
        )
        if not proj_entity_id:
            raise HTTPException(status_code=404, detail=f"Project '{project_slug}' not found")
        if proj_entity_id != entity_id:
            raise HTTPException(
                status_code=403,
                detail=f"Project '{project_slug}' does not belong to the specified entity"
            )
        return entity_id
    
    # Get entity_id from project
    proj_entity_id = await conn.fetchval(
        "SELECT entity_id FROM entity_projects WHERE slug = $1",
        project_slug,
    )
    if not proj_entity_id:
        raise HTTPException(status_code=404, detail=f"Project '{project_slug}' not found")
    return proj_entity_id


async def _create_new_project_requirements_alert(
    conn: asyncpg.Connection, entity_slug: str, project_slug: str, project_name: str
) -> None:
    """
    Create an alert in policy_alerts when a new project is created.
    The alert appears in AI Requirements Register to indicate that requirements
    need to be mapped/assigned to this new project.
    """
    try:
        # Generate policy ID with entity_slug and project_slug
        policy_id = f"{entity_slug}-{project_slug}-system-requirements-policy"
        
        # Find or create a system policy for AI Requirements alerts
        policy_row = await conn.fetchrow(
            """
            SELECT id, title FROM policies
            WHERE id = $1
            """,
            policy_id,
        )
        
        if not policy_row:
            # Create a system policy for requirements alerts if it doesn't exist
            await conn.execute(
                """
                INSERT INTO policies (id, title, owner_role, status, created_at, updated_at)
                VALUES ($1, 'AI Requirements Register', 'system', 'active', NOW(), NOW())
                ON CONFLICT (id) DO NOTHING
                """,
                policy_id,
            )
            policy_title = "AI Requirements Register"
        else:
            policy_id = policy_row["id"]
            policy_title = policy_row["title"]
        
        # Create the alert
        await conn.execute(
            """
            INSERT INTO policy_alerts (
                id, policy_id, policy_title, project_slug,
                alert_type, severity, message, source_type, source_key,
                status, created_at, updated_at, details_json
            )
            VALUES (
                $1, $2, $3, $4,
                'new_project_requirements', 'info',
                $5, 'project', $6,
                'open', NOW(), NOW(), $7::jsonb
            )
            ON CONFLICT (policy_id, alert_type, source_type, source_key, project_slug)
            DO UPDATE SET
                status = 'open',
                updated_at = NOW(),
                resolved_at = NULL,
                message = EXCLUDED.message
            """,
            str(uuid4()),
            policy_id,
            policy_title,
            project_slug,
            f"New project '{project_name}' ({project_slug}) requires requirements mapping in AI Requirements Register",
            project_slug,
            json.dumps({
                "project_slug": project_slug,
                "project_name": project_name,
                "alert_reason": "new_project_created",
            }),
        )
    except Exception:
        # Never break project creation if alert creation fails
        pass


def _as_json(value: object | None) -> str | None:
    if value is None:
        return None
    return json.dumps(value)
LOCAL_EVIDENCE_ROOT = Path(EVIDENCE_FILE_ROOT).resolve()
LOCAL_EVIDENCE_ROOT.mkdir(parents=True, exist_ok=True)


def _normalize_path(p: str) -> str:
    return p.replace("\\", "/").rstrip("/")

def _local_path_from_uri(uri: str) -> Optional[Path]:
    if not uri.startswith("file://"):
        return None
    raw = uri[len("file://") :].lstrip("/")
    if not raw:
        return None

    raw_norm = _normalize_path(raw)
    root_norm = _normalize_path(str(LOCAL_EVIDENCE_ROOT))
    rel = ""

    if EVIDENCE_FILE_PREFIX:
        pref_norm = _normalize_path(EVIDENCE_FILE_PREFIX)
        if raw_norm.lower().startswith(pref_norm.lower()):
            rel = raw_norm[len(pref_norm) :].lstrip("/")

    if not rel:
        parts = [p for p in raw_norm.split("/") if p]
        if "leadai-evidence" in parts:
            idx = parts.index("leadai-evidence")
            rel = "/".join(parts[idx + 1 :])

    if not rel and raw_norm.lower().startswith(root_norm.lower()):
        rel = raw_norm[len(root_norm) :].lstrip("/")

    if not rel:
        rel = raw_norm

    try:
        p = Path(root_norm, rel).resolve()
    except Exception:
        return None

    root = LOCAL_EVIDENCE_ROOT.resolve()
    if p != root and root not in p.parents:
        return None
    return p


def _s3_key_from_uri(uri: str) -> str:
    if not uri.startswith("s3://"):
        return ""
    without = uri[len("s3://"):]
    parts = without.split("/", 1)
    if len(parts) < 2:
        return ""
    return parts[1]


def _s3_bucket_from_uri(uri: str) -> str:
    if not uri.startswith("s3://"):
        return ""
    without = uri[len("s3://"):]
    parts = without.split("/", 1)
    return parts[0] if parts else ""

def _build_evidence_url(base: str, slug: str, project: str) -> str:
    base_clean = base.rstrip("/")
    return f"{base_clean}/{slug}/scorecard/{project}/dashboard/kpis_admin"

def _clean_optional_str(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    trimmed = value.strip()
    return trimmed if trimmed else None


def _slugify(value: str) -> str:
    raw = value.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "-", raw).strip("-")
    return slug or "use-case"


async def _generate_unique_uc_id(
    conn: asyncpg.Connection,
    project_slug: str,
    name: str,
    max_len: int = 64,
) -> str:
    base = _slugify(f"{project_slug}-{name}")
    if len(base) > max_len:
        base = base[:max_len].rstrip("-")
        if not base:
            base = "use-case"

    candidate = base
    suffix = 2
    while True:
        exists = await conn.fetchval(
            "SELECT 1 FROM ai_system_registry WHERE uc_id = $1",
            candidate,
        )
        if not exists:
            return candidate

        suffix_text = f"-{suffix}"
        trimmed = base
        if len(base) + len(suffix_text) > max_len:
            trimmed = base[: max_len - len(suffix_text)].rstrip("-")
            if not trimmed:
                trimmed = "use-case"
        candidate = f"{trimmed}{suffix_text}"
        suffix += 1
        if suffix > 99:
            short = uuid4().hex[:6]
            suffix_text = f"-{short}"
            trimmed = base
            if len(base) + len(suffix_text) > max_len:
                trimmed = base[: max_len - len(suffix_text)].rstrip("-")
                if not trimmed:
                    trimmed = "use-case"
            candidate = f"{trimmed}{suffix_text}"
            exists = await conn.fetchval(
                "SELECT 1 FROM ai_system_registry WHERE uc_id = $1",
                candidate,
            )
            if not exists:
                return candidate
            raise HTTPException(
                status_code=409,
                detail="Unable to generate unique use case reference",
            )


_PROVENANCE_ENTITY_TABLES = {
    "dataset": "provenance_datasets",
    "model": "provenance_models",
    "artifact": "provenance_artifacts",
    "evidence": "provenance_evidence",
}


def _validate_provenance_type(value: str) -> str:
    if value not in _PROVENANCE_ENTITY_TABLES:
        raise HTTPException(status_code=400, detail="Invalid provenance type")
    return value


async def _insert_provenance_audit(
    conn: asyncpg.Connection,
    entity_type: str,
    entity_id: str,
    action: str,
    actor: Optional[str],
    details: Optional[dict],
) -> None:
    payload = json.dumps(details or {})
    await conn.execute(
        """
        INSERT INTO provenance_audit (entity_type, entity_id, action, actor, details_json)
        VALUES ($1, $2, $3, $4, $5)
        """,
        entity_type,
        entity_id,
        action,
        actor,
        payload,
    )


# ---------- Models ----------
class ControlIn(BaseModel):
    control_id: str = Field(min_length=1)
    name: str = Field(min_length=1)
    pillar: Optional[str] = None
    unit: Optional[str] = None
    norm_min: Optional[float] = None
    norm_max: Optional[float] = None
    higher_is_better: bool = True
    weight: float = 1.0
    axis_key: Optional[str] = Field(
        default=None,
        pattern="^(safety|compliance|provenance)$",
    )
    # extended columns from controls table
    target_text: Optional[str] = None
    target_numeric: Optional[int] = None
    evidence_source: Optional[str] = None
    owner_role: Optional[str] = None
    frequency: Optional[int] = None
    failure_action: Optional[int] = None
    maturity_anchor_l3: Optional[int] = None
    current_value: Optional[int] = None
    as_of: Optional[int] = None
    notes: Optional[str] = None
    kpi_score: Optional[int] = None
    description: Optional[str] = None
    example: Optional[str] = None



class ControlOut(ControlIn):
    pass


class TrustAxisMapIn(BaseModel):
    axis_key: Optional[str] = Field(
        default=None,
        pattern="^(safety|compliance|provenance)$",
    )
    notes: Optional[str] = None


class TrustAxisMapOut(TrustAxisMapIn):
    pillar_key: str
    pillar_name: Optional[str] = None


class ProvenanceArtifactIn(BaseModel):
    project_slug: str = Field(min_length=1)
    name: str = Field(min_length=1)
    uri: str = Field(min_length=1)
    size_bytes: Optional[int] = None
    mime: Optional[str] = None
    license_name: Optional[str] = None
    license_url: Optional[str] = None
    usage_rights: Optional[str] = None


class ProvenanceArtifactOut(ProvenanceArtifactIn):
    id: str
    sha256: str
    created_at: Optional[datetime] = None


class EntityProviderArtifactIn(BaseModel):
    provider_key: str = Field(min_length=1)
    name: str = Field(min_length=1)
    uri: str = Field(min_length=1)
    sha256: Optional[str] = None
    type: Optional[str] = None
    status: Optional[str] = None
    valid_from: Optional[datetime] = None
    valid_to: Optional[datetime] = None


class EntityProviderArtifactOut(EntityProviderArtifactIn):
    id: str
    entity_id: UUID
    updated_at: Optional[datetime] = None


class ProvenanceDatasetIn(BaseModel):
    project_slug: str = Field(min_length=1)
    name: str = Field(min_length=1)
    description: Optional[str] = None
    artifact_id: Optional[str] = None


class ProvenanceDatasetOut(ProvenanceDatasetIn):
    id: str
    created_at: Optional[datetime] = None


class ProvenanceModelIn(BaseModel):
    project_slug: str = Field(min_length=1)
    name: str = Field(min_length=1)
    version: Optional[str] = None
    framework: Optional[str] = None
    description: Optional[str] = None
    artifact_id: Optional[str] = None


class ProvenanceModelOut(ProvenanceModelIn):
    id: str
    created_at: Optional[datetime] = None


class ProvenanceEvidenceIn(BaseModel):
    project_slug: str = Field(min_length=1)
    name: str = Field(min_length=1)
    description: Optional[str] = None
    artifact_id: Optional[str] = None


class ProvenanceEvidenceOut(ProvenanceEvidenceIn):
    id: str
    created_at: Optional[datetime] = None


class ProvenanceLineageIn(BaseModel):
    project_slug: str = Field(min_length=1)
    parent_type: str = Field(min_length=1)
    parent_id: str = Field(min_length=1)
    child_type: str = Field(min_length=1)
    child_id: str = Field(min_length=1)
    relationship: Optional[str] = None


class ProvenanceLineageOut(ProvenanceLineageIn):
    id: int
    created_at: Optional[datetime] = None


class ProvenanceAuditOut(BaseModel):
    id: int
    entity_type: str
    entity_id: str
    action: str
    actor: Optional[str] = None
    at: Optional[datetime] = None
    details_json: Optional[dict] = None


class ProvenanceValidateOut(BaseModel):
    artifact_id: str
    sha256: str
    computed_sha256: str
    match: bool


class ProjectIn(BaseModel):
    slug: str = Field(min_length=1)
    name: str = Field(min_length=1)
    risk_level: Optional[str] = None
    target_threshold: float = 0.75  # 0..1
    priority: Optional[str] = None
    sponsor: Optional[str] = None
    owner: Optional[str] = None
    status: Optional[str] = None
    creation_date: Optional[date] = None
    update_date: Optional[datetime] = None
    update_date: Optional[datetime] = None


class ProjectOut(ProjectIn):
    id: str


# ---------- Controls CRUD (global) ----------
@router.get("/controls", response_model=List[ControlOut])  # or drop response_model for flexibility
async def list_controls():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        rows = await conn.fetch(
            """
            SELECT
              id::text         AS control_id,
              kpi_key,
              name,
              pillar,
              unit,
              norm_min,
              norm_max,
              higher_is_better,
              weight,
              axis_key,
              target_text,
              target_numeric,
              evidence_source,
              owner_role,
              frequency,
              failure_action,
              maturity_anchor_l3,
              current_value,
              as_of,
              kpi_score,
              description,
              example,
              notes
            FROM controls
            ORDER BY pillar ASC NULLS LAST, kpi_key ASC
            """
        )
        return [dict(r) for r in rows]



@router.post("/controls", response_model=ControlOut)
async def create_control(body: ControlIn) -> ControlOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        try:
            await conn.execute(
                """
                INSERT INTO controls
                  (id, name, pillar, unit, norm_min, norm_max, higher_is_better, weight, axis_key)
                VALUES ($1::uuid, $2, $3, $4, $5, $6, $7, $8, $9)
                """,
                body.control_id, body.name, body.pillar, body.unit,
                body.norm_min, body.norm_max, body.higher_is_better, body.weight,
                body.axis_key,
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(status_code=409, detail="control_id already exists")
        return ControlOut(**body.model_dump())


@router.put("/controls/{control_id}", response_model=ControlOut)
async def update_control(control_id: str, body: ControlIn) -> ControlOut:
    if control_id != body.control_id:
        raise HTTPException(status_code=400, detail="control_id mismatch between path and body")

    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        result = await conn.execute(
            """
            UPDATE controls
            SET name=$2, pillar=$3, unit=$4, norm_min=$5, norm_max=$6,
                higher_is_better=$7, weight=$8, axis_key=$9
            WHERE id=$1::uuid
            """,
            body.control_id, body.name, body.pillar, body.unit,
            body.norm_min, body.norm_max, body.higher_is_better, body.weight,
            body.axis_key,
        )
        if result.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="control not found")
        return ControlOut(**body.model_dump())


@router.delete("/controls/{control_id}")
async def delete_control(control_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        await conn.execute("DELETE FROM control_values WHERE id=$1::uuid", control_id)
        res = await conn.execute("DELETE FROM controls WHERE id=$1::uuid", control_id)
        if int(res.split()[-1]) == 0:
            raise HTTPException(status_code=404, detail="control not found")
        return {"deleted": control_id}


# ---------- Trust Axis Mapping (global) ----------
@router.get("/trust-axis-mapping", response_model=List[TrustAxisMapOut])
async def list_trust_axis_mapping() -> List[TrustAxisMapOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
              p.key AS pillar_key,
              p.name AS pillar_name,
              m.axis_key,
              m.notes
            FROM pillars p
            LEFT JOIN trust_axis_pillar_map m
              ON m.pillar_key = p.key
            ORDER BY p.key
            """
        )
        return [TrustAxisMapOut(**dict(r)) for r in rows]


@router.put("/trust-axis-mapping/{pillar_key}", response_model=TrustAxisMapOut)
async def upsert_trust_axis_mapping(
    pillar_key: str,
    body: TrustAxisMapIn,
) -> TrustAxisMapOut:
    axis_key = _clean_optional_str(body.axis_key)
    notes = _clean_optional_str(body.notes)

    pool = await get_pool()
    async with pool.acquire() as conn:
        if axis_key is None:
            await conn.execute(
                "DELETE FROM trust_axis_pillar_map WHERE pillar_key=$1",
                pillar_key,
            )
        else:
            await conn.execute(
                """
                INSERT INTO trust_axis_pillar_map (pillar_key, axis_key, notes)
                VALUES ($1, $2, $3)
                ON CONFLICT (pillar_key)
                DO UPDATE SET axis_key=$2, notes=$3
                """,
                pillar_key,
                axis_key,
                notes,
            )

        row = await conn.fetchrow(
            """
            SELECT
              p.key AS pillar_key,
              p.name AS pillar_name,
              m.axis_key,
              m.notes
            FROM pillars p
            LEFT JOIN trust_axis_pillar_map m
              ON m.pillar_key = p.key
            WHERE p.key = $1
            """,
            pillar_key,
        )
        if not row:
            row = await conn.fetchrow(
                """
                SELECT
                  pillar_key,
                  NULL::text AS pillar_name,
                  axis_key,
                  notes
                FROM trust_axis_pillar_map
                WHERE pillar_key = $1
                """,
                pillar_key,
            )
        if not row:
            raise HTTPException(status_code=404, detail="Pillar not found")

        return TrustAxisMapOut(**dict(row))

# ---------- Provenance ----------
@router.get("/provenance/artifacts")
async def list_provenance_artifacts(
    project_slug: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    entity_id: UUID = Depends(get_entity_id_or_first_for_viewer),
) -> dict:
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    clauses: List[str] = []
    params: List[object] = []
    idx = 1

    if entity_id:
        clauses.append(f"entity_id=${idx}")
        params.append(entity_id)
        idx += 1
    if project_slug:
        clauses.append(f"project_slug=${idx}")
        params.append(project_slug)
        idx += 1
    if q:
        clauses.append(
            f"(name ILIKE ${idx} OR uri ILIKE ${idx} OR license_name ILIKE ${idx})"
        )
        params.append(f"%{q}%")
        idx += 1

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM provenance_artifacts {where}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT
              id,
              project_slug,
              name,
              uri,
              sha256,
              size_bytes,
              mime,
              license_name,
              license_url,
              usage_rights,
              created_at
            FROM provenance_artifacts
            {where}
            ORDER BY created_at DESC
            LIMIT ${idx} OFFSET ${idx + 1}
            """,
            *params,
            limit,
            offset,
        )
        return {"items": [dict(r) for r in rows], "total": int(total or 0)}


@router.post("/provenance/artifacts", response_model=ProvenanceArtifactOut)
async def create_provenance_artifact(
    body: ProvenanceArtifactIn,
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> ProvenanceArtifactOut:
    if not body.uri.startswith("s3://"):
        raise HTTPException(status_code=400, detail="Artifact URI must be s3://")
    
    pool = await get_pool()
    async with pool.acquire() as conn:
        effective_entity_id = await _get_entity_id_from_project_slug(conn, body.project_slug, entity_id)

    try:
        sha256_hex = compute_s3_sha256(body.uri)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Artifact read failed: {exc}")

    artifact_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        effective_entity_id = await _get_entity_id_from_project_slug(conn, body.project_slug, entity_id)
        
        row = await conn.fetchrow(
            """
            INSERT INTO provenance_artifacts (
              id, entity_id, project_slug, name, uri, sha256, size_bytes, mime,
              license_name, license_url, usage_rights, created_by
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, 'system')
            RETURNING
              id, project_slug, name, uri, sha256, size_bytes, mime,
              license_name, license_url, usage_rights, created_at
            """,
            artifact_id,
            effective_entity_id,
            body.project_slug,
            body.name,
            body.uri,
            sha256_hex.lower(),
            body.size_bytes,
            _clean_optional_str(body.mime),
            _clean_optional_str(body.license_name),
            _clean_optional_str(body.license_url),
            _clean_optional_str(body.usage_rights),
        )
        await _insert_provenance_audit(
            conn,
            "artifact",
            artifact_id,
            "created",
            "system",
            {"sha256": sha256_hex.lower(), "uri": body.uri},
        )
        return ProvenanceArtifactOut(**dict(row))


@router.post(
    "/provenance/artifacts/from-evidence/{evidence_id}",
    response_model=ProvenanceArtifactOut,
)
async def create_provenance_artifact_from_evidence(
    evidence_id: int,
) -> ProvenanceArtifactOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        ev = await conn.fetchrow(
            """
            SELECT project_slug, name, uri, sha256, size_bytes, mime
            FROM evidence
            WHERE id=$1
            """,
            evidence_id,
        )
        if not ev:
            raise HTTPException(status_code=404, detail="Evidence not found")

        uri = ev["uri"] or ""
        if not uri.startswith("s3://"):
            raise HTTPException(status_code=400, detail="Evidence must be stored in S3")

        try:
            computed = compute_s3_sha256(uri)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Artifact read failed: {exc}")

        stored = (ev["sha256"] or "").lower()
        if stored and stored != computed:
            raise HTTPException(status_code=400, detail="Evidence sha256 mismatch")

        artifact_id = str(uuid4())
        row = await conn.fetchrow(
            """
            INSERT INTO provenance_artifacts (
              id, project_slug, name, uri, sha256, size_bytes, mime,
              license_name, license_url, usage_rights, created_by
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7, NULL, NULL, NULL, 'system')
            RETURNING
              id, project_slug, name, uri, sha256, size_bytes, mime,
              license_name, license_url, usage_rights, created_at
            """,
            artifact_id,
            ev["project_slug"],
            ev["name"] or f"evidence-{evidence_id}",
            uri,
            computed.lower(),
            ev["size_bytes"],
            ev["mime"],
        )
        await _insert_provenance_audit(
            conn,
            "artifact",
            artifact_id,
            "created_from_evidence",
            "system",
            {"evidence_id": evidence_id, "sha256": computed.lower()},
        )
        return ProvenanceArtifactOut(**dict(row))


@router.post("/provenance/artifacts/{artifact_id}:validate", response_model=ProvenanceValidateOut)
async def validate_provenance_artifact(artifact_id: str) -> ProvenanceValidateOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, uri, sha256 FROM provenance_artifacts WHERE id=$1",
            artifact_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Artifact not found")

        try:
            computed = compute_s3_sha256(row["uri"])
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Artifact read failed: {exc}")

        stored = (row["sha256"] or "").lower()
        match = computed == stored
        await _insert_provenance_audit(
            conn,
            "artifact",
            artifact_id,
            "integrity_check",
            "system",
            {"computed_sha256": computed, "match": match},
        )
        return ProvenanceValidateOut(
            artifact_id=artifact_id,
            sha256=stored,
            computed_sha256=computed,
            match=match,
        )


@router.get("/entity-provider-artifacts")
async def list_entity_provider_artifacts(
    provider_key: Optional[str] = None,
    limit: int = 200,
    offset: int = 0,
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    limit = max(1, min(limit, 500))
    offset = max(0, offset)
    clauses: List[str] = ["entity_id=$1"]
    params: List[object] = [entity_id]
    idx = 2

    if provider_key:
        clauses.append(f"provider_key=${idx}")
        params.append(provider_key.strip().lower())
        idx += 1

    where = f"WHERE {' AND '.join(clauses)}"
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM entity_provider_artifacts {where}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT
              id,
              entity_id,
              provider_key,
              name,
              uri,
              sha256,
              type,
              status,
              valid_from,
              valid_to,
              updated_at
            FROM entity_provider_artifacts
            {where}
            ORDER BY updated_at DESC
            LIMIT ${idx} OFFSET ${idx + 1}
            """,
            *params,
            limit,
            offset,
        )
        return {"items": [dict(r) for r in rows], "total": int(total or 0)}


@router.post("/entity-provider-artifacts", response_model=EntityProviderArtifactOut)
async def create_entity_provider_artifact(
    body: EntityProviderArtifactIn,
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> EntityProviderArtifactOut:
    provider_key = body.provider_key.strip().lower()
    allowed = {"openai", "anthropic", "google", "meta"}
    if provider_key not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"provider_key must be one of {sorted(allowed)}",
        )
    artifact_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO entity_provider_artifacts (
              id, entity_id, provider_key, name, uri, sha256, type, status,
              valid_from, valid_to, updated_at
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, now())
            RETURNING
              id, entity_id, provider_key, name, uri, sha256, type, status,
              valid_from, valid_to, updated_at
            """,
            artifact_id,
            entity_id,
            provider_key,
            body.name.strip(),
            body.uri.strip(),
            _clean_optional_str(body.sha256),
            _clean_optional_str(body.type),
            _clean_optional_str(body.status),
            body.valid_from,
            body.valid_to,
        )
        return EntityProviderArtifactOut(**dict(row))


@router.put("/entity-provider-artifacts/{artifact_id}", response_model=EntityProviderArtifactOut)
async def update_entity_provider_artifact(
    artifact_id: str,
    body: EntityProviderArtifactIn,
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> EntityProviderArtifactOut:
    provider_key = body.provider_key.strip().lower()
    allowed = {"openai", "anthropic", "google", "meta"}
    if provider_key not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"provider_key must be one of {sorted(allowed)}",
        )
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE entity_provider_artifacts
            SET provider_key=$1,
                name=$2,
                uri=$3,
                sha256=$4,
                type=$5,
                status=$6,
                valid_from=$7,
                valid_to=$8,
                updated_at=now()
            WHERE id=$9 AND entity_id=$10
            RETURNING
              id, entity_id, provider_key, name, uri, sha256, type, status,
              valid_from, valid_to, updated_at
            """,
            provider_key,
            body.name.strip(),
            body.uri.strip(),
            _clean_optional_str(body.sha256),
            _clean_optional_str(body.type),
            _clean_optional_str(body.status),
            body.valid_from,
            body.valid_to,
            artifact_id,
            entity_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Artifact not found")
        return EntityProviderArtifactOut(**dict(row))


@router.delete("/entity-provider-artifacts/{artifact_id}")
async def delete_entity_provider_artifact(
    artifact_id: str,
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        deleted = await conn.execute(
            "DELETE FROM entity_provider_artifacts WHERE id=$1 AND entity_id=$2",
            artifact_id,
            entity_id,
        )
    return {"deleted": deleted}


@router.get("/provenance/datasets")
async def list_provenance_datasets(
    project_slug: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    entity_id: UUID = Depends(get_entity_id_or_first_for_viewer),
) -> dict:
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    clauses: List[str] = []
    params: List[object] = []
    idx = 1

    if entity_id:
        clauses.append(f"entity_id=${idx}")
        params.append(entity_id)
        idx += 1
    if project_slug:
        clauses.append(f"project_slug=${idx}")
        params.append(project_slug)
        idx += 1
    if q:
        clauses.append(f"(name ILIKE ${idx} OR description ILIKE ${idx})")
        params.append(f"%{q}%")
        idx += 1

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM provenance_datasets {where}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT id, project_slug, name, description, artifact_id, created_at
            FROM provenance_datasets
            {where}
            ORDER BY created_at DESC
            LIMIT ${idx} OFFSET ${idx + 1}
            """,
            *params,
            limit,
            offset,
        )
        return {"items": [dict(r) for r in rows], "total": int(total or 0)}


@router.post("/provenance/datasets", response_model=ProvenanceDatasetOut)
async def create_provenance_dataset(
    body: ProvenanceDatasetIn,
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> ProvenanceDatasetOut:
    dataset_id = str(uuid4())
    artifact_id = _clean_optional_str(body.artifact_id)
    pool = await get_pool()
    async with pool.acquire() as conn:
        effective_entity_id = await _get_entity_id_from_project_slug(conn, body.project_slug, entity_id)
        
        if artifact_id:
            exists = await conn.fetchval(
                "SELECT 1 FROM provenance_artifacts WHERE id=$1 AND entity_id=$2",
                artifact_id,
                effective_entity_id,
            )
            if not exists:
                raise HTTPException(status_code=404, detail="Artifact not found")

        row = await conn.fetchrow(
            """
            INSERT INTO provenance_datasets (
              id, entity_id, project_slug, name, description, artifact_id, created_by
            )
            VALUES ($1, $2, $3, $4, $5, $6, 'system')
            RETURNING id, project_slug, name, description, artifact_id, created_at
            """,
            dataset_id,
            effective_entity_id,
            body.project_slug,
            body.name,
            _clean_optional_str(body.description),
            artifact_id,
        )
        await _insert_provenance_audit(
            conn,
            "dataset",
            dataset_id,
            "created",
            "system",
            {"artifact_id": artifact_id},
        )
        return ProvenanceDatasetOut(**dict(row))


@router.put("/provenance/datasets/{dataset_id}", response_model=ProvenanceDatasetOut)
async def update_provenance_dataset(
    dataset_id: str,
    body: ProvenanceDatasetIn,
) -> ProvenanceDatasetOut:
    artifact_id = _clean_optional_str(body.artifact_id)
    pool = await get_pool()
    async with pool.acquire() as conn:
        if artifact_id:
            exists = await conn.fetchval(
                "SELECT 1 FROM provenance_artifacts WHERE id=$1",
                artifact_id,
            )
            if not exists:
                raise HTTPException(status_code=404, detail="Artifact not found")

        row = await conn.fetchrow(
            """
            UPDATE provenance_datasets
            SET project_slug=$2, name=$3, description=$4, artifact_id=$5
            WHERE id=$1
            RETURNING id, project_slug, name, description, artifact_id, created_at
            """,
            dataset_id,
            body.project_slug,
            body.name,
            _clean_optional_str(body.description),
            artifact_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Dataset not found")

        await _insert_provenance_audit(
            conn,
            "dataset",
            dataset_id,
            "updated",
            "system",
            {"artifact_id": artifact_id},
        )
        return ProvenanceDatasetOut(**dict(row))


@router.delete("/provenance/datasets/{dataset_id}")
async def delete_provenance_dataset(dataset_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM provenance_datasets WHERE id=$1",
            dataset_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Dataset not found")

        await conn.execute(
            """
            DELETE FROM provenance_lineage
            WHERE (parent_type='dataset' AND parent_id=$1)
               OR (child_type='dataset' AND child_id=$1)
            """,
            dataset_id,
        )
        await _insert_provenance_audit(
            conn,
            "dataset",
            dataset_id,
            "deleted",
            "system",
            {},
        )
        return {"ok": True, "deleted": dataset_id}


@router.get("/provenance/models")
async def list_provenance_models(
    project_slug: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
) -> dict:
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    clauses: List[str] = []
    params: List[object] = []
    idx = 1

    if project_slug:
        clauses.append(f"project_slug=${idx}")
        params.append(project_slug)
        idx += 1
    if q:
        clauses.append(
            f"(name ILIKE ${idx} OR description ILIKE ${idx} OR framework ILIKE ${idx})"
        )
        params.append(f"%{q}%")
        idx += 1

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM provenance_models {where}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT id, project_slug, name, version, framework, description, artifact_id, created_at
            FROM provenance_models
            {where}
            ORDER BY created_at DESC
            LIMIT ${idx} OFFSET ${idx + 1}
            """,
            *params,
            limit,
            offset,
        )
        return {"items": [dict(r) for r in rows], "total": int(total or 0)}


@router.post("/provenance/models", response_model=ProvenanceModelOut)
async def create_provenance_model(
    body: ProvenanceModelIn,
) -> ProvenanceModelOut:
    model_id = str(uuid4())
    artifact_id = _clean_optional_str(body.artifact_id)
    pool = await get_pool()
    async with pool.acquire() as conn:
        if artifact_id:
            exists = await conn.fetchval(
                "SELECT 1 FROM provenance_artifacts WHERE id=$1",
                artifact_id,
            )
            if not exists:
                raise HTTPException(status_code=404, detail="Artifact not found")

        row = await conn.fetchrow(
            """
            INSERT INTO provenance_models (
              id, project_slug, name, version, framework, description, artifact_id, created_by
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7, 'system')
            RETURNING id, project_slug, name, version, framework, description, artifact_id, created_at
            """,
            model_id,
            body.project_slug,
            body.name,
            _clean_optional_str(body.version),
            _clean_optional_str(body.framework),
            _clean_optional_str(body.description),
            artifact_id,
        )
        await _insert_provenance_audit(
            conn,
            "model",
            model_id,
            "created",
            "system",
            {"artifact_id": artifact_id},
        )
        return ProvenanceModelOut(**dict(row))


@router.put("/provenance/models/{model_id}", response_model=ProvenanceModelOut)
async def update_provenance_model(
    model_id: str,
    body: ProvenanceModelIn,
) -> ProvenanceModelOut:
    artifact_id = _clean_optional_str(body.artifact_id)
    pool = await get_pool()
    async with pool.acquire() as conn:
        if artifact_id:
            exists = await conn.fetchval(
                "SELECT 1 FROM provenance_artifacts WHERE id=$1",
                artifact_id,
            )
            if not exists:
                raise HTTPException(status_code=404, detail="Artifact not found")

        row = await conn.fetchrow(
            """
            UPDATE provenance_models
            SET project_slug=$2, name=$3, version=$4, framework=$5,
                description=$6, artifact_id=$7
            WHERE id=$1
            RETURNING id, project_slug, name, version, framework, description, artifact_id, created_at
            """,
            model_id,
            body.project_slug,
            body.name,
            _clean_optional_str(body.version),
            _clean_optional_str(body.framework),
            _clean_optional_str(body.description),
            artifact_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Model not found")

        await _insert_provenance_audit(
            conn,
            "model",
            model_id,
            "updated",
            "system",
            {"artifact_id": artifact_id},
        )
        return ProvenanceModelOut(**dict(row))


@router.delete("/provenance/models/{model_id}")
async def delete_provenance_model(model_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM provenance_models WHERE id=$1",
            model_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Model not found")

        await conn.execute(
            """
            DELETE FROM provenance_lineage
            WHERE (parent_type='model' AND parent_id=$1)
               OR (child_type='model' AND child_id=$1)
            """,
            model_id,
        )
        await _insert_provenance_audit(
            conn,
            "model",
            model_id,
            "deleted",
            "system",
            {},
        )
        return {"ok": True, "deleted": model_id}


@router.get("/provenance/evidence")
async def list_provenance_evidence(
    project_slug: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
) -> dict:
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    clauses: List[str] = []
    params: List[object] = []
    idx = 1

    if project_slug:
        clauses.append(f"project_slug=${idx}")
        params.append(project_slug)
        idx += 1
    if q:
        clauses.append(f"(name ILIKE ${idx} OR description ILIKE ${idx})")
        params.append(f"%{q}%")
        idx += 1

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM provenance_evidence {where}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT id, project_slug, name, description, artifact_id, created_at
            FROM provenance_evidence
            {where}
            ORDER BY created_at DESC
            LIMIT ${idx} OFFSET ${idx + 1}
            """,
            *params,
            limit,
            offset,
        )
        return {"items": [dict(r) for r in rows], "total": int(total or 0)}


@router.post("/provenance/evidence", response_model=ProvenanceEvidenceOut)
async def create_provenance_evidence(
    body: ProvenanceEvidenceIn,
) -> ProvenanceEvidenceOut:
    evidence_id = str(uuid4())
    artifact_id = _clean_optional_str(body.artifact_id)
    pool = await get_pool()
    async with pool.acquire() as conn:
        if artifact_id:
            exists = await conn.fetchval(
                "SELECT 1 FROM provenance_artifacts WHERE id=$1",
                artifact_id,
            )
            if not exists:
                raise HTTPException(status_code=404, detail="Artifact not found")

        row = await conn.fetchrow(
            """
            INSERT INTO provenance_evidence (
              id, project_slug, name, description, artifact_id, created_by
            )
            VALUES ($1, $2, $3, $4, $5, 'system')
            RETURNING id, project_slug, name, description, artifact_id, created_at
            """,
            evidence_id,
            body.project_slug,
            body.name,
            _clean_optional_str(body.description),
            artifact_id,
        )
        await _insert_provenance_audit(
            conn,
            "evidence",
            evidence_id,
            "created",
            "system",
            {"artifact_id": artifact_id},
        )
        return ProvenanceEvidenceOut(**dict(row))


@router.put("/provenance/evidence/{evidence_id}", response_model=ProvenanceEvidenceOut)
async def update_provenance_evidence(
    evidence_id: str,
    body: ProvenanceEvidenceIn,
) -> ProvenanceEvidenceOut:
    artifact_id = _clean_optional_str(body.artifact_id)
    pool = await get_pool()
    async with pool.acquire() as conn:
        if artifact_id:
            exists = await conn.fetchval(
                "SELECT 1 FROM provenance_artifacts WHERE id=$1",
                artifact_id,
            )
            if not exists:
                raise HTTPException(status_code=404, detail="Artifact not found")

        row = await conn.fetchrow(
            """
            UPDATE provenance_evidence
            SET project_slug=$2, name=$3, description=$4, artifact_id=$5
            WHERE id=$1
            RETURNING id, project_slug, name, description, artifact_id, created_at
            """,
            evidence_id,
            body.project_slug,
            body.name,
            _clean_optional_str(body.description),
            artifact_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Evidence not found")

        await _insert_provenance_audit(
            conn,
            "evidence",
            evidence_id,
            "updated",
            "system",
            {"artifact_id": artifact_id},
        )
        return ProvenanceEvidenceOut(**dict(row))


@router.delete("/provenance/evidence/{evidence_id}")
async def delete_provenance_evidence(evidence_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM provenance_evidence WHERE id=$1",
            evidence_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Evidence not found")

        await conn.execute(
            """
            DELETE FROM provenance_lineage
            WHERE (parent_type='evidence' AND parent_id=$1)
               OR (child_type='evidence' AND child_id=$1)
            """,
            evidence_id,
        )
        await _insert_provenance_audit(
            conn,
            "evidence",
            evidence_id,
            "deleted",
            "system",
            {},
        )
        return {"ok": True, "deleted": evidence_id}


@router.get("/provenance/lineage")
async def list_provenance_lineage(
    project_slug: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
) -> dict:
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    clauses: List[str] = []
    params: List[object] = []
    idx = 1

    if project_slug:
        clauses.append(f"project_slug=${idx}")
        params.append(project_slug)
        idx += 1
    if q:
        clauses.append(
            f"(relationship ILIKE ${idx} OR parent_id ILIKE ${idx} OR child_id ILIKE ${idx})"
        )
        params.append(f"%{q}%")
        idx += 1

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM provenance_lineage {where}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT id, project_slug, parent_type, parent_id, child_type, child_id, relationship, created_at
            FROM provenance_lineage
            {where}
            ORDER BY created_at DESC
            LIMIT ${idx} OFFSET ${idx + 1}
            """,
            *params,
            limit,
            offset,
        )
        return {"items": [dict(r) for r in rows], "total": int(total or 0)}


@router.post("/provenance/lineage", response_model=ProvenanceLineageOut)
async def create_provenance_lineage(
    body: ProvenanceLineageIn,
) -> ProvenanceLineageOut:
    parent_type = _validate_provenance_type(body.parent_type)
    child_type = _validate_provenance_type(body.child_type)
    parent_table = _PROVENANCE_ENTITY_TABLES[parent_type]
    child_table = _PROVENANCE_ENTITY_TABLES[child_type]

    pool = await get_pool()
    async with pool.acquire() as conn:
        parent_exists = await conn.fetchval(
            f"SELECT 1 FROM {parent_table} WHERE id=$1",
            body.parent_id,
        )
        if not parent_exists:
            raise HTTPException(status_code=404, detail="Parent not found")

        child_exists = await conn.fetchval(
            f"SELECT 1 FROM {child_table} WHERE id=$1",
            body.child_id,
        )
        if not child_exists:
            raise HTTPException(status_code=404, detail="Child not found")

        row = await conn.fetchrow(
            """
            INSERT INTO provenance_lineage (
              project_slug, parent_type, parent_id, child_type, child_id, relationship
            )
            VALUES ($1, $2, $3, $4, $5, $6)
            RETURNING id, project_slug, parent_type, parent_id, child_type, child_id, relationship, created_at
            """,
            body.project_slug,
            parent_type,
            body.parent_id,
            child_type,
            body.child_id,
            _clean_optional_str(body.relationship),
        )
        await _insert_provenance_audit(
            conn,
            parent_type,
            body.parent_id,
            "linked_to",
            "system",
            {"child_type": child_type, "child_id": body.child_id},
        )
        await _insert_provenance_audit(
            conn,
            child_type,
            body.child_id,
            "linked_from",
            "system",
            {"parent_type": parent_type, "parent_id": body.parent_id},
        )
        return ProvenanceLineageOut(**dict(row))


@router.put("/provenance/lineage/{lineage_id}", response_model=ProvenanceLineageOut)
async def update_provenance_lineage(
    lineage_id: int,
    body: ProvenanceLineageIn,
) -> ProvenanceLineageOut:
    parent_type = _validate_provenance_type(body.parent_type)
    child_type = _validate_provenance_type(body.child_type)
    parent_table = _PROVENANCE_ENTITY_TABLES[parent_type]
    child_table = _PROVENANCE_ENTITY_TABLES[child_type]

    pool = await get_pool()
    async with pool.acquire() as conn:
        parent_exists = await conn.fetchval(
            f"SELECT 1 FROM {parent_table} WHERE id=$1",
            body.parent_id,
        )
        if not parent_exists:
            raise HTTPException(status_code=404, detail="Parent not found")

        child_exists = await conn.fetchval(
            f"SELECT 1 FROM {child_table} WHERE id=$1",
            body.child_id,
        )
        if not child_exists:
            raise HTTPException(status_code=404, detail="Child not found")

        row = await conn.fetchrow(
            """
            UPDATE provenance_lineage
            SET project_slug=$2, parent_type=$3, parent_id=$4,
                child_type=$5, child_id=$6, relationship=$7
            WHERE id=$1
            RETURNING id, project_slug, parent_type, parent_id, child_type, child_id, relationship, created_at
            """,
            lineage_id,
            body.project_slug,
            parent_type,
            body.parent_id,
            child_type,
            body.child_id,
            _clean_optional_str(body.relationship),
        )
        if not row:
            raise HTTPException(status_code=404, detail="Lineage not found")

        await _insert_provenance_audit(
            conn,
            parent_type,
            body.parent_id,
            "lineage_updated",
            "system",
            {"child_type": child_type, "child_id": body.child_id},
        )
        return ProvenanceLineageOut(**dict(row))


@router.delete("/provenance/lineage/{lineage_id}")
async def delete_provenance_lineage(lineage_id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            DELETE FROM provenance_lineage
            WHERE id=$1
            RETURNING parent_type, parent_id, child_type, child_id
            """,
            lineage_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Lineage not found")

        await _insert_provenance_audit(
            conn,
            row["parent_type"],
            row["parent_id"],
            "lineage_deleted",
            "system",
            {"child_type": row["child_type"], "child_id": row["child_id"]},
        )
        return {"ok": True, "deleted": lineage_id}


@router.get("/provenance/audit", response_model=List[ProvenanceAuditOut])
async def list_provenance_audit(
    entity_type: str,
    entity_id: str,
) -> List[ProvenanceAuditOut]:
    _validate_provenance_type(entity_type)
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, entity_type, entity_id, action, actor, at, details_json
            FROM provenance_audit
            WHERE entity_type=$1 AND entity_id=$2
            ORDER BY at DESC
            """,
            entity_type,
            entity_id,
        )
        return [ProvenanceAuditOut(**dict(r)) for r in rows]

# ---------- kpis handler ----------
@router.get("/kpis")
async def list_kpis():
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
              k.id,
              k.key,
              k.name,
              pl.key  AS pillar_key,
              pl.name AS pillar_name,
              k.unit,
              k.weight,
              k.min_ideal,
              k.max_ideal,
              k.invert,
              k.description,
              k.example
            FROM public.kpis k
            JOIN public.pillars pl ON pl.id = k.pillar_id
            ORDER BY pillar_name ASC NULLS LAST, k.key
            """
        )
        return [dict(r) for r in rows]


@router.get("/kpis.xlsx")
async def export_kpis():
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
              k.key  AS kpi_key,
              k.name AS kpi_name,
              pl.name AS pillar_name,
              k.unit,
              k.description,
              k.example,
              k.weight,
              k.min_ideal,
              k.max_ideal,
              k.invert
            FROM public.kpis k
            JOIN public.pillars pl ON pl.id = k.pillar_id
            ORDER BY pillar_name ASC NULLS LAST, k.key
            """
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "kpis"
        ws.append(
            [
                "pillar",
                "kpi_key",
                "name",
                "unit",
                "description",
                "example",
                "weight",
                "min_ideal",
                "max_ideal",
                "invert",
            ]
        )
        if rows:
            for r in rows:
                ws.append(
                    [
                        r["pillar_name"],
                        r["kpi_key"],
                        r["kpi_name"],
                        r["unit"],
                        r["description"],
                        r["example"],
                        r["weight"],
                        r["min_ideal"],
                        r["max_ideal"],
                        r["invert"],
                    ]
                )

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="kpis.xlsx"'},
    )


# ---------- Projects (create/upsert) ----------
@router.post("/projects", response_model=ProjectOut)
async def create_project(
    body: ProjectIn,
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> ProjectOut:
    """
    Create a new project. Requires editor role or higher.
    Entity ID is validated by authorization.
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        entity_slug = await conn.fetchval(
            "SELECT slug FROM entity WHERE id = $1",
            entity_id,
        )
        if not entity_slug:
            raise HTTPException(status_code=400, detail="Entity slug is missing")
        existing = await conn.fetchrow(
            """
            SELECT id, risk_level, priority, sponsor, owner, creation_date, status, entity_id, is_archived
            FROM entity_projects
            WHERE slug=$1 AND entity_id=$2
            """,
            body.slug,
            entity_id,
        )

        base_risk = _clean_optional_str(body.risk_level) or (existing["risk_level"] if existing and existing["risk_level"] else "medium")
        priority = _clean_optional_str(body.priority) or base_risk
        sponsor = _clean_optional_str(body.sponsor)
        owner = _clean_optional_str(body.owner)
        status_value = _clean_optional_str(body.status) or (
            existing["status"] if existing and existing["status"] else "Planned"
        )

        if existing:
            if sponsor is None and existing["sponsor"] is not None and body.sponsor is None:
                sponsor = existing["sponsor"]
            if owner is None and existing["owner"] is not None and body.owner is None:
                owner = existing["owner"]
            creation_date = (
                body.creation_date if body.creation_date is not None else existing["creation_date"]
            )
        else:
            creation_date = body.creation_date or date.today()

        update_ts = body.update_date or datetime.now(timezone.utc)

        if existing:
            await conn.execute(
                """
                UPDATE entity_projects
                SET name=$2,
                    risk_level=$3,
                    target_threshold=$4,
                    priority=$5,
                    sponsor=$6,
                    owner=$7,
                    status=$8,
                    creation_date=$9,
                    update_date=$10,
                    entity_slug=$11,
                    is_archived=FALSE,
                    archived_at=NULL
                WHERE slug=$1 AND entity_id=$12
                """,
                body.slug,
                body.name,
                base_risk,
                body.target_threshold,
                priority,
                sponsor,
                owner,
                status_value,
                creation_date,
                update_ts,
                entity_slug,
                entity_id,
            )
            row = await conn.fetchrow(
                """
                SELECT id, slug, name, risk_level, target_threshold, priority, sponsor, owner, status, creation_date, update_date
                FROM entity_projects
                WHERE slug=$1 AND entity_id=$2
                """,
                body.slug,
                entity_id,
            )
            return ProjectOut(**dict(row))

        try:
            row = await conn.fetchrow(
                """
                INSERT INTO entity_projects (
                    id,
                    entity_id,
                    entity_slug,
                    slug,
                    name,
                    risk_level,
                    target_threshold,
                    priority,
                    sponsor,
                    owner,
                    status,
                    creation_date,
                    update_date
                )
                VALUES (gen_random_uuid()::text, $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
                RETURNING id, slug, name, risk_level, target_threshold, priority, sponsor, owner, status, creation_date, update_date
                """,
                entity_id,
                entity_slug,
                body.slug,
                body.name,
                base_risk,
                body.target_threshold,
                priority,
                sponsor,
                owner,
                status_value,
                creation_date,
                update_ts,
            )
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to create project: {e}")
        
        # Create alert for new project in AI Requirements Register
        await _create_new_project_requirements_alert(conn, entity_slug, body.slug, body.name)
        
        return ProjectOut(**dict(row))



# ---------- Evidence helper for UI ----------
@router.get("/projects/{project_slug}/kpis/{kpi_key}/control-id")
async def get_control_id_for_kpi(
    project_slug: str,
    kpi_key: str,
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    """
    Used by Evidence() button:
    returns the control_id (uuid text) for a given {project_slug, kpi_key}.
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Verify project exists and get entity_id
        effective_entity_id = await _get_entity_id_from_project_slug(conn, project_slug, entity_id)
        
        # Resolve once to fail fast if the project doesn't exist
        await get_project_id_by_slug(conn, project_slug)

        row = await conn.fetchrow(
            """
            SELECT v.control_id::text AS control_id
            FROM control_values v
            LEFT JOIN controls c ON c.id = v.control_id
            WHERE v.project_slug = $1
              AND (v.kpi_key = $2 OR c.kpi_key = $2)
            LIMIT 1
            """,
            project_slug, kpi_key,
        )
        if not row or not row["control_id"]:
            raise HTTPException(status_code=404, detail="Control not found for KPI")
        return {"control_id": row["control_id"]}


# ---------- Evidence endpoints (using current schema) ----------
class EvidenceUpdate(BaseModel):
    evidence_source: Optional[str] = None
    owner_role: Optional[str] = None
    notes: Optional[str] = None


class EvidenceFinalize(BaseModel):
    evidence_id: int            # adjust to str if your id is UUID
    sha256_hex: Optional[str] = None
    updated_by: Optional[str] = None  # person who completed the upload (for audit)


class EvidenceStatusUpdate(BaseModel):
    status: Optional[str] = None
    action: Optional[str] = None
    comment: Optional[str] = None
    updated_by: Optional[str] = None
    approval_status: Optional[str] = None
    approved_by: Optional[str] = None


class AiSystemIn(BaseModel):
    project_slug: Optional[str] = None
    name: str
    uc_id: Optional[str] = None
    description: Optional[str] = None
    owner: Optional[str] = None
    system_owner_email: Optional[str] = None
    business_unit: Optional[str] = None
    risk_owner_role: Optional[str] = None
    decision_authority: Optional[str] = None
    model_provider: Optional[str] = None
    provider_type: Optional[str] = None
    intended_use: Optional[str] = None
    intended_users: Optional[str] = None
    system_boundary: Optional[str] = None
    model_type: Optional[str] = None
    model_version: Optional[str] = None
    deployment_environment: Optional[str] = None
    lifecycle_stage: Optional[str] = None
    training_data_sources: Optional[str] = None
    personal_data_flag: Optional[bool] = None
    sensitive_attributes_flag: Optional[bool] = None
    risk_tier: Optional[str] = None
    status: Optional[str] = None
    region_scope: Optional[str] = None
    data_sensitivity: Optional[str] = None
    # New governance fields
    model_name: Optional[str] = None
    technical_lead: Optional[str] = None
    target_users: Optional[str] = None
    intended_purpose: Optional[str] = None
    out_of_scope_uses: Optional[str] = None
    deployment_method: Optional[str] = None
    data_residency: Optional[str] = None
    base_model_type: Optional[str] = None
    input_output_modality: Optional[str] = None
    fine_tuning_data: Optional[str] = None
    data_minimization: Optional[str] = None
    human_oversight_mechanism: Optional[str] = None
    impact_assessment_reference: Optional[str] = None
    known_limitations: Optional[str] = None
    langfuse_project_id: Optional[str] = None
    langfuse_base_url: Optional[str] = None


class AiSystemUpdate(BaseModel):
    project_slug: Optional[str] = None
    name: Optional[str] = None
    description: Optional[str] = None
    owner: Optional[str] = None
    system_owner_email: Optional[str] = None
    business_unit: Optional[str] = None
    risk_owner_role: Optional[str] = None
    decision_authority: Optional[str] = None
    model_provider: Optional[str] = None
    provider_type: Optional[str] = None
    intended_use: Optional[str] = None
    intended_users: Optional[str] = None
    system_boundary: Optional[str] = None
    model_type: Optional[str] = None
    model_version: Optional[str] = None
    deployment_environment: Optional[str] = None
    lifecycle_stage: Optional[str] = None
    training_data_sources: Optional[str] = None
    personal_data_flag: Optional[bool] = None
    sensitive_attributes_flag: Optional[bool] = None
    risk_tier: Optional[str] = None
    status: Optional[str] = None
    region_scope: Optional[str] = None
    data_sensitivity: Optional[str] = None
    model_name: Optional[str] = None
    technical_lead: Optional[str] = None
    target_users: Optional[str] = None
    intended_purpose: Optional[str] = None
    out_of_scope_uses: Optional[str] = None
    deployment_method: Optional[str] = None
    data_residency: Optional[str] = None
    base_model_type: Optional[str] = None
    input_output_modality: Optional[str] = None
    fine_tuning_data: Optional[str] = None
    data_minimization: Optional[str] = None
    human_oversight_mechanism: Optional[str] = None
    impact_assessment_reference: Optional[str] = None
    known_limitations: Optional[str] = None
    langfuse_project_id: Optional[str] = None
    langfuse_base_url: Optional[str] = None


class ModelCardUpsert(BaseModel):
    summary_md: Optional[str] = None
    limitations: Optional[str] = None
    out_of_scope: Optional[str] = None
    review_cadence: Optional[str] = None
    status: Optional[str] = None
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None


class LangfuseProjectLookup(BaseModel):
    public_key: str
    secret_key: str
    base_url: Optional[str] = None


class RequirementIn(BaseModel):
    project_slug: Optional[str] = None
    uc_id: Optional[str] = None
    framework: str
    requirement_code: str
    title: Optional[str] = None
    description: Optional[str] = None
    applicability: Optional[str] = None
    owner_role: Optional[str] = None
    status: Optional[str] = None
    evidence_ids: Optional[list] = None
    mapped_controls: Optional[list] = None
    notes: Optional[str] = None


class RequirementUpdate(BaseModel):
    project_slug: Optional[str] = None
    uc_id: Optional[str] = None
    framework: Optional[str] = None
    requirement_code: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    applicability: Optional[str] = None
    owner_role: Optional[str] = None
    status: Optional[str] = None
    evidence_ids: Optional[list] = None
    mapped_controls: Optional[list] = None
    notes: Optional[str] = None


class AimsScopeIn(BaseModel):
    scope_name: Optional[str] = None
    scope_statement: Optional[str] = None
    context_internal: Optional[str] = None
    context_external: Optional[str] = None
    interested_parties: Optional[list] = None
    scope_boundaries: Optional[str] = None
    lifecycle_coverage: Optional[list] = None
    cloud_platforms: Optional[list] = None
    regulatory_requirements: Optional[list] = None
    isms_pms_integration: Optional[str] = None
    exclusions: Optional[str] = None
    owner: Optional[str] = None
    status: Optional[str] = None
    updated_by: Optional[str] = None


class PolicyIn(BaseModel):
    title: str
    owner_role: Optional[str] = None
    status: Optional[str] = None
    iso42001_requirement: Optional[str] = None
    iso42001_status: Optional[str] = None
    euaiact_requirements: Optional[str] = None
    nistairmf_requirements: Optional[str] = None
    comment: Optional[str] = None
    action: Optional[str] = None
    template: Optional[str] = None
    version_label: Optional[str] = None
    content: Optional[str] = None


class PolicyUpdate(BaseModel):
    title: Optional[str] = None
    owner_role: Optional[str] = None
    status: Optional[str] = None
    iso42001_requirement: Optional[str] = None
    iso42001_status: Optional[str] = None
    euaiact_requirements: Optional[str] = None
    nistairmf_requirements: Optional[str] = None
    comment: Optional[str] = None
    action: Optional[str] = None
    template: Optional[str] = None


class PolicyControlMapUpdate(BaseModel):
    policy_id: str
    project_slug: str
    control_ids: List[str] = Field(default_factory=list)


class PolicyVersionIn(BaseModel):
    version_label: str
    content: Optional[str] = None


class PolicyVersionUpdate(BaseModel):
    status: Optional[str] = None
    approved_by: Optional[str] = None
    content: Optional[str] = None


class DataSourceConnectorIn(BaseModel):
    name: str = Field(min_length=1)
    type: str = Field(default="postgres", min_length=1)
    host: str = Field(min_length=1)
    port: int = 5432
    database: str = Field(min_length=1)
    username: str = Field(min_length=1)
    password: Optional[str] = None
    ssl_mode: Optional[str] = None
    status: Optional[str] = None


class DataSourceConnectorUpdate(BaseModel):
    name: Optional[str] = None
    host: Optional[str] = None
    port: Optional[int] = None
    database: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    ssl_mode: Optional[str] = None
    status: Optional[str] = None


class DataSourceConnectorOut(BaseModel):
    id: str
    name: str
    type: str
    host: str
    port: int
    database: str
    username: str
    ssl_mode: Optional[str] = None
    status: str
    last_tested_at: Optional[datetime] = None
    last_test_status: Optional[str] = None
    last_test_error: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None


class DataClassificationTagIn(BaseModel):
    tag_name: str = Field(min_length=1)
    sensitivity_tier: Optional[str] = None
    pii_flag: bool = False
    allowed_usage: Optional[str] = None
    retention_class: Optional[str] = None


class DataClassificationTagUpdate(BaseModel):
    tag_name: Optional[str] = None
    sensitivity_tier: Optional[str] = None
    pii_flag: Optional[bool] = None
    allowed_usage: Optional[str] = None
    retention_class: Optional[str] = None


class DataClassificationTagOut(DataClassificationTagIn):
    id: str
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None


class DataClassificationAssignmentIn(BaseModel):
    connector_id: str = Field(min_length=1)
    schema_name: Optional[str] = None
    table_name: Optional[str] = None
    name: Optional[str] = None
    id_number: Optional[str] = None
    tag_id: str = Field(min_length=1)


class DataClassificationAssignmentUpdate(BaseModel):
    connector_id: Optional[str] = None
    schema_name: Optional[str] = None
    table_name: Optional[str] = None
    name: Optional[str] = None
    id_number: Optional[str] = None
    tag_id: Optional[str] = None


class DataClassificationAssignmentOut(BaseModel):
    id: str
    connector_id: str
    schema_name: Optional[str] = None
    table_name: Optional[str] = None
    name: Optional[str] = None
    id_number: Optional[str] = None
    tag_id: str
    tag_name: Optional[str] = None
    sensitivity_tier: Optional[str] = None
    pii_flag: Optional[bool] = None
    allowed_usage: Optional[str] = None
    retention_class: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None


class DataRetentionPolicyIn(BaseModel):
    retention_class: str = Field(min_length=1)
    archive_after_days: Optional[int] = None
    delete_after_days: int = Field(ge=1)
    notes: Optional[str] = None


class DataRetentionPolicyUpdate(BaseModel):
    retention_class: Optional[str] = None
    archive_after_days: Optional[int] = None
    delete_after_days: Optional[int] = None
    notes: Optional[str] = None


class DataRetentionPolicyOut(DataRetentionPolicyIn):
    id: str
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None


class DataRetentionRecordIn(BaseModel):
    assignment_id: str = Field(min_length=1)
    retention_class: str = Field(min_length=1)
    start_date: date
    last_accessed_at: Optional[datetime] = None
    status: Optional[str] = None
    notes: Optional[str] = None


class DataRetentionRecordUpdate(BaseModel):
    assignment_id: Optional[str] = None
    retention_class: Optional[str] = None
    start_date: Optional[date] = None
    last_accessed_at: Optional[datetime] = None
    status: Optional[str] = None
    notes: Optional[str] = None


class DataRetentionRecordOut(BaseModel):
    id: str
    assignment_id: str
    retention_class: str
    start_date: date
    last_accessed_at: Optional[datetime] = None
    status: str
    archived_at: Optional[datetime] = None
    deleted_at: Optional[datetime] = None
    notes: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None


class DataUsageRecordIn(BaseModel):
    assignment_id: str = Field(min_length=1)
    usage_type: str = Field(min_length=1)
    purpose: Optional[str] = None


class DataUsageRecordOut(BaseModel):
    id: str
    assignment_id: str
    usage_type: str
    purpose: Optional[str] = None
    recorded_at: Optional[datetime] = None


class DataGovernanceWarningOut(BaseModel):
    id: str
    assignment_id: str
    warning_type: str
    severity: str
    message: str
    created_at: Optional[datetime] = None
    resolved_at: Optional[datetime] = None


class PolicyAlertOut(BaseModel):
    id: str
    policy_id: str
    policy_title: str
    project_slug: Optional[str] = None
    alert_type: str
    severity: str
    message: str
    source_type: Optional[str] = None
    source_key: Optional[str] = None
    status: str
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    resolved_at: Optional[datetime] = None
    details_json: Optional[dict] = None


@router.get("/projects/{project_slug}/controls/{control_id}/evidence")
async def get_control_evidence(
    project_slug: str,
    control_id: str,
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    """
    Evidence() drawer fetch.
    Returns current metadata for this control in this project.
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        effective_entity_id = await _get_entity_id_from_project_slug(conn, project_slug, entity_id)
        
        row = await conn.fetchrow(
            """
            SELECT v.project_slug,
                   v.control_id::text              AS control_id,
                   COALESCE(c.kpi_key, v.kpi_key)  AS kpi_key,
                   v.evidence_source,
                   v.owner_role,
                   v.notes,
                   v.updated_at
            FROM control_values v
            LEFT JOIN controls c ON c.id = v.control_id
            WHERE v.project_slug = $1 AND v.entity_id = $3 AND (v.control_id::text = $2 OR v.control_id = $2::uuid)
            """,
            project_slug, control_id, effective_entity_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Evidence not found for control")

        files = await conn.fetch(
            """
            SELECT id, name, mime, size_bytes, sha256, uri, status, created_at, updated_at
            FROM evidence
            WHERE project_slug=$1 AND entity_id=$3 AND (control_id::text=$2 OR control_id=$2::uuid)
            ORDER BY updated_at DESC NULLS LAST, id DESC
            """,
            project_slug, control_id, effective_entity_id
        )

        attachments = []
        for f in files:
            attachments.append({
                "id": int(f["id"]),
                "filename": f["name"],        # alias for UI
                "name": f["name"],
                "content_type": f["mime"],
                "size_bytes": f["size_bytes"],
                "sha256_hex": f["sha256"],
                "storage_url": f["uri"],
                "status": f["status"],
                "created_at": f["created_at"].isoformat() if f["created_at"] else None,
                "updated_at": f["updated_at"].isoformat() if f["updated_at"] else None,
            })

        return {
            "project_slug": row["project_slug"],
            "control_id": row["control_id"],
            "kpi_key": row["kpi_key"],
            "evidence_source": row["evidence_source"],
            "owner_role": row["owner_role"],
            "notes": row["notes"],
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
            "attachments": attachments,
        }


@router.get("/evidences")
async def list_all_evidences(
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        query = """
            SELECT
              e.id,
              e.project_slug,
              e.control_id::text AS control_id,
              e.name,
              e.mime,
              e.size_bytes,
              e.sha256,
              e.uri,
              e.status,
              e.approval_status,
              e.approved_by,
              e.approved_at,
              e.created_at,
              e.updated_at,
              e.last_comment,
              e.last_action,
              e.attachment_name,
              e.attachment_uri,
              e.attachment_mime,
              e.attachment_size,
              e.attachment_sha256,
              e.updated_by,
              e.last_update,
              COALESCE(v.evidence_source, c.evidence_source) AS evidence_source,
              COALESCE(v.owner_role, c.owner_role) AS owner_role
            FROM evidence e
            LEFT JOIN control_values v
              ON v.project_slug = e.project_slug
             AND v.control_id = e.control_id
             AND v.entity_id = e.entity_id
            LEFT JOIN controls c
              ON c.id = e.control_id
        """
        if entity_id:
            query += " WHERE e.entity_id = $1"
            params = [entity_id]
        else:
            params = []
        query += " ORDER BY e.created_at DESC NULLS LAST, e.id DESC"
        
        rows = await conn.fetch(query, *params)

    items = []
    for r in rows:
        uri = r["uri"] or ""
        download_url = ""
        if uri.startswith("s3://") or uri.startswith("file://"):
            download_url = f"/admin/evidence/{r['id']}:download"
        elif uri:
            download_url = uri
        attachment_uri = r["attachment_uri"] or ""
        attachment_download_url = ""
        if attachment_uri.startswith("s3://") or attachment_uri.startswith("file://"):
            attachment_download_url = f"/admin/evidences/{r['id']}/attachment:download"
        elif attachment_uri:
            attachment_download_url = attachment_uri

        last_update = r["last_update"] or r["updated_at"]
        items.append(
            {
                "id": int(r["id"]),
                "project_slug": r["project_slug"],
                "control_id": r["control_id"],
                "name": r["name"],
                "mime": r["mime"],
                "size_bytes": r["size_bytes"],
                "sha256": r["sha256"],
                "uri": uri,
                "status": r["status"],
                "approval_status": r["approval_status"],
                "approved_by": r["approved_by"],
                "approved_at": r["approved_at"].isoformat() if r["approved_at"] else None,
                "evidence_source": r["evidence_source"],
                "owner_role": r["owner_role"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
                "last_comment": r["last_comment"],
                "last_action": r["last_action"],
                "attachment_name": r["attachment_name"],
                "attachment_uri": attachment_uri or None,
                "attachment_mime": r["attachment_mime"],
                "attachment_size": r["attachment_size"],
                "attachment_sha256": r["attachment_sha256"],
                "updated_by": r["updated_by"],
                "last_update": last_update.isoformat() if last_update else None,
                "download_url": download_url,
                "attachment_download_url": attachment_download_url or None,
            }
        )

    return {"items": items}


@router.post("/projects/{project_slug}/controls/{control_id}/evidence:init")
async def init_control_evidence_upload(
    request: Request,
    project_slug: str,
    control_id: str,
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    """
    Import Evidence button preflight/init.

    Creates an 'evidence' row with status 'pending' and returns a PUT URL
    that the UI can upload to (this API).
    """
    payload = {}
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    filename = (
        payload.get("filename")
        or payload.get("name")
        or "upload.bin"
    )
    filename = os.path.basename(str(filename)) or "upload.bin"
    content_type = (
        payload.get("contentType")
        or payload.get("mime")
        or "application/octet-stream"
    )
    size_bytes = payload.get("sizeBytes") or payload.get("size_bytes") or 0
    try:
        size_bytes = int(size_bytes)
    except Exception:
        size_bytes = 0

    # Person who is uploading (from UI session); store in evidence and audit
    created_by = (
        payload.get("createdBy")
        or payload.get("created_by")
        or payload.get("user_email")
        or payload.get("user_name")
    )
    if isinstance(created_by, str):
        created_by = created_by.strip() or "system"
    else:
        created_by = "system"

    object_key = f"evidence/{project_slug}/{control_id}/{uuid4()}-{filename}"
    uri = object_uri(object_key)
    upload_url, upload_headers = presign_put(object_key, content_type)

    pool = await get_pool()
    async with pool.acquire() as conn:
        effective_entity_id = await _get_entity_id_from_project_slug(
            conn, project_slug, entity_id
        )
        # Validate control exists for the project
        exists = await conn.fetchval(
            """
            SELECT 1
            FROM control_values
            WHERE project_slug=$1
              AND entity_id=$3
              AND (control_id::text=$2 OR control_id=$2::uuid)
            """,
            project_slug, control_id, effective_entity_id
        )
        if not exists:
            raise HTTPException(status_code=404, detail="Control not found")

        row = await conn.fetchrow(
            """
            INSERT INTO evidence (
              project_slug, control_id, name, mime,
              size_bytes, sha256, uri, status, created_by, entity_id,
              created_at, updated_at
            )
            VALUES ($1, $2::uuid, $3, $4, $5, NULL, $6, 'pending', $7, $8, NOW(), NOW())
            RETURNING id, uri, status
            """,
            project_slug, control_id, filename, content_type, size_bytes, uri, created_by, effective_entity_id
        )
        evidence_id = row["id"]

    await _log_audit(
        event_type="evidence_init",
        actor=created_by,
        source_service="core-svc",
        object_type="evidence",
        object_id=str(evidence_id),
        project_slug=project_slug,
        details={
            "control_id": control_id,
            "filename": filename,
            "created_by": created_by,
        },
    )

    return {
        "ok": True,
        "mode": "put",
        # new key expected by frontend:
        "upload_url": upload_url,
        # legacy alias for completeness:
        "put_url": upload_url,
        "headers": upload_headers,
        "max_size_mb": 100,
        "accepted": ["application/pdf", "text/csv", "image/*", "text/plain"],
        "evidence_id": evidence_id,
        # expose storage fields for UI (aliases too)
        "storage_url": uri,
        "uri": uri,
        "status": row["status"] or "pending",
        "method": "PUT",
    }


@router.put("/projects/{project_slug}/controls/{control_id}/evidence:upload/{evidence_id}")
@router.put("/projects/{project_slug}/controls/{control_id}/evidence/upload/{evidence_id}")
async def upload_control_evidence_binary(
    request: Request,
    project_slug: str,
    control_id: str,
    evidence_id: int,  # change to str if UUID
):
    """
    Step 2: UI does a PUT to this URL with the raw file body.
    We stream the body to S3 (preferred) and update the DB row with size/mime,
    sha256 and an s3:// URI. Falls back to local disk if S3 upload fails.
    """
    original_filename = request.headers.get("X-Original-Filename", "upload.bin")
    content_type = request.headers.get("Content-Type", "application/octet-stream")

    pool = await get_pool()
    storage_uri = None
    sha_hex = None
    size = 0

    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT uri
            FROM evidence
            WHERE id=$1 AND project_slug=$2 AND (control_id::text=$3 OR control_id=$3::uuid)
            """,
            evidence_id,
            project_slug,
            control_id,
        )
        if row:
            storage_uri = row["uri"]

    if storage_uri and storage_uri.startswith("s3://") and s3_ready():
        key = _s3_key_from_uri(storage_uri)
        if not key:
            raise HTTPException(status_code=400, detail="Invalid evidence URI")
        try:
            size, sha_hex = await upload_stream_to_s3(
                key, request.stream(), content_type
            )
        except Exception:
            raise HTTPException(status_code=500, detail="S3 upload failed")

    if not storage_uri or not storage_uri.startswith("s3://"):
        bucket_dir = LOCAL_EVIDENCE_ROOT / project_slug / control_id
        bucket_dir.mkdir(parents=True, exist_ok=True)
        target_path = bucket_dir / f"{evidence_id}-{original_filename}"

        hasher = hashlib.sha256()
        size = 0
        with target_path.open("wb") as f:
            async for chunk in request.stream():
                if not chunk:
                    continue
                f.write(chunk)
                hasher.update(chunk)
                size += len(chunk)
        sha_hex = hasher.hexdigest()
        storage_uri = f"file://{target_path.as_posix()}"

    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE evidence
            SET name         = $4,
                mime         = $5,
                size_bytes   = $6,
                sha256       = $7,
                uri          = $8,
                status       = 'uploaded',
                entity_id    = COALESCE(
                                 entity_id,
                                 (SELECT entity_id
                                  FROM control_values
                                  WHERE project_slug = $1
                                    AND (control_id::text=$2 OR control_id=$2::uuid)
                                  LIMIT 1)
                               ),
                updated_at   = NOW()
            WHERE id = $3 AND project_slug = $1 AND (control_id::text=$2 OR control_id=$2::uuid)
            """,
            project_slug, control_id, evidence_id,
            original_filename, content_type, size, sha_hex, storage_uri
        )
        if res.split()[-1] == "0":
            if storage_uri and storage_uri.startswith("file://"):
                try:
                    target_path.unlink(missing_ok=True)
                except Exception:
                    pass
            raise HTTPException(status_code=404, detail="Evidence row not found")

    await _log_audit(
        event_type="evidence_uploaded",
        actor="system",
        source_service="core-svc",
        object_type="evidence",
        object_id=str(evidence_id),
        project_slug=project_slug,
        details={"control_id": control_id, "filename": original_filename},
    )
    return Response(status_code=200)


@router.post("/projects/{project_slug}/controls/{control_id}/evidence:finalize")
async def finalize_control_evidence_upload(
    project_slug: str,
    control_id: str,
    body: EvidenceFinalize
) -> dict:
    """
    Step 3: UI sends sha256_hex (optional) after the PUT finishes.
    We persist the hash (if provided) and set status to 'ready'.
    Optionally accepts updated_by for audit and evidence.updated_by.
    """
    updated_by = (body.updated_by or "").strip() or None
    pool = await get_pool()
    async with pool.acquire() as conn:
        if updated_by:
            res = await conn.execute(
                """
                UPDATE evidence
                SET sha256     = COALESCE($4, sha256),
                    status     = 'ready',
                    updated_at = NOW(),
                    updated_by = $5,
                    entity_id  = COALESCE(
                                 entity_id,
                                 (SELECT entity_id
                                  FROM control_values
                                  WHERE project_slug = $1
                                    AND (control_id::text=$2 OR control_id=$2::uuid)
                                  LIMIT 1)
                               )
                WHERE id = $3 AND project_slug = $1 AND (control_id::text=$2 OR control_id=$2::uuid)
                """,
                project_slug, control_id, body.evidence_id, body.sha256_hex, updated_by
            )
        else:
            res = await conn.execute(
                """
                UPDATE evidence
                SET sha256     = COALESCE($4, sha256),
                    status     = 'ready',
                    updated_at = NOW(),
                    entity_id  = COALESCE(
                                 entity_id,
                                 (SELECT entity_id
                                  FROM control_values
                                  WHERE project_slug = $1
                                    AND (control_id::text=$2 OR control_id=$2::uuid)
                                  LIMIT 1)
                               )
                WHERE id = $3 AND project_slug = $1 AND (control_id::text=$2 OR control_id=$2::uuid)
                """,
                project_slug, control_id, body.evidence_id, body.sha256_hex
            )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Evidence row not found")
        try:
            from app.services.control_reminders import advance_cycle_for_control
            await advance_cycle_for_control(conn, project_slug, control_id)
        except Exception:
            pass
    await _log_audit(
        event_type="evidence_finalized",
        actor=updated_by or "system",
        source_service="core-svc",
        object_type="evidence",
        object_id=str(body.evidence_id),
        project_slug=project_slug,
        details={
            "control_id": control_id,
            "updated_by": updated_by,
        },
    )
    return {"ok": True, "evidence_id": body.evidence_id}


# ---- Finalize shims to match frontend path style (/.../evidence:finalize/{id}) ----
@router.post("/projects/{project_slug}/controls/{control_id}/evidence:finalize/{evidence_id}")
@router.post("/projects/{project_slug}/controls/{control_id}/evidence/finalize/{evidence_id}")
async def finalize_control_evidence_upload_with_path_id(
    project_slug: str,
    control_id: str,
    evidence_id: int,
    body: EvidenceFinalize
) -> dict:
    body.evidence_id = evidence_id
    return await finalize_control_evidence_upload(project_slug, control_id, body)


# ---- Download URL helper expected by frontend (/admin/evidence/{id}:download-url) ----
@router.post("/evidence/{evidence_id}:download-url")
async def evidence_download_url(evidence_id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT uri FROM evidence WHERE id=$1", evidence_id)
        if not row:
            raise HTTPException(status_code=404, detail="Evidence not found")
        uri = row["uri"] or ""
        if uri.startswith("file://") or uri.startswith("s3://"):
            return {"url": f"/admin/evidence/{evidence_id}:download"}
        return {"url": uri}


@router.delete("/evidences/{evidence_id}")
async def delete_evidence_by_id(evidence_id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT uri, attachment_uri FROM evidence WHERE id=$1", evidence_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Evidence not found")
        uri = row["uri"] or ""
        attachment_uri = row["attachment_uri"] or ""

        # Best-effort delete from storage
        if uri.startswith("s3://"):
            key = _s3_key_from_uri(uri)
            if key:
                try:
                    delete_object(key)
                except Exception:
                    pass
        elif uri.startswith("file://"):
            local_path = _local_path_from_uri(uri)
            if local_path and local_path.exists():
                try:
                    local_path.unlink()
                except Exception:
                    pass

        # Best-effort delete attachment
        if attachment_uri.startswith("s3://"):
            key = _s3_key_from_uri(attachment_uri)
            if key:
                try:
                    delete_object(key)
                except Exception:
                    pass
        elif attachment_uri.startswith("file://"):
            local_path = _local_path_from_uri(attachment_uri)
            if local_path and local_path.exists():
                try:
                    local_path.unlink()
                except Exception:
                    pass

        # Remove DB rows (audit first to avoid FK issues)
        await conn.execute("DELETE FROM evidence_audit WHERE evidence_id=$1", evidence_id)
        await conn.execute("DELETE FROM evidence WHERE id=$1", evidence_id)

    await _log_audit(
        event_type="evidence_deleted",
        actor="system",
        source_service="core-svc",
        object_type="evidence",
        object_id=str(evidence_id),
    )
    return {"ok": True, "deleted": evidence_id}


@router.patch("/evidences/{evidence_id}")
async def update_evidence_status(evidence_id: int, body: EvidenceStatusUpdate) -> dict:
    if not any(
        [
            body.status,
            body.action,
            body.comment,
            body.updated_by,
            body.approval_status,
            body.approved_by,
        ]
    ):
        raise HTTPException(status_code=400, detail="Nothing to update")
    pool = await get_pool()
    async with pool.acquire() as conn:
        approval_status = body.approval_status
        approved_by = body.approved_by or body.updated_by
        res = await conn.execute(
            """
            UPDATE evidence
            SET status = COALESCE($1, status),
                last_action = COALESCE($2, last_action),
                last_comment = COALESCE($3, last_comment),
                updated_by = COALESCE($4, updated_by),
                approval_status = COALESCE($5, approval_status),
                approved_by = CASE
                    WHEN $5 IS NULL THEN approved_by
                    WHEN $5 = 'approved' THEN COALESCE($6, approved_by, $4)
                    ELSE NULL
                END,
                approved_at = CASE
                    WHEN $5 IS NULL THEN approved_at
                    WHEN $5 = 'approved' THEN NOW()
                    ELSE NULL
                END,
                last_update = NOW(),
                updated_at = NOW()
            WHERE id = $7
            """,
            body.status,
            body.action,
            body.comment,
            body.updated_by,
            approval_status,
            approved_by,
            evidence_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Evidence not found")
        action_label = body.action or body.status or "updated"
        details_json = None
        if body.comment or approval_status or approved_by:
            details_json = json.dumps(
                {
                    "comment": body.comment,
                    "approval_status": approval_status,
                    "approved_by": approved_by,
                }
            )
        await conn.execute(
            """
            INSERT INTO evidence_audit (evidence_id, action, actor, details_json)
            VALUES ($1, $2, $3, $4::jsonb)
            """,
            evidence_id,
            action_label,
            body.updated_by or "system",
            details_json,
        )
    return {
        "ok": True,
        "evidence_id": evidence_id,
        "status": body.status,
        "action": body.action,
        "updated_by": body.updated_by,
        "approval_status": approval_status,
        "approved_by": approved_by,
    }


@router.post("/evidences/{evidence_id}/note")
async def update_evidence_note(
    evidence_id: int,
    comment: Optional[str] = Form(None),
    updated_by: Optional[str] = Form(None),
    action: Optional[str] = Form(None),
    attachment: Optional[UploadFile] = File(None),
) -> dict:
    if not comment and not attachment and not action and not updated_by:
        raise HTTPException(status_code=400, detail="Nothing to update")

    attachment_uri = None
    attachment_name = None
    attachment_mime = None
    attachment_size = None
    attachment_sha256 = None

    if attachment is not None:
        attachment_name = os.path.basename(attachment.filename or "attachment.bin")
        attachment_mime = attachment.content_type or "application/octet-stream"
        object_key = f"evidence-notes/{evidence_id}/{uuid4()}-{attachment_name}"

        async def _stream_file():
            while True:
                chunk = await attachment.read(1024 * 1024)
                if not chunk:
                    break
                yield chunk

        if s3_ready():
            try:
                attachment_size, attachment_sha256 = await upload_stream_to_s3(
                    object_key, _stream_file(), attachment_mime
                )
                attachment_uri = object_uri(object_key)
            except Exception:
                attachment_uri = None
                try:
                    await attachment.seek(0)
                except Exception:
                    pass

        if not attachment_uri:
            local_dir = LOCAL_EVIDENCE_ROOT / "evidence-notes" / str(evidence_id)
            local_dir.mkdir(parents=True, exist_ok=True)
            target_path = local_dir / f"{uuid4()}-{attachment_name}"
            hasher = hashlib.sha256()
            size = 0
            with target_path.open("wb") as f:
                async for chunk in _stream_file():
                    if not chunk:
                        continue
                    f.write(chunk)
                    hasher.update(chunk)
                    size += len(chunk)
            attachment_size = size
            attachment_sha256 = hasher.hexdigest()
            attachment_uri = f"file://{target_path.as_posix()}"

    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE evidence
            SET last_comment = COALESCE($1, last_comment),
                last_action = COALESCE($2, last_action),
                updated_by = COALESCE($3, updated_by),
                last_update = NOW(),
                updated_at = NOW(),
                attachment_name = COALESCE($4, attachment_name),
                attachment_uri = COALESCE($5, attachment_uri),
                attachment_mime = COALESCE($6, attachment_mime),
                attachment_size = COALESCE($7, attachment_size),
                attachment_sha256 = COALESCE($8, attachment_sha256)
            WHERE id = $9
            """,
            comment,
            action or ("attachment_uploaded" if attachment else "note_updated"),
            updated_by,
            attachment_name,
            attachment_uri,
            attachment_mime,
            attachment_size,
            attachment_sha256,
            evidence_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Evidence not found")

        details = None
        if comment or attachment_name or attachment_uri:
            details = json.dumps(
                {
                    "comment": comment,
                    "attachment_name": attachment_name,
                    "attachment_uri": attachment_uri,
                }
            )
        await conn.execute(
            """
            INSERT INTO evidence_audit (evidence_id, action, actor, details_json)
            VALUES ($1, $2, $3, $4::jsonb)
            """,
            evidence_id,
            action or ("attachment_uploaded" if attachment else "note_updated"),
            updated_by or "system",
            details,
        )

        row = await conn.fetchrow(
            """
            SELECT last_comment, last_action, updated_by, last_update,
                   attachment_name, attachment_uri, attachment_mime,
                   attachment_size, attachment_sha256
            FROM evidence
            WHERE id = $1
            """,
            evidence_id,
        )

    attachment_download_url = None
    if row and row["attachment_uri"]:
        uri = row["attachment_uri"]
        if uri.startswith("s3://") or uri.startswith("file://"):
            attachment_download_url = f"/admin/evidences/{evidence_id}/attachment:download"
        else:
            attachment_download_url = uri

    return {
        "ok": True,
        "evidence_id": evidence_id,
        "last_comment": row["last_comment"] if row else None,
        "last_action": row["last_action"] if row else None,
        "updated_by": row["updated_by"] if row else None,
        "last_update": row["last_update"].isoformat() if row and row["last_update"] else None,
        "attachment_name": row["attachment_name"] if row else None,
        "attachment_uri": row["attachment_uri"] if row else None,
        "attachment_mime": row["attachment_mime"] if row else None,
        "attachment_size": row["attachment_size"] if row else None,
        "attachment_sha256": row["attachment_sha256"] if row else None,
        "attachment_download_url": attachment_download_url,
    }


@router.get("/evidences/{evidence_id}/attachment:download")
async def evidence_attachment_download(evidence_id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT attachment_uri, attachment_name, attachment_mime
            FROM evidence
            WHERE id=$1
            """,
            evidence_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Evidence not found")
        uri = row["attachment_uri"] or ""
        if not uri:
            raise HTTPException(status_code=404, detail="Attachment not found")
        filename = row["attachment_name"] or f"evidence-attachment-{evidence_id}"
        mime = row["attachment_mime"] or "application/octet-stream"

    if uri.startswith("file://"):
        local_path = _local_path_from_uri(uri)
        if not local_path or not local_path.exists():
            raise HTTPException(status_code=404, detail="Attachment file not found")
        return FileResponse(
            local_path,
            media_type=mime,
            filename=filename,
        )

    if uri.startswith("s3://"):
        key = _s3_key_from_uri(uri)
        if not key:
            raise HTTPException(status_code=400, detail="Invalid attachment URI")
        try:
            obj = get_object(key)
        except Exception:
            raise HTTPException(status_code=404, detail="Attachment not found in S3")
        return StreamingResponse(
            obj["Body"],
            media_type=mime,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return StreamingResponse(
        io.BytesIO(),
        media_type="text/plain",
        status_code=400,
    )


# ----------------- AI System Registry -----------------
@router.get("/ai-systems/helper")
async def list_ai_system_registry_helper() -> dict:
    """Return field_name, description, and helper_values (dropdown options) for AI system register form."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT field_name, description, helper_values
            FROM ai_system_registry_helper
            ORDER BY id
            """
        )
    def _parse_helper_values(value: Any) -> list:
        if isinstance(value, list):
            return value
        if value is None:
            return []
        if isinstance(value, (bytes, bytearray)):
            try:
                value = value.decode("utf-8")
            except Exception:
                return []
        if isinstance(value, str):
            raw = value.strip()
            if not raw:
                return []
            try:
                parsed = json.loads(raw)
            except Exception:
                try:
                    parsed = json.loads(raw.replace('""', '"'))
                except Exception:
                    return []
            if isinstance(parsed, list):
                return parsed
            if isinstance(parsed, dict) and isinstance(parsed.get("items"), list):
                return parsed["items"]
            return []
        return []
    items = [
        {
            "field_name": r["field_name"],
            "description": r["description"] or "",
            "helper_values": _parse_helper_values(r["helper_values"]),
        }
        for r in rows
    ]
    return {"items": items}


def _parse_helper_values_raw(value: Any) -> list:
    """Parse helper_values from DB (list, json string, or bytes) to list of strings."""
    if isinstance(value, list):
        return [str(x).strip() for x in value if x is not None and str(x).strip()]
    if value is None:
        return []
    if isinstance(value, (bytes, bytearray)):
        try:
            value = value.decode("utf-8")
        except Exception:
            return []
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return []
        try:
            parsed = json.loads(raw)
        except Exception:
            try:
                parsed = json.loads(raw.replace('""', '"'))
            except Exception:
                return []
        if isinstance(parsed, list):
            return [str(x).strip() for x in parsed if x is not None and str(x).strip()]
        if isinstance(parsed, dict) and isinstance(parsed.get("items"), list):
            return [str(x).strip() for x in parsed["items"] if x is not None and str(x).strip()]
        return []
    return []


_MODEL_PROVIDER_FIELD = "model_provider"


@router.get("/ai-systems/helper/model-providers")
async def list_model_providers() -> dict:
    """Return the list of model provider options for the AI system register (field model_provider)."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT helper_values FROM ai_system_registry_helper
            WHERE field_name = $1
            """,
            _MODEL_PROVIDER_FIELD,
        )
    if not row:
        return {"values": []}
    values = _parse_helper_values_raw(row["helper_values"])
    return {"values": values}


class ModelProvidersPutBody(BaseModel):
    values: List[str] = Field(..., description="Full list of model provider names")


class ModelProviderAddBody(BaseModel):
    value: str = Field(..., min_length=1, description="Model provider name to add")


class ModelProviderRemoveBody(BaseModel):
    value: str = Field(..., min_length=1, description="Model provider name to remove")


class ModelProviderUpdateBody(BaseModel):
    old_value: str = Field(..., min_length=1, description="Current name")
    new_value: str = Field(..., min_length=1, description="New name")


@router.put("/ai-systems/helper/model-providers")
async def replace_model_providers(body: ModelProvidersPutBody) -> dict:
    """Replace the full list of model provider options."""
    pool = await get_pool()
    values = [v.strip() for v in body.values if v and v.strip()]
    async with pool.acquire() as conn:
        await conn.execute(
            """
            UPDATE ai_system_registry_helper
            SET helper_values = $1::jsonb
            WHERE field_name = $2
            """,
            json.dumps(values),
            _MODEL_PROVIDER_FIELD,
        )
        row = await conn.fetchrow(
            "SELECT helper_values FROM ai_system_registry_helper WHERE field_name = $1",
            _MODEL_PROVIDER_FIELD,
        )
    if not row:
        raise HTTPException(status_code=404, detail="model_provider helper row not found")
    return {"values": _parse_helper_values_raw(row["helper_values"])}


@router.post("/ai-systems/helper/model-providers")
async def add_model_provider(body: ModelProviderAddBody) -> dict:
    """Append a model provider if not already present."""
    pool = await get_pool()
    add_val = body.value.strip()
    if not add_val:
        raise HTTPException(status_code=400, detail="value must be non-empty")
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT helper_values FROM ai_system_registry_helper WHERE field_name = $1",
            _MODEL_PROVIDER_FIELD,
        )
        if not row:
            raise HTTPException(status_code=404, detail="model_provider helper row not found")
        current = _parse_helper_values_raw(row["helper_values"])
        if add_val in current:
            return {"values": current}
        new_values = current + [add_val]
        await conn.execute(
            "UPDATE ai_system_registry_helper SET helper_values = $1::jsonb WHERE field_name = $2",
            json.dumps(new_values),
            _MODEL_PROVIDER_FIELD,
        )
    return {"values": new_values}


@router.delete("/ai-systems/helper/model-providers")
async def remove_model_provider(body: ModelProviderRemoveBody) -> dict:
    """Remove a model provider by value."""
    pool = await get_pool()
    remove_val = body.value.strip()
    if not remove_val:
        raise HTTPException(status_code=400, detail="value must be non-empty")
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT helper_values FROM ai_system_registry_helper WHERE field_name = $1",
            _MODEL_PROVIDER_FIELD,
        )
        if not row:
            raise HTTPException(status_code=404, detail="model_provider helper row not found")
        current = _parse_helper_values_raw(row["helper_values"])
        new_values = [v for v in current if v != remove_val]
        if len(new_values) == len(current):
            return {"values": current}
        await conn.execute(
            "UPDATE ai_system_registry_helper SET helper_values = $1::jsonb WHERE field_name = $2",
            json.dumps(new_values),
            _MODEL_PROVIDER_FIELD,
        )
    return {"values": new_values}


@router.patch("/ai-systems/helper/model-providers")
async def update_model_provider(body: ModelProviderUpdateBody) -> dict:
    """Rename one model provider (old_value -> new_value)."""
    pool = await get_pool()
    old_val = body.old_value.strip()
    new_val = body.new_value.strip()
    if not old_val or not new_val:
        raise HTTPException(status_code=400, detail="old_value and new_value must be non-empty")
    if old_val == new_val:
        async with pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT helper_values FROM ai_system_registry_helper WHERE field_name = $1",
                _MODEL_PROVIDER_FIELD,
            )
        if not row:
            raise HTTPException(status_code=404, detail="model_provider helper row not found")
        return {"values": _parse_helper_values_raw(row["helper_values"])}
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT helper_values FROM ai_system_registry_helper WHERE field_name = $1",
            _MODEL_PROVIDER_FIELD,
        )
        if not row:
            raise HTTPException(status_code=404, detail="model_provider helper row not found")
        current = _parse_helper_values_raw(row["helper_values"])
        try:
            idx = current.index(old_val)
        except ValueError:
            return {"values": current}
        new_values = list(current)
        new_values[idx] = new_val
        await conn.execute(
            "UPDATE ai_system_registry_helper SET helper_values = $1::jsonb WHERE field_name = $2",
            json.dumps(new_values),
            _MODEL_PROVIDER_FIELD,
        )
    return {"values": new_values}


@router.get("/ai-systems")
async def list_ai_systems(
    limit: int = 50,
    offset: int = 0,
    q: Optional[str] = None,
    project_slug: Optional[str] = None,
    risk_tier: Optional[str] = None,
    status: Optional[str] = None,
    locale: Optional[str] = None,
    entity_id: Optional[UUID] = Query(default=None, description="Filter by entity; only systems for this entity are returned"),
) -> dict:
    normalized_locale = _normalize_locale(locale)
    where = []
    params = []
    join_sql = ""
    if normalized_locale:
        params.append(normalized_locale)
        locale_param = len(params)
        join_sql = (
            f"LEFT JOIN ai_system_translations t ON t.system_id = s.id AND t.locale = ${locale_param}"
        )
    if entity_id is not None:
        params.append(entity_id)
        where.append(f"s.entity_id = ${len(params)}")
    if project_slug:
        params.append(project_slug)
        where.append(f"s.project_slug = ${len(params)}")
    if risk_tier:
        params.append(risk_tier)
        where.append(f"s.risk_tier = ${len(params)}")
    if status:
        params.append(status)
        where.append(f"s.status = ${len(params)}")
    if q:
        params.append(f"%{q}%")
        name_expr = "COALESCE(t.name, s.name)" if normalized_locale else "s.name"
        model_provider_expr = "COALESCE(t.model_provider, s.model_provider)" if normalized_locale else "s.model_provider"
        where.append(
            f"(s.uc_id ILIKE ${len(params)} OR {name_expr} ILIKE ${len(params)} OR {model_provider_expr} ILIKE ${len(params)})"
        )

    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    from_sql = f"FROM ai_system_registry s {join_sql}"
    name_expr = "COALESCE(t.name, s.name)" if normalized_locale else "s.name"
    description_expr = "COALESCE(t.description, s.description)" if normalized_locale else "s.description"
    owner_expr = "COALESCE(t.owner, s.owner)" if normalized_locale else "s.owner"
    business_unit_expr = "COALESCE(t.business_unit, s.business_unit)" if normalized_locale else "s.business_unit"
    model_provider_expr = "COALESCE(t.model_provider, s.model_provider)" if normalized_locale else "s.model_provider"
    provider_type_expr = "COALESCE(t.provider_type, s.provider_type)" if normalized_locale else "s.provider_type"
    intended_use_expr = "COALESCE(t.intended_use, s.intended_use)" if normalized_locale else "s.intended_use"
    intended_users_expr = "COALESCE(t.intended_users, s.intended_users)" if normalized_locale else "s.intended_users"
    system_boundary_expr = "COALESCE(t.system_boundary, s.system_boundary)" if normalized_locale else "s.system_boundary"
    model_type_expr = "COALESCE(t.model_type, s.model_type)" if normalized_locale else "s.model_type"
    model_version_expr = "COALESCE(t.model_version, s.model_version)" if normalized_locale else "s.model_version"
    deployment_environment_expr = (
        "COALESCE(t.deployment_environment, s.deployment_environment)"
        if normalized_locale
        else "s.deployment_environment"
    )
    lifecycle_stage_expr = (
        "COALESCE(t.lifecycle_stage, s.lifecycle_stage)"
        if normalized_locale
        else "s.lifecycle_stage"
    )
    training_data_sources_expr = (
        "COALESCE(t.training_data_sources, s.training_data_sources)"
        if normalized_locale
        else "s.training_data_sources"
    )
    risk_tier_expr = "COALESCE(t.risk_tier, s.risk_tier)" if normalized_locale else "s.risk_tier"
    status_expr = "COALESCE(t.status, s.status)" if normalized_locale else "s.status"
    region_scope_expr = "COALESCE(t.region_scope, s.region_scope)" if normalized_locale else "s.region_scope"
    data_sensitivity_expr = (
        "COALESCE(t.data_sensitivity, s.data_sensitivity)"
        if normalized_locale
        else "s.data_sensitivity"
    )
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) {from_sql} {where_sql}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT s.id, s.uc_id, s.project_slug,
                   {name_expr} AS name,
                   {description_expr} AS description,
                   {owner_expr} AS owner,
                   s.system_owner_email,
                   {business_unit_expr} AS business_unit,
                   s.risk_owner_role,
                   s.decision_authority,
                   {model_provider_expr} AS model_provider,
                   {provider_type_expr} AS provider_type,
                   {intended_use_expr} AS intended_use,
                   {intended_users_expr} AS intended_users,
                   {system_boundary_expr} AS system_boundary,
                   {model_type_expr} AS model_type,
                   {model_version_expr} AS model_version,
                   {deployment_environment_expr} AS deployment_environment,
                   {lifecycle_stage_expr} AS lifecycle_stage,
                   {training_data_sources_expr} AS training_data_sources,
                   s.personal_data_flag,
                   s.sensitive_attributes_flag,
                   {risk_tier_expr} AS risk_tier,
                   {status_expr} AS status,
                   {region_scope_expr} AS region_scope,
                   {data_sensitivity_expr} AS data_sensitivity,
                   s.model_name, s.technical_lead, s.target_users, s.intended_purpose,
                   s.out_of_scope_uses, s.deployment_method, s.data_residency,
                   s.base_model_type, s.input_output_modality, s.fine_tuning_data,
                   s.data_minimization, s.human_oversight_mechanism,
                   s.impact_assessment_reference, s.known_limitations,
                   s.langfuse_project_id, s.langfuse_base_url,
                   s.entity_id, s.entity_slug,
                   s.created_at, s.updated_at
            {from_sql}
            {where_sql}
            ORDER BY s.updated_at DESC NULLS LAST, s.created_at DESC
            LIMIT ${len(params)+1} OFFSET ${len(params)+2}
            """,
            *params,
            limit,
            offset,
        )

    items = []
    for r in rows:
        items.append(
            {
                "id": r["id"],
                "uc_id": r["uc_id"],
                "project_slug": r["project_slug"],
                "name": r["name"],
                "description": r["description"],
                "owner": r["owner"],
                "system_owner_email": r["system_owner_email"],
                "business_unit": r["business_unit"],
                "risk_owner_role": r["risk_owner_role"],
                "decision_authority": r["decision_authority"],
                "model_provider": r["model_provider"],
                "provider_type": r["provider_type"],
                "intended_use": r["intended_use"],
                "intended_users": r["intended_users"],
                "system_boundary": r["system_boundary"],
                "model_type": r["model_type"],
                "model_version": r["model_version"],
                "deployment_environment": r["deployment_environment"],
                "lifecycle_stage": r["lifecycle_stage"],
                "training_data_sources": r["training_data_sources"],
                "personal_data_flag": r["personal_data_flag"],
                "sensitive_attributes_flag": r["sensitive_attributes_flag"],
                "risk_tier": r["risk_tier"],
                "status": r["status"],
                "region_scope": r["region_scope"],
                "data_sensitivity": r["data_sensitivity"],
                "model_name": r.get("model_name"),
                "technical_lead": r.get("technical_lead"),
                "target_users": r.get("target_users"),
                "intended_purpose": r.get("intended_purpose"),
                "out_of_scope_uses": r.get("out_of_scope_uses"),
                "deployment_method": r.get("deployment_method"),
                "data_residency": r.get("data_residency"),
                "base_model_type": r.get("base_model_type"),
                "input_output_modality": r.get("input_output_modality"),
                "fine_tuning_data": r.get("fine_tuning_data"),
                "data_minimization": r.get("data_minimization"),
                "human_oversight_mechanism": r.get("human_oversight_mechanism"),
                "impact_assessment_reference": r.get("impact_assessment_reference"),
                "known_limitations": r.get("known_limitations"),
                "langfuse_project_id": r.get("langfuse_project_id"),
                "langfuse_base_url": r.get("langfuse_base_url"),
                "entity_id": str(r["entity_id"]) if r["entity_id"] else None,
                "entity_slug": r["entity_slug"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
            }
        )
    return {"items": items, "total": int(total or 0)}


@router.post("/ai-systems")
async def create_ai_system(body: AiSystemIn) -> dict:
    system_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        project_slug = _clean_optional_str(body.project_slug)
        entity_id = None
        entity_slug = None
        if project_slug:
            entity_row = await conn.fetchrow(
                """
                SELECT e.id, e.slug
                FROM entity_projects p
                JOIN entity e ON e.id = p.entity_id
                WHERE p.slug = $1
                """,
                project_slug,
            )
            if entity_row:
                entity_id = entity_row["id"]
                entity_slug = entity_row["slug"]
        name = _clean_optional_str(body.name)
        if not name:
            raise HTTPException(status_code=400, detail="Name is required")
        uc_project_slug = project_slug or "general"
        uc_id = await _generate_unique_uc_id(conn, uc_project_slug, name)
        await conn.execute(
            """
            INSERT INTO ai_system_registry (
              id, uc_id, project_slug, name, description, owner, system_owner_email,
              business_unit, risk_owner_role, decision_authority, model_provider, provider_type,
              intended_use, intended_users, system_boundary, model_type, model_version,
              deployment_environment, lifecycle_stage, training_data_sources,
              personal_data_flag, sensitive_attributes_flag,
              risk_tier, status, region_scope, data_sensitivity,
              model_name, technical_lead, target_users, intended_purpose, out_of_scope_uses,
              deployment_method, data_residency, base_model_type, input_output_modality,
              fine_tuning_data, data_minimization, human_oversight_mechanism,
              impact_assessment_reference, known_limitations,
              langfuse_project_id, langfuse_base_url,
              entity_id, entity_slug, created_at, updated_at
            )
            VALUES (
              $1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18,$19,
              $20,$21,$22,$23,COALESCE($24,'active'),$25,$26,
              $27,$28,$29,$30,$31,$32,$33,$34,$35,$36,$37,$38,$39,$40,$41,$42,$43,$44,NOW(),NOW()
            )
            """,
            system_id,
            uc_id,
            project_slug,
            name,
            body.description,
            body.owner,
            body.system_owner_email,
            body.business_unit,
            body.risk_owner_role,
            body.decision_authority,
            body.model_provider,
            body.provider_type,
            body.intended_use,
            body.intended_users,
            body.system_boundary,
            body.model_type,
            body.model_version,
            body.deployment_environment,
            body.lifecycle_stage,
            body.training_data_sources,
            body.personal_data_flag,
            body.sensitive_attributes_flag,
            body.risk_tier,
            body.status,
            body.region_scope,
            body.data_sensitivity,
            body.model_name,
            body.technical_lead,
            body.target_users,
            body.intended_purpose,
            body.out_of_scope_uses,
            body.deployment_method,
            body.data_residency,
            body.base_model_type,
            body.input_output_modality,
            body.fine_tuning_data,
            body.data_minimization,
            body.human_oversight_mechanism,
            body.impact_assessment_reference,
            body.known_limitations,
            body.langfuse_project_id,
            body.langfuse_base_url,
            entity_id,
            entity_slug,
        )
    await _log_audit(
        event_type="ai_system_created",
        actor="system",
        source_service="core-svc",
        object_type="ai_system",
        object_id=system_id,
        project_slug=project_slug,
        details={"uc_id": uc_id, "name": name},
    )
    return {"ok": True, "id": system_id}


@router.put("/ai-systems/{system_id}")
async def update_ai_system(system_id: str, body: AiSystemUpdate) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        entity_id = None
        entity_slug = None
        project_slug = _clean_optional_str(body.project_slug)
        if project_slug:
            entity_row = await conn.fetchrow(
                """
                SELECT e.id, e.slug
                FROM entity_projects p
                JOIN entity e ON e.id = p.entity_id
                WHERE p.slug = $1
                """,
                project_slug,
            )
            if entity_row:
                entity_id = entity_row["id"]
                entity_slug = entity_row["slug"]
        res = await conn.execute(
            """
            UPDATE ai_system_registry
            SET project_slug = COALESCE($1, project_slug),
                name = COALESCE($2, name),
                description = COALESCE($3, description),
                owner = COALESCE($4, owner),
                system_owner_email = COALESCE($5, system_owner_email),
                business_unit = COALESCE($6, business_unit),
                risk_owner_role = COALESCE($7, risk_owner_role),
                decision_authority = COALESCE($8, decision_authority),
                model_provider = COALESCE($9, model_provider),
                provider_type = COALESCE($10, provider_type),
                intended_use = COALESCE($11, intended_use),
                intended_users = COALESCE($12, intended_users),
                system_boundary = COALESCE($13, system_boundary),
                model_type = COALESCE($14, model_type),
                model_version = COALESCE($15, model_version),
                deployment_environment = COALESCE($16, deployment_environment),
                lifecycle_stage = COALESCE($17, lifecycle_stage),
                training_data_sources = COALESCE($18, training_data_sources),
                personal_data_flag = COALESCE($19, personal_data_flag),
                sensitive_attributes_flag = COALESCE($20, sensitive_attributes_flag),
                risk_tier = COALESCE($21, risk_tier),
                status = COALESCE($22, status),
                region_scope = COALESCE($23, region_scope),
                data_sensitivity = COALESCE($24, data_sensitivity),
                model_name = COALESCE($25, model_name),
                technical_lead = COALESCE($26, technical_lead),
                target_users = COALESCE($27, target_users),
                intended_purpose = COALESCE($28, intended_purpose),
                out_of_scope_uses = COALESCE($29, out_of_scope_uses),
                deployment_method = COALESCE($30, deployment_method),
                data_residency = COALESCE($31, data_residency),
                base_model_type = COALESCE($32, base_model_type),
                input_output_modality = COALESCE($33, input_output_modality),
                fine_tuning_data = COALESCE($34, fine_tuning_data),
                data_minimization = COALESCE($35, data_minimization),
                human_oversight_mechanism = COALESCE($36, human_oversight_mechanism),
                impact_assessment_reference = COALESCE($37, impact_assessment_reference),
                known_limitations = COALESCE($38, known_limitations),
                langfuse_project_id = COALESCE($39, langfuse_project_id),
                langfuse_base_url = COALESCE($40, langfuse_base_url),
                entity_id = COALESCE($41, entity_id),
                entity_slug = COALESCE($42, entity_slug),
                updated_at = NOW()
            WHERE id = $43
            """,
            project_slug,
            body.name,
            body.description,
            body.owner,
            body.system_owner_email,
            body.business_unit,
            body.risk_owner_role,
            body.decision_authority,
            body.model_provider,
            body.provider_type,
            body.intended_use,
            body.intended_users,
            body.system_boundary,
            body.model_type,
            body.model_version,
            body.deployment_environment,
            body.lifecycle_stage,
            body.training_data_sources,
            body.personal_data_flag,
            body.sensitive_attributes_flag,
            body.risk_tier,
            body.status,
            body.region_scope,
            body.data_sensitivity,
            body.model_name,
            body.technical_lead,
            body.target_users,
            body.intended_purpose,
            body.out_of_scope_uses,
            body.deployment_method,
            body.data_residency,
            body.base_model_type,
            body.input_output_modality,
            body.fine_tuning_data,
            body.data_minimization,
            body.human_oversight_mechanism,
            body.impact_assessment_reference,
            body.known_limitations,
            body.langfuse_project_id,
            body.langfuse_base_url,
            entity_id,
            entity_slug,
            system_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="System not found")
    await _log_audit(
        event_type="ai_system_updated",
        actor="system",
        source_service="core-svc",
        object_type="ai_system",
        object_id=system_id,
        project_slug=body.project_slug,
        details={"fields": body.model_dump(exclude_none=True)},
    )
    return {"ok": True, "id": system_id}


@router.post("/ai-systems/{system_id}/retire")
async def retire_ai_system(system_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE ai_system_registry
            SET status = 'retired',
                updated_at = NOW()
            WHERE id = $1
            RETURNING project_slug, uc_id, name
            """,
            system_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="System not found")
    await _log_audit(
        event_type="ai_system_retired",
        actor="system",
        source_service="core-svc",
        object_type="ai_system",
        object_id=system_id,
        project_slug=row["project_slug"],
        details={"uc_id": row["uc_id"], "name": row["name"]},
    )
    return {"ok": True, "id": system_id}


# ----------------- Model Cards -----------------
def _langfuse_prompt_env_name(prompt_key: str) -> str:
    safe = prompt_key.upper().replace("-", "_")
    return f"LANGFUSE_PROMPT_{safe}"


def _extract_langfuse_project(payload: Any) -> dict:
    if isinstance(payload, dict):
        data = payload.get("data")
        if isinstance(data, list) and data:
            if isinstance(data[0], dict):
                return data[0]
        if isinstance(payload.get("project"), dict):
            return payload["project"]
        return payload
    if isinstance(payload, list) and payload:
        if isinstance(payload[0], dict):
            return payload[0]
    return {}


def _extract_langfuse_project_id(project: dict) -> Optional[str]:
    for key in ("id", "projectId", "project_id"):
        value = project.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


@router.post("/langfuse/project-id")
async def resolve_langfuse_project_id(body: LangfuseProjectLookup) -> dict:
    base_url = (body.base_url or os.getenv("LANGFUSE_BASE_URL", "")).strip()
    if not base_url:
        raise HTTPException(status_code=400, detail="Langfuse base URL is required")
    base_url = base_url.rstrip("/")
    url = f"{base_url}/api/public/projects"
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            resp = await client.get(url, auth=(body.public_key, body.secret_key))
        if resp.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Invalid Langfuse project key")
        resp.raise_for_status()
        project = _extract_langfuse_project(resp.json())
        project_id = _extract_langfuse_project_id(project)
        if not project_id:
            raise HTTPException(status_code=502, detail="Langfuse project id missing")
        return {
            "ok": True,
            "project_id": project_id,
            "project_name": project.get("name"),
            "base_url": base_url,
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/langfuse/prompts/{prompt_key}/versions")
async def list_langfuse_prompt_versions(prompt_key: str, limit: int = 5) -> dict:
    prompt_name = os.getenv(_langfuse_prompt_env_name(prompt_key))
    if not prompt_name:
        return {"ok": False, "message": "Prompt key not configured"}

    adapter = LangfuseAdapter()
    if not adapter.is_configured():
        return {"ok": False, "message": "Langfuse not configured"}

    prompt_data = adapter.fetch_prompt_versions(prompt_name, limit=limit)
    if not prompt_data:
        return {"ok": False, "message": "Prompt not found"}

    return {
        "ok": True,
        "prompt_key": prompt_key,
        "prompt_name": prompt_name,
        "prompt": prompt_data,
    }


@router.get("/langfuse/status")
async def langfuse_status() -> dict:
    adapter = LangfuseAdapter()
    if not adapter.base_url:
        return {
            "ok": False,
            "configured": False,
            "reachable": False,
            "message": "Langfuse base URL not set",
        }
    reachable, message = adapter.check_connection()
    return {
        "ok": reachable,
        "configured": adapter.is_configured(),
        "reachable": reachable,
        "message": message,
    }


@router.get("/model-cards")
async def list_model_cards(
    entity_id: Optional[UUID] = None,
    entity_slug: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
) -> dict:
    where = []
    params: list = []
    if entity_id:
        params.append(entity_id)
        where.append(f"s.entity_id = ${len(params)}")
    if entity_slug:
        params.append(entity_slug)
        where.append(f"s.entity_slug = ${len(params)}")
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM ai_system_registry s {where_sql}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT s.id AS system_id, s.name, s.project_slug, s.model_provider,
                   s.model_type, s.model_version, s.owner, s.system_owner_email,
                   s.risk_tier, s.status, s.intended_use, s.intended_users,
                   s.system_boundary, s.training_data_sources, s.personal_data_flag,
                   s.sensitive_attributes_flag, s.lifecycle_stage, s.deployment_environment,
                   s.langfuse_project_id, s.langfuse_base_url, s.entity_id, s.entity_slug,
                   s.created_at, s.updated_at,
                   mc.id AS card_id, mc.version, mc.status AS card_status,
                   mc.summary_md, mc.limitations, mc.out_of_scope, mc.review_cadence,
                   mc.approved_by, mc.approved_at, mc.created_at AS card_created_at,
                   mc.updated_at AS card_updated_at
            FROM ai_system_registry s
            LEFT JOIN LATERAL (
              SELECT *
              FROM model_card
              WHERE system_id = s.id
              ORDER BY version DESC, created_at DESC
              LIMIT 1
            ) mc ON TRUE
            {where_sql}
            ORDER BY s.updated_at DESC NULLS LAST, s.created_at DESC
            LIMIT ${len(params)+1} OFFSET ${len(params)+2}
            """,
            *params,
            limit,
            offset,
        )

        card_ids = [r["card_id"] for r in rows if r.get("card_id")]
        evidence_rows = []
        if card_ids:
            evidence_rows = await conn.fetch(
                """
                SELECT model_card_id, source, metric_key, metric_value, last_seen_at
                FROM model_card_evidence
                WHERE model_card_id = ANY($1::uuid[])
                ORDER BY last_seen_at DESC
                """,
                card_ids,
            )

    evidence_map: dict[str, list] = {}
    for r in evidence_rows:
        key = str(r["model_card_id"])
        evidence_map.setdefault(key, []).append(
            {
                "source": r["source"],
                "metric_key": r["metric_key"],
                "metric_value": r["metric_value"],
                "last_seen_at": r["last_seen_at"].isoformat() if r["last_seen_at"] else None,
            }
        )

    items = []
    for r in rows:
        card_id = r.get("card_id")
        card_id_str = str(card_id) if card_id else None
        items.append(
            {
                "system": {
                    "id": r["system_id"],
                    "name": r["name"],
                    "project_slug": r["project_slug"],
                    "model_provider": r["model_provider"],
                    "model_type": r["model_type"],
                    "model_version": r["model_version"],
                    "owner": r["owner"],
                    "system_owner_email": r["system_owner_email"],
                    "risk_tier": r["risk_tier"],
                    "status": r["status"],
                    "intended_use": r["intended_use"],
                    "intended_users": r["intended_users"],
                    "system_boundary": r["system_boundary"],
                    "training_data_sources": r["training_data_sources"],
                    "personal_data_flag": r["personal_data_flag"],
                    "sensitive_attributes_flag": r["sensitive_attributes_flag"],
                    "lifecycle_stage": r["lifecycle_stage"],
                    "deployment_environment": r["deployment_environment"],
                    "langfuse_project_id": r["langfuse_project_id"],
                    "langfuse_base_url": r["langfuse_base_url"],
                    "entity_id": str(r["entity_id"]) if r["entity_id"] else None,
                    "entity_slug": r["entity_slug"],
                    "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                    "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
                },
                "model_card": {
                    "id": card_id_str,
                    "version": r["version"],
                    "status": r["card_status"],
                    "summary_md": r["summary_md"],
                    "limitations": r["limitations"],
                    "out_of_scope": r["out_of_scope"],
                    "review_cadence": r["review_cadence"],
                    "approved_by": r["approved_by"],
                    "approved_at": r["approved_at"].isoformat() if r["approved_at"] else None,
                    "created_at": r["card_created_at"].isoformat() if r["card_created_at"] else None,
                    "updated_at": r["card_updated_at"].isoformat() if r["card_updated_at"] else None,
                }
                if card_id
                else None,
                "evidence": evidence_map.get(card_id_str, []),
            }
        )

    return {"items": items, "total": int(total or 0)}


@router.get("/model-cards/{system_id}")
async def get_model_card(system_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT s.id AS system_id, s.name, s.project_slug, s.model_provider,
                   s.model_type, s.model_version, s.owner, s.system_owner_email,
                   s.risk_tier, s.status, s.intended_use, s.intended_users,
                   s.system_boundary, s.training_data_sources, s.personal_data_flag,
                   s.sensitive_attributes_flag, s.lifecycle_stage, s.deployment_environment,
                   s.langfuse_project_id, s.langfuse_base_url, s.entity_id, s.entity_slug,
                   s.created_at, s.updated_at,
                   mc.id AS card_id, mc.version, mc.status AS card_status,
                   mc.summary_md, mc.limitations, mc.out_of_scope, mc.review_cadence,
                   mc.approved_by, mc.approved_at, mc.created_at AS card_created_at,
                   mc.updated_at AS card_updated_at
            FROM ai_system_registry s
            LEFT JOIN LATERAL (
              SELECT *
              FROM model_card
              WHERE system_id = s.id
              ORDER BY version DESC, created_at DESC
              LIMIT 1
            ) mc ON TRUE
            WHERE s.id = $1
            """,
            system_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="System not found")

        evidence_rows = []
        if row.get("card_id"):
            evidence_rows = await conn.fetch(
                """
                SELECT source, metric_key, metric_value, last_seen_at
                FROM model_card_evidence
                WHERE model_card_id = $1
                ORDER BY last_seen_at DESC
                """,
                row["card_id"],
            )

    return {
        "system": {
            "id": row["system_id"],
            "name": row["name"],
            "project_slug": row["project_slug"],
            "model_provider": row["model_provider"],
            "model_type": row["model_type"],
            "model_version": row["model_version"],
            "owner": row["owner"],
            "system_owner_email": row["system_owner_email"],
            "risk_tier": row["risk_tier"],
            "status": row["status"],
            "intended_use": row["intended_use"],
            "intended_users": row["intended_users"],
            "system_boundary": row["system_boundary"],
            "training_data_sources": row["training_data_sources"],
            "personal_data_flag": row["personal_data_flag"],
            "sensitive_attributes_flag": row["sensitive_attributes_flag"],
            "lifecycle_stage": row["lifecycle_stage"],
            "deployment_environment": row["deployment_environment"],
            "langfuse_project_id": row["langfuse_project_id"],
            "langfuse_base_url": row["langfuse_base_url"],
            "entity_id": str(row["entity_id"]) if row["entity_id"] else None,
            "entity_slug": row["entity_slug"],
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
        },
        "model_card": {
            "id": str(row["card_id"]) if row["card_id"] else None,
            "version": row["version"],
            "status": row["card_status"],
            "summary_md": row["summary_md"],
            "limitations": row["limitations"],
            "out_of_scope": row["out_of_scope"],
            "review_cadence": row["review_cadence"],
            "approved_by": row["approved_by"],
            "approved_at": row["approved_at"].isoformat() if row["approved_at"] else None,
            "created_at": row["card_created_at"].isoformat() if row["card_created_at"] else None,
            "updated_at": row["card_updated_at"].isoformat() if row["card_updated_at"] else None,
        }
        if row.get("card_id")
        else None,
        "evidence": [
            {
                "source": r["source"],
                "metric_key": r["metric_key"],
                "metric_value": r["metric_value"],
                "last_seen_at": r["last_seen_at"].isoformat() if r["last_seen_at"] else None,
            }
            for r in evidence_rows
        ],
    }


@router.post("/model-cards/{system_id}")
async def upsert_model_card(system_id: str, body: ModelCardUpsert) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        exists = await conn.fetchrow(
            "SELECT id FROM ai_system_registry WHERE id = $1",
            system_id,
        )
        if not exists:
            raise HTTPException(status_code=404, detail="System not found")

        latest = await conn.fetchrow(
            """
            SELECT id, version
            FROM model_card
            WHERE system_id = $1
            ORDER BY version DESC, created_at DESC
            LIMIT 1
            """,
            system_id,
        )

        if latest:
            updates = {
                "summary_md": body.summary_md,
                "limitations": body.limitations,
                "out_of_scope": body.out_of_scope,
                "review_cadence": body.review_cadence,
                "status": body.status,
                "approved_by": body.approved_by,
                "approved_at": body.approved_at,
            }
            set_parts = []
            params: list = [latest["id"]]
            for key, value in updates.items():
                if value is None:
                    continue
                params.append(value)
                set_parts.append(f"{key} = ${len(params)}")
            if set_parts:
                set_parts.append("updated_at = now()")
                await conn.execute(
                    f"UPDATE model_card SET {', '.join(set_parts)} WHERE id = $1",
                    *params,
                )
            card_id = latest["id"]
            version = latest["version"]
        else:
            card_id = await conn.fetchval(
                """
                INSERT INTO model_card (
                  system_id, version, status, summary_md, limitations,
                  out_of_scope, review_cadence, approved_by, approved_at,
                  created_at, updated_at
                )
                VALUES ($1, 1, COALESCE($2, 'draft'), $3, $4, $5, $6, $7, $8, NOW(), NOW())
                RETURNING id
                """,
                system_id,
                body.status,
                body.summary_md,
                body.limitations,
                body.out_of_scope,
                body.review_cadence,
                body.approved_by,
                body.approved_at,
            )
            version = 1

    return {"ok": True, "id": str(card_id), "version": version}


@router.post("/model-cards/{system_id}/sync-langfuse")
async def sync_model_card_langfuse(system_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        system = await conn.fetchrow(
            """
            SELECT id, langfuse_project_id, langfuse_base_url, model_provider, model_version
            FROM ai_system_registry
            WHERE id = $1
            """,
            system_id,
        )
        if not system:
            raise HTTPException(status_code=404, detail="System not found")
        if not system["langfuse_project_id"]:
            return {"ok": False, "message": "Langfuse project id not configured"}

        latest = await conn.fetchrow(
            """
            SELECT id
            FROM model_card
            WHERE system_id = $1
            ORDER BY version DESC, created_at DESC
            LIMIT 1
            """,
            system_id,
        )
        card_id = latest["id"] if latest else None
        if not card_id:
            card_id = await conn.fetchval(
                """
                INSERT INTO model_card (system_id, version, status, created_at, updated_at)
                VALUES ($1, 1, 'draft', NOW(), NOW())
                RETURNING id
                """,
                system_id,
            )

    adapter = LangfuseAdapter(base_url=system["langfuse_base_url"])
    metrics: list[LangfuseMetric] = adapter.fetch_project_metrics(
        system["langfuse_project_id"]
    )
    metadata = adapter.fetch_project_metadata(system["langfuse_project_id"])
    if not metrics:
        return {"ok": False, "message": "No Langfuse metrics available"}

    pool = await get_pool()
    async with pool.acquire() as conn:
        update_fields = {}
        if metadata.get("model_provider") and not system.get("model_provider"):
            update_fields["model_provider"] = metadata.get("model_provider")
        if metadata.get("model_version") and not system.get("model_version"):
            update_fields["model_version"] = metadata.get("model_version")
        if update_fields:
            set_parts = []
            params = [system_id]
            for key, value in update_fields.items():
                params.append(value)
                set_parts.append(f"{key} = ${len(params)}")
            set_parts.append("updated_at = now()")
            await conn.execute(
                f"UPDATE ai_system_registry SET {', '.join(set_parts)} WHERE id = $1",
                *params,
            )

        for metric in metrics:
            await conn.execute(
                """
                INSERT INTO model_card_evidence (
                  model_card_id, source, metric_key, metric_value, last_seen_at
                )
                VALUES ($1, $2, $3, $4::jsonb, NOW())
                ON CONFLICT (model_card_id, source, metric_key)
                DO UPDATE SET metric_value = EXCLUDED.metric_value, last_seen_at = NOW()
                """,
                card_id,
                metric.source,
                metric.metric_key,
                json.dumps(metric.metric_value),
            )
        if metadata.get("model_provider"):
            await conn.execute(
                """
                INSERT INTO model_card_evidence (
                  model_card_id, source, metric_key, metric_value, last_seen_at
                )
                VALUES ($1, $2, $3, $4::jsonb, NOW())
                ON CONFLICT (model_card_id, source, metric_key)
                DO UPDATE SET metric_value = EXCLUDED.metric_value, last_seen_at = NOW()
                """,
                card_id,
                "langfuse_metadata",
                "model_provider",
                json.dumps(metadata.get("model_provider")),
            )
        if metadata.get("model_version"):
            await conn.execute(
                """
                INSERT INTO model_card_evidence (
                  model_card_id, source, metric_key, metric_value, last_seen_at
                )
                VALUES ($1, $2, $3, $4::jsonb, NOW())
                ON CONFLICT (model_card_id, source, metric_key)
                DO UPDATE SET metric_value = EXCLUDED.metric_value, last_seen_at = NOW()
                """,
                card_id,
                "langfuse_metadata",
                "model_version",
                json.dumps(metadata.get("model_version")),
            )

    return {"ok": True, "count": len(metrics)}


# ----------------- Data Source Connectors (Postgres) -----------------
@router.get("/data-sources", response_model=List[DataSourceConnectorOut])
async def list_data_sources() -> List[DataSourceConnectorOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, name, type, host, port, database, username,
                   ssl_mode, status, last_tested_at, last_test_status,
                   last_test_error, created_at, updated_at
            FROM data_source_connectors
            ORDER BY updated_at DESC, created_at DESC
            """
        )
    return [DataSourceConnectorOut(**dict(r)) for r in rows]


@router.post("/data-sources", response_model=DataSourceConnectorOut, status_code=201)
async def create_data_source(body: DataSourceConnectorIn) -> DataSourceConnectorOut:
    connector_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            await conn.execute(
                """
                INSERT INTO data_source_connectors (
                  id, name, type, host, port, database, username, password,
                  ssl_mode, status, created_at, updated_at
                )
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,COALESCE($10,'active'),NOW(),NOW())
                """,
                connector_id,
                body.name,
                body.type,
                body.host,
                body.port,
                body.database,
                body.username,
                _clean_optional_str(body.password),
                _clean_optional_str(body.ssl_mode),
                _clean_optional_str(body.status),
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(status_code=409, detail="Connector name already exists")

        row = await conn.fetchrow(
            """
            SELECT id, name, type, host, port, database, username,
                   ssl_mode, status, last_tested_at, last_test_status,
                   last_test_error, created_at, updated_at
            FROM data_source_connectors
            WHERE id = $1
            """,
            connector_id,
        )
    await _log_audit(
        event_type="data_source_created",
        actor="system",
        source_service="core-svc",
        object_type="data_source",
        object_id=connector_id,
        details={"name": body.name, "type": body.type},
    )
    return DataSourceConnectorOut(**dict(row))


@router.put("/data-sources/{connector_id}", response_model=DataSourceConnectorOut)
async def update_data_source(
    connector_id: str,
    body: DataSourceConnectorUpdate,
) -> DataSourceConnectorOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE data_source_connectors
            SET name = COALESCE($1, name),
                host = COALESCE($2, host),
                port = COALESCE($3, port),
                database = COALESCE($4, database),
                username = COALESCE($5, username),
                password = COALESCE($6, password),
                ssl_mode = COALESCE($7, ssl_mode),
                status = COALESCE($8, status),
                updated_at = NOW()
            WHERE id = $9
            """,
            _clean_optional_str(body.name),
            _clean_optional_str(body.host),
            body.port,
            _clean_optional_str(body.database),
            _clean_optional_str(body.username),
            _clean_optional_str(body.password),
            _clean_optional_str(body.ssl_mode),
            _clean_optional_str(body.status),
            connector_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Connector not found")
        row = await conn.fetchrow(
            """
            SELECT id, name, type, host, port, database, username,
                   ssl_mode, status, last_tested_at, last_test_status,
                   last_test_error, created_at, updated_at
            FROM data_source_connectors
            WHERE id = $1
            """,
            connector_id,
        )
    await _log_audit(
        event_type="data_source_updated",
        actor="system",
        source_service="core-svc",
        object_type="data_source",
        object_id=connector_id,
        details={"fields": body.model_dump(exclude_none=True)},
    )
    return DataSourceConnectorOut(**dict(row))


@router.delete("/data-sources/{connector_id}", status_code=204)
async def delete_data_source(connector_id: str) -> Response:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM data_source_connectors WHERE id = $1",
            connector_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Connector not found")
    await _log_audit(
        event_type="data_source_deleted",
        actor="system",
        source_service="core-svc",
        object_type="data_source",
        object_id=connector_id,
    )
    return Response(status_code=204)


@router.post("/data-sources/{connector_id}:test")
async def test_data_source(connector_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT host, port, database, username, password, ssl_mode
            FROM data_source_connectors
            WHERE id = $1
            """,
            connector_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Connector not found")

    ssl_mode = (row["ssl_mode"] or "").lower()
    ssl_param = None if ssl_mode in ("", "disable", "off", "false") else True

    ok = False
    error = None
    try:
        test_conn = await asyncpg.connect(
            user=row["username"],
            password=row["password"],
            host=row["host"],
            port=int(row["port"]),
            database=row["database"],
            timeout=5,
            ssl=ssl_param,
        )
        await test_conn.execute("SELECT 1")
        await test_conn.close()
        ok = True
    except Exception as exc:
        error = str(exc)

    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """
            UPDATE data_source_connectors
            SET last_tested_at = NOW(),
                last_test_status = $1,
                last_test_error = $2,
                updated_at = NOW()
            WHERE id = $3
            """,
            "success" if ok else "failure",
            error,
            connector_id,
        )

    return {"ok": ok, "error": error}


def _ssl_param(ssl_mode_value: Optional[str]) -> Optional[bool]:
    ssl_mode = (ssl_mode_value or "").lower()
    return None if ssl_mode in ("", "disable", "off", "false") else True


async def _fetch_connector_row(connector_id: str) -> asyncpg.Record:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT id, type, host, port, database, username, password, ssl_mode
            FROM data_source_connectors
            WHERE id = $1
            """,
            connector_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Connector not found")
        return row


@router.get("/data-sources/{connector_id}/schemas")
async def list_data_source_schemas(connector_id: str) -> dict:
    row = await _fetch_connector_row(connector_id)
    if (row["type"] or "").lower() != "postgres":
        raise HTTPException(status_code=400, detail="Unsupported connector type")

    try:
        conn = await asyncpg.connect(
            user=row["username"],
            password=row["password"],
            host=row["host"],
            port=int(row["port"]),
            database=row["database"],
            timeout=5,
            ssl=_ssl_param(row["ssl_mode"]),
        )
        schemas = await conn.fetch(
            """
            SELECT schema_name
            FROM information_schema.schemata
            WHERE schema_name NOT LIKE 'pg_%'
              AND schema_name <> 'information_schema'
            ORDER BY schema_name
            """
        )
        await conn.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Schema discovery failed: {exc}")

    return {"schemas": [r["schema_name"] for r in schemas]}


@router.get("/data-sources/{connector_id}/tables")
async def list_data_source_tables(
    connector_id: str,
    schema: Optional[str] = None,
) -> dict:
    row = await _fetch_connector_row(connector_id)
    if (row["type"] or "").lower() != "postgres":
        raise HTTPException(status_code=400, detail="Unsupported connector type")

    try:
        conn = await asyncpg.connect(
            user=row["username"],
            password=row["password"],
            host=row["host"],
            port=int(row["port"]),
            database=row["database"],
            timeout=5,
            ssl=_ssl_param(row["ssl_mode"]),
        )
        if schema:
            tables = await conn.fetch(
                """
                SELECT table_schema, table_name
                FROM information_schema.tables
                WHERE table_type = 'BASE TABLE'
                  AND table_schema = $1
                ORDER BY table_schema, table_name
                """,
                schema,
            )
        else:
            tables = await conn.fetch(
                """
                SELECT table_schema, table_name
                FROM information_schema.tables
                WHERE table_type = 'BASE TABLE'
                  AND table_schema NOT LIKE 'pg_%'
                  AND table_schema <> 'information_schema'
                ORDER BY table_schema, table_name
                """
            )
        await conn.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Table discovery failed: {exc}")

    return {
        "tables": [
            {"schema": r["table_schema"], "table": r["table_name"]}
            for r in tables
        ]
    }


# ----------------- Data Classification -----------------
@router.get(
    "/data-classification/tags",
    response_model=List[DataClassificationTagOut],
)
async def list_classification_tags() -> List[DataClassificationTagOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, tag_name, sensitivity_tier, pii_flag, allowed_usage,
                   retention_class, created_at, updated_at
            FROM data_classification_tags
            ORDER BY updated_at DESC, created_at DESC
            """
        )
    return [DataClassificationTagOut(**dict(r)) for r in rows]


@router.post(
    "/data-classification/tags",
    response_model=DataClassificationTagOut,
    status_code=201,
)
async def create_classification_tag(
    body: DataClassificationTagIn,
) -> DataClassificationTagOut:
    tag_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            await conn.execute(
                """
                INSERT INTO data_classification_tags (
                  id, tag_name, sensitivity_tier, pii_flag, allowed_usage,
                  retention_class, created_at, updated_at
                )
                VALUES ($1,$2,$3,$4,$5,$6,NOW(),NOW())
                """,
                tag_id,
                body.tag_name,
                body.sensitivity_tier,
                body.pii_flag,
                body.allowed_usage,
                body.retention_class,
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(status_code=409, detail="Tag name already exists")
        row = await conn.fetchrow(
            """
            SELECT id, tag_name, sensitivity_tier, pii_flag, allowed_usage,
                   retention_class, created_at, updated_at
            FROM data_classification_tags
            WHERE id = $1
            """,
            tag_id,
        )
    await _log_audit(
        event_type="data_classification_tag_created",
        actor="system",
        source_service="core-svc",
        object_type="data_classification_tag",
        object_id=tag_id,
        details={"tag_name": body.tag_name},
    )
    return DataClassificationTagOut(**dict(row))


@router.put(
    "/data-classification/tags/{tag_id}",
    response_model=DataClassificationTagOut,
)
async def update_classification_tag(
    tag_id: str,
    body: DataClassificationTagUpdate,
) -> DataClassificationTagOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE data_classification_tags
            SET tag_name = COALESCE($1, tag_name),
                sensitivity_tier = COALESCE($2, sensitivity_tier),
                pii_flag = COALESCE($3, pii_flag),
                allowed_usage = COALESCE($4, allowed_usage),
                retention_class = COALESCE($5, retention_class),
                updated_at = NOW()
            WHERE id = $6
            """,
            _clean_optional_str(body.tag_name),
            _clean_optional_str(body.sensitivity_tier),
            body.pii_flag,
            _clean_optional_str(body.allowed_usage),
            _clean_optional_str(body.retention_class),
            tag_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Tag not found")
        row = await conn.fetchrow(
            """
            SELECT id, tag_name, sensitivity_tier, pii_flag, allowed_usage,
                   retention_class, created_at, updated_at
            FROM data_classification_tags
            WHERE id = $1
            """,
            tag_id,
        )
    await _log_audit(
        event_type="data_classification_tag_updated",
        actor="system",
        source_service="core-svc",
        object_type="data_classification_tag",
        object_id=tag_id,
        details={"fields": body.model_dump(exclude_none=True)},
    )
    return DataClassificationTagOut(**dict(row))


@router.delete("/data-classification/tags/{tag_id}", status_code=204)
async def delete_classification_tag(tag_id: str) -> Response:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM data_classification_tags WHERE id = $1",
            tag_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Tag not found")
    await _log_audit(
        event_type="data_classification_tag_deleted",
        actor="system",
        source_service="core-svc",
        object_type="data_classification_tag",
        object_id=tag_id,
    )
    return Response(status_code=204)


@router.get(
    "/data-classification/assignments",
    response_model=List[DataClassificationAssignmentOut],
)
async def list_classification_assignments(
    connector_id: Optional[str] = None,
) -> List[DataClassificationAssignmentOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        if connector_id:
            rows = await conn.fetch(
                """
                SELECT a.id, a.connector_id, a.schema_name, a.table_name, a.name,
                       a.id_number, a.tag_id, a.created_at, a.updated_at,
                       t.tag_name, t.sensitivity_tier, t.pii_flag,
                       t.allowed_usage, t.retention_class
                FROM data_classification_assignments a
                JOIN data_classification_tags t ON t.id = a.tag_id
                WHERE a.connector_id = $1
                ORDER BY a.updated_at DESC, a.created_at DESC
                """,
                connector_id,
            )
        else:
            rows = await conn.fetch(
                """
                SELECT a.id, a.connector_id, a.schema_name, a.table_name, a.name,
                       a.id_number, a.tag_id, a.created_at, a.updated_at,
                       t.tag_name, t.sensitivity_tier, t.pii_flag,
                       t.allowed_usage, t.retention_class
                FROM data_classification_assignments a
                JOIN data_classification_tags t ON t.id = a.tag_id
                ORDER BY a.updated_at DESC, a.created_at DESC
                """
            )
    return [DataClassificationAssignmentOut(**dict(r)) for r in rows]


@router.post(
    "/data-classification/assignments",
    response_model=DataClassificationAssignmentOut,
    status_code=201,
)
async def create_classification_assignment(
    body: DataClassificationAssignmentIn,
) -> DataClassificationAssignmentOut:
    assignment_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO data_classification_assignments (
              id, connector_id, schema_name, table_name, name, id_number, tag_id,
              created_at, updated_at
            )
            VALUES ($1,$2,$3,$4,$5,$6,$7,NOW(),NOW())
            """,
            assignment_id,
            body.connector_id,
            _clean_optional_str(body.schema_name),
            _clean_optional_str(body.table_name),
            _clean_optional_str(body.name),
            _clean_optional_str(body.id_number),
            body.tag_id,
        )
        row = await conn.fetchrow(
            """
            SELECT a.id, a.connector_id, a.schema_name, a.table_name, a.name,
                   a.id_number, a.tag_id, a.created_at, a.updated_at,
                   t.tag_name, t.sensitivity_tier, t.pii_flag,
                   t.allowed_usage, t.retention_class
            FROM data_classification_assignments a
            JOIN data_classification_tags t ON t.id = a.tag_id
            WHERE a.id = $1
            """,
            assignment_id,
        )
    await _log_audit(
        event_type="data_classification_assignment_created",
        actor="system",
        source_service="core-svc",
        object_type="data_classification_assignment",
        object_id=assignment_id,
        details={"connector_id": body.connector_id, "tag_id": body.tag_id},
    )
    return DataClassificationAssignmentOut(**dict(row))


@router.put(
    "/data-classification/assignments/{assignment_id}",
    response_model=DataClassificationAssignmentOut,
)
async def update_classification_assignment(
    assignment_id: str,
    body: DataClassificationAssignmentUpdate,
) -> DataClassificationAssignmentOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE data_classification_assignments
            SET connector_id = COALESCE($1, connector_id),
                schema_name = COALESCE($2, schema_name),
                table_name = COALESCE($3, table_name),
                name = COALESCE($4, name),
                id_number = COALESCE($5, id_number),
                tag_id = COALESCE($6, tag_id),
                updated_at = NOW()
            WHERE id = $7
            """,
            _clean_optional_str(body.connector_id),
            _clean_optional_str(body.schema_name),
            _clean_optional_str(body.table_name),
            _clean_optional_str(body.name),
            _clean_optional_str(body.id_number),
            _clean_optional_str(body.tag_id),
            assignment_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Assignment not found")
        row = await conn.fetchrow(
            """
            SELECT a.id, a.connector_id, a.schema_name, a.table_name, a.name,
                   a.id_number, a.tag_id, a.created_at, a.updated_at,
                   t.tag_name, t.sensitivity_tier, t.pii_flag,
                   t.allowed_usage, t.retention_class
            FROM data_classification_assignments a
            JOIN data_classification_tags t ON t.id = a.tag_id
            WHERE a.id = $1
            """,
            assignment_id,
        )
    await _log_audit(
        event_type="data_classification_assignment_updated",
        actor="system",
        source_service="core-svc",
        object_type="data_classification_assignment",
        object_id=assignment_id,
        details={"fields": body.model_dump(exclude_none=True)},
    )
    return DataClassificationAssignmentOut(**dict(row))


@router.delete(
    "/data-classification/assignments/{assignment_id}",
    status_code=204,
)
async def delete_classification_assignment(assignment_id: str) -> Response:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM data_classification_assignments WHERE id = $1",
            assignment_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Assignment not found")
    await _log_audit(
        event_type="data_classification_assignment_deleted",
        actor="system",
        source_service="core-svc",
        object_type="data_classification_assignment",
        object_id=assignment_id,
    )
    return Response(status_code=204)


# ----------------- Retention, Usage, Governance Warnings -----------------
@router.get(
    "/retention/policies",
    response_model=List[DataRetentionPolicyOut],
)
async def list_retention_policies() -> List[DataRetentionPolicyOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, retention_class, archive_after_days, delete_after_days,
                   notes, created_at, updated_at
            FROM data_retention_policies
            ORDER BY updated_at DESC, created_at DESC
            """
        )
    return [DataRetentionPolicyOut(**dict(r)) for r in rows]


@router.post(
    "/retention/policies",
    response_model=DataRetentionPolicyOut,
    status_code=201,
)
async def create_retention_policy(
    body: DataRetentionPolicyIn,
) -> DataRetentionPolicyOut:
    policy_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            await conn.execute(
                """
                INSERT INTO data_retention_policies (
                  id, retention_class, archive_after_days, delete_after_days,
                  notes, created_at, updated_at
                )
                VALUES ($1,$2,$3,$4,$5,NOW(),NOW())
                """,
                policy_id,
                body.retention_class,
                body.archive_after_days,
                body.delete_after_days,
                body.notes,
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(status_code=409, detail="Retention class already exists")
        row = await conn.fetchrow(
            """
            SELECT id, retention_class, archive_after_days, delete_after_days,
                   notes, created_at, updated_at
            FROM data_retention_policies
            WHERE id = $1
            """,
            policy_id,
        )
    await _log_audit(
        event_type="retention_policy_created",
        actor="system",
        source_service="core-svc",
        object_type="retention_policy",
        object_id=policy_id,
        details={"retention_class": body.retention_class},
    )
    return DataRetentionPolicyOut(**dict(row))


@router.put(
    "/retention/policies/{policy_id}",
    response_model=DataRetentionPolicyOut,
)
async def update_retention_policy(
    policy_id: str,
    body: DataRetentionPolicyUpdate,
) -> DataRetentionPolicyOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE data_retention_policies
            SET retention_class = COALESCE($1, retention_class),
                archive_after_days = COALESCE($2, archive_after_days),
                delete_after_days = COALESCE($3, delete_after_days),
                notes = COALESCE($4, notes),
                updated_at = NOW()
            WHERE id = $5
            """,
            _clean_optional_str(body.retention_class),
            body.archive_after_days,
            body.delete_after_days,
            _clean_optional_str(body.notes),
            policy_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Retention policy not found")
        row = await conn.fetchrow(
            """
            SELECT id, retention_class, archive_after_days, delete_after_days,
                   notes, created_at, updated_at
            FROM data_retention_policies
            WHERE id = $1
            """,
            policy_id,
        )
    await _log_audit(
        event_type="retention_policy_updated",
        actor="system",
        source_service="core-svc",
        object_type="retention_policy",
        object_id=policy_id,
        details={"fields": body.model_dump(exclude_none=True)},
    )
    return DataRetentionPolicyOut(**dict(row))


@router.delete("/retention/policies/{policy_id}", status_code=204)
async def delete_retention_policy(policy_id: str) -> Response:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM data_retention_policies WHERE id = $1",
            policy_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Retention policy not found")
    await _log_audit(
        event_type="retention_policy_deleted",
        actor="system",
        source_service="core-svc",
        object_type="retention_policy",
        object_id=policy_id,
    )
    return Response(status_code=204)


@router.get(
    "/retention/records",
    response_model=List[DataRetentionRecordOut],
)
async def list_retention_records() -> List[DataRetentionRecordOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, assignment_id, retention_class, start_date, last_accessed_at,
                   status, archived_at, deleted_at, notes, created_at, updated_at
            FROM data_retention_records
            ORDER BY updated_at DESC, created_at DESC
            """
        )
    return [DataRetentionRecordOut(**dict(r)) for r in rows]


@router.post(
    "/retention/records",
    response_model=DataRetentionRecordOut,
    status_code=201,
)
async def create_retention_record(
    body: DataRetentionRecordIn,
) -> DataRetentionRecordOut:
    record_id = str(uuid4())
    status_value = _clean_optional_str(body.status) or "active"
    archived_at = datetime.now(timezone.utc) if status_value == "archived" else None
    deleted_at = datetime.now(timezone.utc) if status_value == "deleted" else None
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO data_retention_records (
              id, assignment_id, retention_class, start_date, last_accessed_at,
              status, archived_at, deleted_at, notes, created_at, updated_at
            )
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,NOW(),NOW())
            """,
            record_id,
            body.assignment_id,
            body.retention_class,
            body.start_date,
            body.last_accessed_at,
            status_value,
            archived_at,
            deleted_at,
            _clean_optional_str(body.notes),
        )
        row = await conn.fetchrow(
            """
            SELECT id, assignment_id, retention_class, start_date, last_accessed_at,
                   status, archived_at, deleted_at, notes, created_at, updated_at
            FROM data_retention_records
            WHERE id = $1
            """,
            record_id,
        )
    await _log_audit(
        event_type="retention_record_created",
        actor="system",
        source_service="core-svc",
        object_type="retention_record",
        object_id=record_id,
        details={"assignment_id": body.assignment_id},
    )
    return DataRetentionRecordOut(**dict(row))


@router.put(
    "/retention/records/{record_id}",
    response_model=DataRetentionRecordOut,
)
async def update_retention_record(
    record_id: str,
    body: DataRetentionRecordUpdate,
) -> DataRetentionRecordOut:
    status_value = _clean_optional_str(body.status)
    archived_at = datetime.now(timezone.utc) if status_value == "archived" else None
    deleted_at = datetime.now(timezone.utc) if status_value == "deleted" else None
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE data_retention_records
            SET assignment_id = COALESCE($1, assignment_id),
                retention_class = COALESCE($2, retention_class),
                start_date = COALESCE($3, start_date),
                last_accessed_at = COALESCE($4, last_accessed_at),
                status = COALESCE($5, status),
                archived_at = COALESCE($6, archived_at),
                deleted_at = COALESCE($7, deleted_at),
                notes = COALESCE($8, notes),
                updated_at = NOW()
            WHERE id = $9
            """,
            _clean_optional_str(body.assignment_id),
            _clean_optional_str(body.retention_class),
            body.start_date,
            body.last_accessed_at,
            status_value,
            archived_at,
            deleted_at,
            _clean_optional_str(body.notes),
            record_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Retention record not found")
        row = await conn.fetchrow(
            """
            SELECT id, assignment_id, retention_class, start_date, last_accessed_at,
                   status, archived_at, deleted_at, notes, created_at, updated_at
            FROM data_retention_records
            WHERE id = $1
            """,
            record_id,
        )
    await _log_audit(
        event_type="retention_record_updated",
        actor="system",
        source_service="core-svc",
        object_type="retention_record",
        object_id=record_id,
        details={"fields": body.model_dump(exclude_none=True)},
    )
    return DataRetentionRecordOut(**dict(row))


@router.delete("/retention/records/{record_id}", status_code=204)
async def delete_retention_record(record_id: str) -> Response:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM data_retention_records WHERE id = $1",
            record_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Retention record not found")
    await _log_audit(
        event_type="retention_record_deleted",
        actor="system",
        source_service="core-svc",
        object_type="retention_record",
        object_id=record_id,
    )
    return Response(status_code=204)


@router.get(
    "/data-usage",
    response_model=List[DataUsageRecordOut],
)
async def list_data_usage_records(
    assignment_id: Optional[str] = None,
) -> List[DataUsageRecordOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        if assignment_id:
            rows = await conn.fetch(
                """
                SELECT id, assignment_id, usage_type, purpose, recorded_at
                FROM data_usage_records
                WHERE assignment_id = $1
                ORDER BY recorded_at DESC
                """,
                assignment_id,
            )
        else:
            rows = await conn.fetch(
                """
                SELECT id, assignment_id, usage_type, purpose, recorded_at
                FROM data_usage_records
                ORDER BY recorded_at DESC
                """
            )
    return [DataUsageRecordOut(**dict(r)) for r in rows]


@router.post(
    "/data-usage",
    response_model=DataUsageRecordOut,
    status_code=201,
)
async def create_data_usage_record(
    body: DataUsageRecordIn,
) -> DataUsageRecordOut:
    usage_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO data_usage_records (
              id, assignment_id, usage_type, purpose, recorded_at
            )
            VALUES ($1,$2,$3,$4,NOW())
            """,
            usage_id,
            body.assignment_id,
            body.usage_type,
            _clean_optional_str(body.purpose),
        )
        row = await conn.fetchrow(
            """
            SELECT id, assignment_id, usage_type, purpose, recorded_at
            FROM data_usage_records
            WHERE id = $1
            """,
            usage_id,
        )
    await _log_audit(
        event_type="data_usage_recorded",
        actor="system",
        source_service="core-svc",
        object_type="data_usage",
        object_id=usage_id,
        details={"assignment_id": body.assignment_id, "usage_type": body.usage_type},
    )
    return DataUsageRecordOut(**dict(row))


@router.post("/data-governance/warnings:compute")
async def compute_data_governance_warnings() -> dict:
    rows = await compute_governance_warnings()
    return {"warnings": rows}


@router.get(
    "/data-governance/warnings",
    response_model=List[DataGovernanceWarningOut],
)
async def list_data_governance_warnings() -> List[DataGovernanceWarningOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, assignment_id, warning_type, severity, message,
                   created_at, resolved_at
            FROM data_governance_warnings
            ORDER BY created_at DESC
            """
        )
    return [DataGovernanceWarningOut(**dict(r)) for r in rows]


@router.post("/data-governance/warnings/{warning_id}/resolve")
async def resolve_data_governance_warning(warning_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE data_governance_warnings
            SET resolved_at = NOW()
            WHERE id = $1
            """,
            warning_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Warning not found")
    return {"ok": True, "id": warning_id}


@router.get("/policy-alerts")
async def list_policy_alerts(
    limit: int = 100,
    offset: int = 0,
    project_slug: Optional[str] = None,
    status: Optional[str] = None,
    include_global: bool = True,
) -> dict:
    limit = max(1, min(int(limit), 200))
    offset = max(0, int(offset))
    clauses = []
    params = []
    if status:
        params.append(status)
        clauses.append(f"status = ${len(params)}")
    if project_slug:
        params.append(project_slug)
        if include_global:
            clauses.append(
                f"(project_slug = ${len(params)} OR project_slug = 'global' OR project_slug IS NULL)"
            )
        else:
            clauses.append(f"project_slug = ${len(params)}")
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""

    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """
            UPDATE policy_alerts
            SET status = 'resolved',
                resolved_at = NOW(),
                updated_at = NOW()
            WHERE alert_type = 'new_project_requirements'
              AND status = 'open'
              AND project_slug IS NOT NULL
              AND EXISTS (
                  SELECT 1
                  FROM ai_requirement_register r
                  WHERE r.project_slug = policy_alerts.project_slug
              )
            """
        )
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM policy_alerts {where_sql}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT id, policy_id, policy_title, project_slug, alert_type,
                   severity, message, source_type, source_key, status,
                   created_at, updated_at, resolved_at, details_json
            FROM policy_alerts
            {where_sql}
            ORDER BY created_at DESC
            LIMIT ${len(params)+1} OFFSET ${len(params)+2}
            """,
            *params,
            limit,
            offset,
        )

    items = []
    for r in rows:
        details_json = r["details_json"]
        if isinstance(details_json, str):
            try:
                details_json = json.loads(details_json)
            except json.JSONDecodeError:
                details_json = {"raw": details_json}
        items.append(
            PolicyAlertOut(
                id=r["id"],
                policy_id=r["policy_id"],
                policy_title=r["policy_title"],
                project_slug=r["project_slug"],
                alert_type=r["alert_type"],
                severity=r["severity"],
                message=r["message"],
                source_type=r["source_type"],
                source_key=r["source_key"],
                status=r["status"],
                created_at=r["created_at"],
                updated_at=r["updated_at"],
                resolved_at=r["resolved_at"],
                details_json=details_json,
            )
        )
    return {"items": items, "total": int(total or 0)}


@router.post("/policy-alerts:compute")
async def compute_policy_alerts_now() -> dict:
    rows = await compute_policy_alerts()
    return {"count": len(rows), "items": rows}


# ----------------- Knowledge Base -----------------
@router.get("/knowledgebase/kpis")
async def list_knowledgebase_kpis() -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT kpi_key, kpi_name, description, iso_42001_clause, euaiact_clause, nist_clause
            FROM kpi_definition
            ORDER BY kpi_name ASC NULLS LAST, kpi_key
            """
        )
    return {"items": [dict(r) for r in rows]}


@router.get("/knowledgebase/kpis/{kpi_key}")
async def get_knowledgebase_kpi(kpi_key: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT kpi_key, kpi_name, description, definition, example
            FROM kpi_definition
            WHERE kpi_key = $1
            """,
            kpi_key,
        )
    if not row:
        raise HTTPException(status_code=404, detail="KPI not found")
    return dict(row)


class KpiKnowledgeUpdate(BaseModel):
    nist_clause: Optional[str] = None


@router.put("/knowledgebase/kpis/{kpi_key}")
async def update_knowledgebase_kpi(kpi_key: str, body: KpiKnowledgeUpdate) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE kpi_definition
            SET nist_clause = $1
            WHERE kpi_key = $2
            """,
            body.nist_clause,
            kpi_key,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="KPI not found")
    return {"ok": True, "kpi_key": kpi_key, "nist_clause": body.nist_clause}


# ----------------- LLM Prompt Templates -----------------
class PromptVersionIn(BaseModel):
    prompt_text: str
    language: Optional[str] = "en"
    variables: Optional[dict] = None
    created_by: Optional[str] = None
    set_active: Optional[bool] = True


class KnowledgeVaultSourceIn(BaseModel):
    title: str
    source_type: str
    content: Optional[str] = None
    project_slug: Optional[str] = None
    entity_id: Optional[UUID] = None
    metadata: Optional[dict] = None
    object_key: Optional[str] = None
    file_name: Optional[str] = None
    file_mime: Optional[str] = None
    file_size: Optional[int] = None
    created_by: Optional[str] = None


class KnowledgeVaultSourceUpdate(BaseModel):
    title: Optional[str] = None
    source_type: Optional[str] = None
    content: Optional[str] = None
    project_slug: Optional[str] = None
    metadata: Optional[dict] = None
    object_key: Optional[str] = None
    file_name: Optional[str] = None
    file_mime: Optional[str] = None
    file_size: Optional[int] = None


class KnowledgeVaultPresignIn(BaseModel):
    file_name: str
    content_type: Optional[str] = None
    project_slug: Optional[str] = None
    entity_id: Optional[UUID] = None


class KnowledgeTableToggleIn(BaseModel):
    enabled: bool = True


class ControlValuesSyncIn(BaseModel):
    kpi_keys: list[str]


class ControlValuesExecRowIn(BaseModel):
    control_id: str
    kpi_key: Optional[str] = None
    owner_role: Optional[str] = None
    designated_owner_name: Optional[str] = None
    designated_owner_email: Optional[str] = None
    due_date: Optional[date] = None
    frequency: Optional[int] = None
    reminder_day: Optional[int] = None
    reminder_count: Optional[int] = None
    designated_owner_manager: Optional[str] = None
    designated_owner_manager_email: Optional[str] = None
    provide_url: Optional[str] = None
    forward_request: Optional[bool] = None
    forward_email: Optional[str] = None
    comment_text: Optional[str] = Field(default=None, max_length=400)


class ControlValuesExecUpdateIn(BaseModel):
    items: list[ControlValuesExecRowIn]


@router.get("/llm-prompts/{key}")
async def get_llm_prompt_by_key(key: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        template = await conn.fetchrow(
            """
            SELECT id, key, name, description, active_version_id, is_active, created_at, updated_at
            FROM llm_prompt_templates
            WHERE key = $1
            """,
            key,
        )
        if not template:
            raise HTTPException(status_code=404, detail="Prompt template not found")
        active_version = None
        if template["active_version_id"]:
            active_version = await conn.fetchrow(
                """
                SELECT id, template_id, version, language, prompt_text, variables, created_at, created_by
                FROM llm_prompt_versions
                WHERE id = $1
                """,
                template["active_version_id"],
            )
    return {
        "template": dict(template),
        "active_version": dict(active_version) if active_version else None,
    }


@router.get("/llm-prompts/{key}/versions")
async def list_llm_prompt_versions(key: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        template = await conn.fetchrow(
            "SELECT id FROM llm_prompt_templates WHERE key = $1",
            key,
        )
        if not template:
            raise HTTPException(status_code=404, detail="Prompt template not found")
        rows = await conn.fetch(
            """
            SELECT id, template_id, version, language, prompt_text, variables, created_at, created_by
            FROM llm_prompt_versions
            WHERE template_id = $1
            ORDER BY version DESC
            """,
            template["id"],
        )
    return {"items": [dict(r) for r in rows]}


@router.post("/llm-prompts/{key}/versions")
async def create_llm_prompt_version(key: str, body: PromptVersionIn) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        template = await conn.fetchrow(
            "SELECT id FROM llm_prompt_templates WHERE key = $1",
            key,
        )
        if not template:
            raise HTTPException(status_code=404, detail="Prompt template not found")
        next_version = await conn.fetchval(
            """
            SELECT COALESCE(MAX(version), 0) + 1
            FROM llm_prompt_versions
            WHERE template_id = $1
            """,
            template["id"],
        )
        version_id = str(uuid4())
        await conn.execute(
            """
            INSERT INTO llm_prompt_versions (id, template_id, version, language, prompt_text, variables, created_by)
            VALUES ($1,$2,$3,$4,$5,$6,$7)
            """,
            version_id,
            template["id"],
            int(next_version),
            body.language or "en",
            body.prompt_text,
            _as_json(body.variables) if body.variables is not None else None,
            body.created_by,
        )
        if body.set_active:
            await conn.execute(
                """
                UPDATE llm_prompt_templates
                SET active_version_id = $1, updated_at = NOW()
                WHERE id = $2
                """,
                version_id,
                template["id"],
            )
    return {"ok": True, "id": version_id, "version": int(next_version)}


@router.put("/llm-prompts/{key}/active/{version_id}")
async def set_llm_prompt_active(key: str, version_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        template = await conn.fetchrow(
            "SELECT id FROM llm_prompt_templates WHERE key = $1",
            key,
        )
        if not template:
            raise HTTPException(status_code=404, detail="Prompt template not found")
        version = await conn.fetchrow(
            """
            SELECT id FROM llm_prompt_versions
            WHERE id = $1 AND template_id = $2
            """,
            version_id,
            template["id"],
        )
        if not version:
            raise HTTPException(status_code=404, detail="Prompt version not found")
        await conn.execute(
            """
            UPDATE llm_prompt_templates
            SET active_version_id = $1, updated_at = NOW()
            WHERE id = $2
            """,
            version_id,
            template["id"],
        )
    return {"ok": True, "active_version_id": version_id}


@router.get("/knowledgebase/euaiact")
async def list_knowledgebase_euaiact() -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT r.chapter,
                   r.section,
                   r.article,
                   r.coverage,
                   s.primary_role,
                   s.risk_classification,
                   s.condition
            FROM euaiact_requirements r
            LEFT JOIN euaiact_requirement_scope s
              ON s.article = r.article AND s.coverage = r.coverage
            ORDER BY r.article, r.coverage, s.primary_role, s.risk_classification
            """
        )
    return {"items": [dict(r) for r in rows]}


@router.get("/knowledgebase/iso42001")
async def list_knowledgebase_iso42001() -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT chapter, section, article, content, coverage, maturity_level
            FROM iso42001
            ORDER BY chapter NULLS LAST, section NULLS LAST, article
            """
        )
    return {"items": [dict(r) for r in rows]}


# ----------------- Knowledge Vault -----------------
async def _resolve_entity_id_for_source(
    conn: asyncpg.Connection,
    entity_id: Optional[UUID],
    project_slug: Optional[str],
) -> Optional[UUID]:
    if entity_id:
        return entity_id
    if not project_slug:
        return None
    row = await conn.fetchrow(
        "SELECT entity_id FROM entity_projects WHERE slug = $1", project_slug
    )
    return row["entity_id"] if row else None


def _sanitize_filename(name: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", name.strip())
    return safe or "upload.bin"


@router.get("/knowledge-vault/sources")
async def list_knowledge_vault_sources(
    project_slug: Optional[str] = None,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        params = []
        where_parts = []
        if entity_id:
            params.append(entity_id)
            where_parts.append(f"entity_id = ${len(params)}")
        if project_slug:
            params.append(project_slug)
            where_parts.append(
                f"(project_slug IS NULL OR project_slug = ${len(params)})"
            )
        where = " AND ".join(where_parts) if where_parts else "TRUE"
        rows = await conn.fetch(
            f"""
            SELECT id, entity_id, project_slug, title, source_type, content, metadata,
                   object_key, file_name, file_mime, file_size,
                   created_by, created_at, updated_at
            FROM report_sources
            WHERE {where}
            ORDER BY updated_at DESC NULLS LAST, created_at DESC
            """,
            *params,
        )
    return {"items": [dict(r) for r in rows]}


@router.post("/knowledge-vault/presign")
async def presign_knowledge_vault_upload(
    body: KnowledgeVaultPresignIn,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        resolved_entity_id = await _resolve_entity_id_for_source(
            conn, body.entity_id or entity_id, body.project_slug
        )
        if not resolved_entity_id:
            raise HTTPException(status_code=400, detail="Entity ID is required")
    safe_name = _sanitize_filename(body.file_name)
    project_part = body.project_slug or "global"
    object_key = f"knowledge-vault/{resolved_entity_id}/{project_part}/{uuid4()}_{safe_name}"
    content_type = body.content_type or "application/octet-stream"
    upload_url, headers = presign_put(object_key, content_type)
    return {
        "upload_url": upload_url,
        "headers": headers,
        "object_key": object_key,
    }

@router.post("/knowledge-vault/sources")
async def create_knowledge_vault_source(
    body: KnowledgeVaultSourceIn,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        resolved_entity_id = await _resolve_entity_id_for_source(
            conn, body.entity_id or entity_id, body.project_slug
        )
        if not resolved_entity_id:
            raise HTTPException(status_code=400, detail="Entity ID is required")
        source_id = str(uuid4())
        await conn.execute(
            """
            INSERT INTO report_sources (
              id, entity_id, project_slug, title, source_type, content, metadata,
              object_key, file_name, file_mime, file_size,
              created_by, created_at, updated_at
            )
            VALUES ($1,$2,$3,$4,$5,$6,$7::jsonb,$8,$9,$10,$11,$12,NOW(),NOW())
            """,
            source_id,
            resolved_entity_id,
            body.project_slug,
            body.title,
            body.source_type,
            body.content,
            _as_json(body.metadata),
            body.object_key,
            body.file_name,
            body.file_mime,
            body.file_size,
            body.created_by,
        )
    return {"ok": True, "id": source_id}


@router.put("/knowledge-vault/sources/{source_id}")
async def update_knowledge_vault_source(
    source_id: str, body: KnowledgeVaultSourceUpdate
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE report_sources
            SET title = COALESCE($1, title),
                source_type = COALESCE($2, source_type),
                content = COALESCE($3, content),
                project_slug = COALESCE($4, project_slug),
                metadata = COALESCE($5::jsonb, metadata),
                object_key = COALESCE($6, object_key),
                file_name = COALESCE($7, file_name),
                file_mime = COALESCE($8, file_mime),
                file_size = COALESCE($9, file_size),
                updated_at = NOW()
            WHERE id = $10
            """,
            body.title,
            body.source_type,
            body.content,
            body.project_slug,
            _as_json(body.metadata),
            body.object_key,
            body.file_name,
            body.file_mime,
            body.file_size,
            source_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Source not found")
    return {"ok": True, "id": source_id}


@router.delete("/knowledge-vault/sources/{source_id}")
async def delete_knowledge_vault_source(source_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "DELETE FROM report_sources WHERE id = $1", source_id
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Source not found")
    return {"ok": True, "id": source_id}


@router.get("/knowledge-vault/sources/{source_id}/download")
async def get_knowledge_vault_download_url(
    source_id: str,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT object_key, file_name, entity_id
            FROM report_sources
            WHERE id = $1
            """,
            source_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Source not found")
        if entity_id and row["entity_id"] != entity_id:
            raise HTTPException(status_code=403, detail="Forbidden")
        if not row["object_key"]:
            raise HTTPException(status_code=400, detail="No file available")
    return {
        "url": presign_get(row["object_key"]),
        "file_name": row["file_name"],
    }


@router.post("/knowledge-vault/sources/{source_id}/ingest")
async def ingest_knowledge_vault_source(
    source_id: str,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT entity_id FROM report_sources WHERE id = $1",
            source_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Source not found")
        if entity_id and row["entity_id"] != entity_id:
            raise HTTPException(status_code=403, detail="Forbidden")
    return await ingest_report_source(source_id)


@router.get("/knowledge-vault/search")
async def search_knowledge_vault_sources(
    query: str,
    limit: int = 6,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    return await search_knowledge_vault(
        entity_id=str(entity_id) if entity_id else None,
        query=query,
        limit=limit,
    )


@router.get("/knowledge-vault/tables")
async def list_knowledge_vault_tables(
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    if not entity_id:
        raise HTTPException(status_code=400, detail="Entity ID is required")
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT r.id,
                   r.table_key,
                   r.label,
                   r.description,
                   r.is_active,
                   COALESCE(s.enabled, false) AS enabled
            FROM knowledge_table_registry r
            LEFT JOIN entity_knowledge_table_settings s
              ON s.table_id = r.id AND s.entity_id = $1
            ORDER BY r.label
            """,
            entity_id,
        )
    return {"items": [dict(r) for r in rows]}


@router.put("/knowledge-vault/tables/{table_key}")
async def update_knowledge_vault_table(
    table_key: str,
    body: KnowledgeTableToggleIn,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    if not entity_id:
        raise HTTPException(status_code=400, detail="Entity ID is required")
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id FROM knowledge_table_registry WHERE table_key = $1",
            table_key,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Table not found")
        table_id = row["id"]
        await conn.execute(
            """
            INSERT INTO entity_knowledge_table_settings (
              id, entity_id, table_id, enabled, created_at, updated_at
            )
            VALUES (gen_random_uuid(), $1, $2, $3, NOW(), NOW())
            ON CONFLICT (entity_id, table_id)
            DO UPDATE SET enabled = EXCLUDED.enabled, updated_at = NOW()
            """,
            entity_id,
            table_id,
            body.enabled,
        )
    return {"ok": True, "table_key": table_key, "enabled": body.enabled}


@router.get("/knowledge-vault/tables/{table_key}/rows")
async def list_knowledge_vault_table_rows(
    table_key: str,
    limit: int = 50,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    if not entity_id:
        raise HTTPException(status_code=400, detail="Entity ID is required")
    safe_limit = max(1, min(limit, 200))
    table_queries = {
        "kpi_definition": """
            SELECT kpi_key,
                   kpi_name,
                   description,
                   iso_42001_clause,
                   euaiact_clause,
                   nist_clause
            FROM kpi_definition
            ORDER BY kpi_key
            LIMIT $1
        """,
        "nistairmf": """
            SELECT function,
                   category,
                   subcategory,
                   statement
            FROM nistairmf
            ORDER BY function, category, subcategory
            LIMIT $1
        """,
        "iso42001": """
            SELECT chapter,
                   section,
                   article,
                   content,
                   coverage,
                   maturity_level
            FROM iso42001
            ORDER BY chapter NULLS LAST, section NULLS LAST, article
            LIMIT $1
        """,
        "euaiact_requirements": """
            SELECT chapter,
                   section,
                   article,
                   coverage,
                   content
            FROM euaiact_requirements
            ORDER BY article, coverage
            LIMIT $1
        """,
        "euaiact_requirement_scope": """
            SELECT primary_role,
                   risk_classification,
                   article,
                   coverage,
                   condition
            FROM euaiact_requirement_scope
            ORDER BY primary_role, risk_classification, article, coverage
            LIMIT $1
        """,
    }
    sql = table_queries.get(table_key)
    if not sql:
        raise HTTPException(status_code=404, detail="Table not supported")
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT r.id,
                   r.table_key,
                   r.label,
                   r.description,
                   r.is_active,
                   COALESCE(s.enabled, false) AS enabled
            FROM knowledge_table_registry r
            LEFT JOIN entity_knowledge_table_settings s
              ON s.table_id = r.id AND s.entity_id = $1
            WHERE r.table_key = $2
            """,
            entity_id,
            table_key,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Table not found")
        if not row["is_active"]:
            raise HTTPException(status_code=404, detail="Table is inactive")
        if not row["enabled"]:
            raise HTTPException(status_code=403, detail="Table disabled for entity")
        rows = await conn.fetch(sql, safe_limit)
    return {
        "table_key": table_key,
        "label": row["label"],
        "description": row["description"],
        "columns": list(rows[0].keys()) if rows else [],
        "rows": [dict(r) for r in rows],
        "limit": safe_limit,
    }


# ----------------- Requirement Register -----------------
@router.get("/requirements")
async def list_requirements(
    limit: int = 50,
    offset: int = 0,
    q: Optional[str] = None,
    project_slug: Optional[str] = None,
    uc_id: Optional[str] = None,
    framework: Optional[str] = None,
    requirement_code: Optional[str] = None,
    status: Optional[str] = None,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    where = []
    params = []
    if entity_id:
        params.append(entity_id)
        where.append(
            f"(entity_id = ${len(params)} OR project_slug IN (SELECT slug FROM entity_projects WHERE entity_id = ${len(params)}))"
        )
    if project_slug:
        params.append(project_slug)
        where.append(f"project_slug = ${len(params)}")
    if uc_id:
        params.append(uc_id)
        where.append(f"uc_id = ${len(params)}")
    if framework:
        params.append(framework)
        where.append(f"framework = ${len(params)}")
    if requirement_code:
        params.append(requirement_code)
        where.append(f"requirement_code = ${len(params)}")
    if status:
        params.append(status)
        where.append(f"status = ${len(params)}")
    if q:
        params.append(f"%{q}%")
        where.append(
            f"(requirement_code ILIKE ${len(params)} OR title ILIKE ${len(params)} OR description ILIKE ${len(params)})"
        )

    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    pool = await get_pool()
    async with pool.acquire() as conn:
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM ai_requirement_register {where_sql}",
            *params,
        )
        rows = await conn.fetch(
            f"""
            SELECT id, project_slug, uc_id, framework, requirement_code, title,
                   description, applicability, owner_role, status, evidence_ids,
                   mapped_controls, notes, created_at, updated_at
            FROM ai_requirement_register
            {where_sql}
            ORDER BY updated_at DESC NULLS LAST, created_at DESC
            LIMIT ${len(params)+1} OFFSET ${len(params)+2}
            """,
            *params,
            limit,
            offset,
        )

    items = []
    for r in rows:
        items.append(
            {
                "id": r["id"],
                "project_slug": r["project_slug"],
                "uc_id": r["uc_id"],
                "framework": r["framework"],
                "requirement_code": r["requirement_code"],
                "title": r["title"],
                "description": r["description"],
                "applicability": r["applicability"],
                "owner_role": r["owner_role"],
                "status": r["status"],
                "evidence_ids": r["evidence_ids"],
                "mapped_controls": r["mapped_controls"],
                "notes": r["notes"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
            }
        )
    return {"items": items, "total": int(total or 0)}


@router.get("/requirements/coverage-options")
async def list_requirement_coverage_options(
    framework: Optional[str] = None,
) -> dict:
    if framework and framework != "eu_ai_act":
        return {"items": []}
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT DISTINCT coverage
            FROM euaiact_requirements
            WHERE coverage IS NOT NULL AND coverage <> ''
            ORDER BY coverage
            """
        )
    return {"items": [r["coverage"] for r in rows]}


@router.get("/requirements/project-kpis")
async def list_requirement_project_kpis(
    project_slug: str,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    def normalize_role_risk(label: str | None) -> tuple[Optional[str], Optional[str]]:
        if not label:
            return None, None
        text = label.lower()
        role = None
        for candidate in ("provider", "deployer", "importer"):
            if candidate in text:
                role = candidate
                break
        risk = None
        if "non" in text:
            risk = "non-high"
        elif "high" in text:
            risk = "high"
        return role, risk

    pool = await get_pool()
    async with pool.acquire() as conn:
        params = [project_slug]
        where = "project_slug = $1"
        if entity_id:
            params.append(entity_id)
            where += f" AND entity_id = ${len(params)}"
        req_rows = await conn.fetch(
            f"""
            SELECT framework, requirement_code
            FROM ai_requirement_register
            WHERE {where}
            ORDER BY framework, requirement_code
            """,
            *params,
        )

        items: list[dict] = []
        summary: list[dict] = []
        seen = set()

        for req in req_rows:
            framework = req["framework"]
            requirement_code = req["requirement_code"]

            if framework == "eu_ai_act":
                role, risk = normalize_role_risk(requirement_code)
                if not role or not risk:
                    summary.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "count": 0,
                            "note": "Could not infer role/risk from requirement code",
                        }
                    )
                    continue
                scope_rows = await conn.fetch(
                    """
                    SELECT article, coverage
                    FROM euaiact_requirement_scope
                    WHERE primary_role = $1 AND risk_classification = $2
                    """,
                    role,
                    risk,
                )
                article_nums: list[str] = []
                for row in scope_rows:
                    match = re.search(r"(\d+)", row["article"] or "")
                    if match:
                        article_nums.append(match.group(1))
                patterns = [f"%Art {num}%" for num in sorted(set(article_nums))]
                if not patterns:
                    summary.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "count": 0,
                            "note": "No mapped EU AI Act clauses found for this scope",
                        }
                    )
                    continue
                kpi_rows = await conn.fetch(
                    """
                    SELECT kd.kpi_key,
                           COALESCE(k.name, kd.kpi_name) AS kpi_name,
                           kd.euaiact_clause AS clause
                    FROM kpi_definition kd
                    LEFT JOIN kpis k ON k.key = kd.kpi_key
                    WHERE kd.euaiact_clause IS NOT NULL
                      AND kd.euaiact_clause ILIKE ANY($1::text[])
                    ORDER BY kd.kpi_name ASC NULLS LAST, kd.kpi_key
                    """,
                    patterns,
                )
                count = 0
                for row in kpi_rows:
                    key = (framework, requirement_code, row["kpi_key"])
                    if key in seen:
                        continue
                    seen.add(key)
                    count += 1
                    items.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "kpi_key": row["kpi_key"],
                            "kpi_name": row["kpi_name"],
                            "clause": row["clause"],
                        }
                    )
                summary.append(
                    {
                        "framework": framework,
                        "requirement_code": requirement_code,
                        "count": count,
                        "note": None,
                    }
                )
                continue

            if framework == "iso_42001":
                kpi_rows = await conn.fetch(
                    """
                    SELECT kd.kpi_key,
                           COALESCE(k.name, kd.kpi_name) AS kpi_name,
                           kd.iso_42001_clause AS clause
                    FROM kpi_definition kd
                    LEFT JOIN kpis k ON k.key = kd.kpi_key
                    WHERE kd.iso_42001_clause IS NOT NULL
                    ORDER BY kd.kpi_name ASC NULLS LAST, kd.kpi_key
                    """
                )
                count = 0
                for row in kpi_rows:
                    key = (framework, requirement_code, row["kpi_key"])
                    if key in seen:
                        continue
                    seen.add(key)
                    count += 1
                    items.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "kpi_key": row["kpi_key"],
                            "kpi_name": row["kpi_name"],
                            "clause": row["clause"],
                        }
                    )
                summary.append(
                    {
                        "framework": framework,
                        "requirement_code": requirement_code,
                        "count": count,
                        "note": None,
                    }
                )
                continue

            if framework == "nist_ai_rmf":
                if not requirement_code:
                    summary.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "count": 0,
                            "note": "No NIST AI RMF function provided",
                        }
                    )
                    continue
                scope_rows = await conn.fetch(
                    """
                    SELECT subcategory
                    FROM nistairmf
                    WHERE LOWER(function) = LOWER($1)
                    """,
                    requirement_code,
                )
                patterns = [f"%{row['subcategory']}%" for row in scope_rows]
                if not patterns:
                    summary.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "count": 0,
                            "note": "No mapped NIST AI RMF clauses found for this function",
                        }
                    )
                    continue
                kpi_rows = await conn.fetch(
                    """
                    SELECT kd.kpi_key,
                           COALESCE(k.name, kd.kpi_name) AS kpi_name,
                           kd.nist_clause AS clause
                    FROM kpi_definition kd
                    LEFT JOIN kpis k ON k.key = kd.kpi_key
                    WHERE kd.nist_clause IS NOT NULL
                      AND kd.nist_clause ILIKE ANY($1::text[])
                    ORDER BY kd.kpi_name ASC NULLS LAST, kd.kpi_key
                    """,
                    patterns,
                )
                count = 0
                for row in kpi_rows:
                    key = (framework, requirement_code, row["kpi_key"])
                    if key in seen:
                        continue
                    seen.add(key)
                    count += 1
                    items.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "kpi_key": row["kpi_key"],
                            "kpi_name": row["kpi_name"],
                            "clause": row["clause"],
                        }
                    )
                summary.append(
                    {
                        "framework": framework,
                        "requirement_code": requirement_code,
                        "count": count,
                        "note": None,
                    }
                )
                continue

            if framework == "company_specific":
                if not requirement_code:
                    summary.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "count": 0,
                            "note": "No KPI key provided for INTERNAL mapping",
                        }
                    )
                    continue
                row = await conn.fetchrow(
                    """
                    SELECT k.key AS kpi_key,
                           COALESCE(k.name, kd.kpi_name) AS kpi_name
                    FROM kpis k
                    LEFT JOIN kpi_definition kd ON kd.kpi_key = k.key
                    WHERE k.key = $1
                    """,
                    requirement_code,
                )
                if not row:
                    summary.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "count": 0,
                            "note": "No KPI found for provided key",
                        }
                    )
                    continue
                key = (framework, requirement_code, row["kpi_key"])
                if key not in seen:
                    seen.add(key)
                    items.append(
                        {
                            "framework": framework,
                            "requirement_code": requirement_code,
                            "kpi_key": row["kpi_key"],
                            "kpi_name": row["kpi_name"],
                            "clause": requirement_code,
                        }
                    )
                summary.append(
                    {
                        "framework": framework,
                        "requirement_code": requirement_code,
                        "count": 1,
                        "note": None,
                    }
                )
                continue

            summary.append(
                {
                    "framework": framework,
                    "requirement_code": requirement_code,
                    "count": 0,
                    "note": "No KPI mapping defined yet for this framework",
                }
            )

    return {"project_slug": project_slug, "items": items, "summary": summary}


@router.post("/requirements")
async def create_requirement(body: RequirementIn) -> dict:
    req_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        entity_id = None
        entity_slug = None
        if body.project_slug:
            entity_row = await conn.fetchrow(
                """
                SELECT entity_id, entity_slug
                FROM entity_projects
                WHERE slug = $1
                """,
                body.project_slug,
            )
            if entity_row:
                entity_id = entity_row["entity_id"]
                entity_slug = entity_row["entity_slug"]
        await conn.execute(
            """
            INSERT INTO ai_requirement_register (
              id, project_slug, uc_id, framework, requirement_code, title,
              description, applicability, owner_role, status, evidence_ids,
              mapped_controls, notes, entity_id, entity_slug, created_at, updated_at
            )
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,COALESCE($10,'not_started'),$11::jsonb,$12::jsonb,$13,$14,$15,NOW(),NOW())
            """,
            req_id,
            body.project_slug,
            body.uc_id,
            body.framework,
            body.requirement_code,
            body.title,
            body.description,
            body.applicability,
            body.owner_role,
            body.status,
            _as_json(body.evidence_ids),
            _as_json(body.mapped_controls),
            body.notes,
            entity_id,
            entity_slug,
        )
    await _log_audit(
        event_type="requirement_created",
        actor="system",
        source_service="core-svc",
        object_type="requirement",
        object_id=req_id,
        project_slug=body.project_slug,
        details={"framework": body.framework, "code": body.requirement_code},
    )
    return {"ok": True, "id": req_id}


@router.put("/requirements/{requirement_id}")
async def update_requirement(requirement_id: str, body: RequirementUpdate) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        entity_id = None
        entity_slug = None
        if body.project_slug:
            entity_row = await conn.fetchrow(
                """
                SELECT entity_id, entity_slug
                FROM entity_projects
                WHERE slug = $1
                """,
                body.project_slug,
            )
            if entity_row:
                entity_id = entity_row["entity_id"]
                entity_slug = entity_row["entity_slug"]
        res = await conn.execute(
            """
            UPDATE ai_requirement_register
            SET project_slug = COALESCE($1, project_slug),
                uc_id = COALESCE($2, uc_id),
                framework = COALESCE($3, framework),
                requirement_code = COALESCE($4, requirement_code),
                title = COALESCE($5, title),
                description = COALESCE($6, description),
                applicability = COALESCE($7, applicability),
                owner_role = COALESCE($8, owner_role),
                status = COALESCE($9, status),
                evidence_ids = COALESCE($10::jsonb, evidence_ids),
                mapped_controls = COALESCE($11::jsonb, mapped_controls),
                notes = COALESCE($12, notes),
                entity_id = COALESCE($13, entity_id),
                entity_slug = COALESCE($14, entity_slug),
                updated_at = NOW()
            WHERE id = $15
            """,
            body.project_slug,
            body.uc_id,
            body.framework,
            body.requirement_code,
            body.title,
            body.description,
            body.applicability,
            body.owner_role,
            body.status,
            _as_json(body.evidence_ids),
            _as_json(body.mapped_controls),
            body.notes,
            entity_id,
            entity_slug,
            requirement_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Requirement not found")
    await _log_audit(
        event_type="requirement_updated",
        actor="system",
        source_service="core-svc",
        object_type="requirement",
        object_id=requirement_id,
        project_slug=body.project_slug,
        details={"fields": body.model_dump(exclude_none=True)},
    )
    return {"ok": True, "id": requirement_id}


@router.delete("/requirements/{requirement_id}")
async def delete_requirement(requirement_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        req_row = await conn.fetchrow(
            """
            SELECT id, entity_id, project_slug, framework, requirement_code
            FROM ai_requirement_register
            WHERE id = $1
            """,
            requirement_id,
        )
        if not req_row:
            raise HTTPException(status_code=404, detail="Requirement not found")

        entity_id = req_row["entity_id"]
        project_slug = req_row["project_slug"]
        framework = req_row["framework"]
        requirement_code = req_row["requirement_code"]

        def normalize_role_risk(label: str | None) -> tuple[Optional[str], Optional[str]]:
            if not label:
                return None, None
            text = label.lower()
            role = None
            for candidate in ("provider", "deployer", "importer"):
                if candidate in text:
                    role = candidate
                    break
            risk = None
            if "non" in text:
                risk = "non-high"
            elif "high" in text:
                risk = "high"
            return role, risk

        kpi_keys: list[str] = []
        if framework == "eu_ai_act":
            role, risk = normalize_role_risk(requirement_code)
            if role and risk:
                scope_rows = await conn.fetch(
                    """
                    SELECT article, coverage
                    FROM euaiact_requirement_scope
                    WHERE primary_role = $1 AND risk_classification = $2
                    """,
                    role,
                    risk,
                )
                article_nums: list[str] = []
                for row in scope_rows:
                    match = re.search(r"(\\d+)", row["article"] or "")
                    if match:
                        article_nums.append(match.group(1))
                patterns = [f"%Art {num}%" for num in sorted(set(article_nums))]
                if patterns:
                    kpi_rows = await conn.fetch(
                        """
                        SELECT kd.kpi_key
                        FROM kpi_definition kd
                        WHERE kd.euaiact_clause IS NOT NULL
                          AND kd.euaiact_clause ILIKE ANY($1::text[])
                        """,
                        patterns,
                    )
                    kpi_keys = [row["kpi_key"] for row in kpi_rows if row["kpi_key"]]
        elif framework == "iso_42001":
            kpi_rows = await conn.fetch(
                """
                SELECT kd.kpi_key
                FROM kpi_definition kd
                WHERE kd.iso_42001_clause IS NOT NULL
                """
            )
            kpi_keys = [row["kpi_key"] for row in kpi_rows if row["kpi_key"]]
        elif framework == "nist_ai_rmf":
            if requirement_code:
                kpi_rows = await conn.fetch(
                    """
                    SELECT kd.kpi_key
                    FROM kpi_definition kd
                    WHERE kd.nist_clause IS NOT NULL
                      AND kd.nist_clause ILIKE $1
                    """,
                    f"%{requirement_code}%",
                )
                kpi_keys = [row["kpi_key"] for row in kpi_rows if row["kpi_key"]]
        elif framework == "internal":
            if requirement_code:
                kpi_rows = await conn.fetch(
                    """
                    SELECT k.key AS kpi_key
                    FROM kpis k
                    WHERE k.key = $1
                    """,
                    requirement_code,
                )
                kpi_keys = [row["kpi_key"] for row in kpi_rows if row["kpi_key"]]

        if project_slug and kpi_keys:
            await conn.execute(
                """
                DELETE FROM control_values
                WHERE project_slug = $1 AND entity_id = $2 AND kpi_key = ANY($3::text[])
                """,
                project_slug,
                entity_id,
                kpi_keys,
            )

        res = await conn.execute(
            "DELETE FROM ai_requirement_register WHERE id=$1", requirement_id
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Requirement not found")
    await _log_audit(
        event_type="requirement_deleted",
        actor="system",
        source_service="core-svc",
        object_type="requirement",
        object_id=requirement_id,
    )
    return {"ok": True, "id": requirement_id}


# ----------------- AIMS Scope (ISO 42001 Clause 4) -----------------
@router.get("/aims-scope")
async def get_aims_scope() -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT *
            FROM aims_scope
            ORDER BY updated_at DESC NULLS LAST, created_at DESC
            LIMIT 1
            """
        )
        if not row:
            return {"item": None}
        return {
            "item": {
                "id": row["id"],
                "scope_name": row["scope_name"],
                "scope_statement": row["scope_statement"],
                "context_internal": row["context_internal"],
                "context_external": row["context_external"],
                "interested_parties": row["interested_parties"],
                "scope_boundaries": row["scope_boundaries"],
                "lifecycle_coverage": row["lifecycle_coverage"],
                "cloud_platforms": row["cloud_platforms"],
                "regulatory_requirements": row["regulatory_requirements"],
                "isms_pms_integration": row["isms_pms_integration"],
                "exclusions": row["exclusions"],
                "owner": row["owner"],
                "status": row["status"],
                "updated_by": row["updated_by"],
                "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
            }
        }


@router.put("/aims-scope")
async def upsert_aims_scope(body: AimsScopeIn) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id FROM aims_scope ORDER BY updated_at DESC NULLS LAST, created_at DESC LIMIT 1"
        )
        if row:
            scope_id = row["id"]
            await conn.execute(
                """
                UPDATE aims_scope
                SET scope_name = COALESCE($1, scope_name),
                    scope_statement = COALESCE($2, scope_statement),
                    context_internal = COALESCE($3, context_internal),
                    context_external = COALESCE($4, context_external),
                    interested_parties = COALESCE($5::jsonb, interested_parties),
                    scope_boundaries = COALESCE($6, scope_boundaries),
                    lifecycle_coverage = COALESCE($7::jsonb, lifecycle_coverage),
                    cloud_platforms = COALESCE($8::jsonb, cloud_platforms),
                    regulatory_requirements = COALESCE($9::jsonb, regulatory_requirements),
                    isms_pms_integration = COALESCE($10, isms_pms_integration),
                    exclusions = COALESCE($11, exclusions),
                    owner = COALESCE($12, owner),
                    status = COALESCE($13, status),
                    updated_by = COALESCE($14, updated_by),
                    updated_at = NOW()
                WHERE id = $15
                """,
                body.scope_name,
                body.scope_statement,
                body.context_internal,
                body.context_external,
                _as_json(body.interested_parties),
                body.scope_boundaries,
                _as_json(body.lifecycle_coverage),
                _as_json(body.cloud_platforms),
                _as_json(body.regulatory_requirements),
                body.isms_pms_integration,
                body.exclusions,
                body.owner,
                body.status,
                body.updated_by,
                scope_id,
            )
        else:
            scope_id = str(uuid4())
            await conn.execute(
                """
                INSERT INTO aims_scope (
                    id, scope_name, scope_statement, context_internal, context_external,
                    interested_parties, scope_boundaries, lifecycle_coverage, cloud_platforms,
                    regulatory_requirements, isms_pms_integration, exclusions, owner, status,
                    updated_by, created_at, updated_at
                )
                VALUES (
                    $1,$2,$3,$4,$5,$6::jsonb,$7,$8::jsonb,$9::jsonb,$10::jsonb,
                    $11,$12,$13,COALESCE($14,'draft'),$15,NOW(),NOW()
                )
                """,
                scope_id,
                body.scope_name,
                body.scope_statement,
                body.context_internal,
                body.context_external,
                _as_json(body.interested_parties),
                body.scope_boundaries,
                _as_json(body.lifecycle_coverage),
                _as_json(body.cloud_platforms),
                _as_json(body.regulatory_requirements),
                body.isms_pms_integration,
                body.exclusions,
                body.owner,
                body.status,
                body.updated_by,
            )

    await _log_audit(
        event_type="aims_scope_updated",
        actor=body.updated_by or "system",
        source_service="core-svc",
        object_type="aims_scope",
        object_id=scope_id,
        details={"status": body.status, "owner": body.owner},
    )
    return {"ok": True, "id": scope_id}


# ----------------- Policy Manager -----------------
@router.get("/policies")
async def list_policies(
    limit: int = 50,
    offset: int = 0,
    q: Optional[str] = None,
    status: Optional[str] = None,
    locale: Optional[str] = None,
    entity_id: Optional[UUID] = None,
) -> dict:
    limit = max(1, min(int(limit), 200))
    offset = max(0, int(offset))
    clauses = []
    params = []
    pool = await get_pool()
    async with pool.acquire() as conn:
        if entity_id:
            params.append(entity_id)
            clauses.append(f"epr.entity_id = ${len(params)}")
            if q:
                params.append(f"%{q}%")
                clauses.append(f"epr.policy_title ILIKE ${len(params)}")
            if status:
                params.append(status)
                clauses.append(f"epr.policy_status = ${len(params)}")
            where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""

            total = await conn.fetchval(
                f"SELECT COUNT(*) FROM entity_policy_register epr {where_sql}",
                *params,
            )
            rows = await conn.fetch(
                f"""
                SELECT epr.policy_id AS id,
                       epr.policy_title AS title,
                       p.owner_role,
                       epr.policy_status AS status,
                       p.iso42001_requirement AS iso42001_requirement,
                       p.iso42001_status,
                       p.euaiact_requirements,
                       p.nistairmf_requirements,
                       epr.comment,
                       epr.action,
                       epr.kpi_keys,
                       p.template,
                       epr.captured_at AS created_at,
                       COALESCE(epr.updated_at, p.updated_at) AS updated_at,
                       epr.version_id AS version_id,
                       epr.version_label AS version_label,
                       epr.version_status AS version_status,
                       NULL::text AS approved_by,
                       NULL::timestamp AS approved_at,
                       NULL::timestamp AS version_created_at
                FROM entity_policy_register epr
                LEFT JOIN policies p ON p.id = epr.policy_id
                {where_sql}
                ORDER BY epr.updated_at DESC NULLS LAST,
                         p.updated_at DESC NULLS LAST,
                         epr.captured_at DESC
                LIMIT ${len(params)+1} OFFSET ${len(params)+2}
                """,
                *params,
                limit,
                offset,
            )
        else:
            normalized_locale = _normalize_locale(locale)
            join_sql = ""
            title_expr = "p.title"
            if normalized_locale:
                params.append(normalized_locale)
                locale_param = len(params)
                join_sql = f"LEFT JOIN policy_translations pt ON pt.policy_id = p.id AND pt.locale = ${locale_param}"
                title_expr = "COALESCE(pt.title, p.title)"
            if q:
                params.append(f"%{q}%")
                clauses.append(f"{title_expr} ILIKE ${len(params)}")
            if status:
                params.append(status)
                clauses.append(f"p.status = ${len(params)}")
            where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""

            total = await conn.fetchval(
                f"SELECT COUNT(*) FROM policies p {join_sql} {where_sql}",
                *params,
            )
            rows = await conn.fetch(
                f"""
                SELECT p.id, {title_expr} AS title, p.owner_role, p.status,
                       p.iso42001_requirement, p.iso42001_status, p.comment, p.action, p.template,
                       p.euaiact_requirements, p.nistairmf_requirements,
                       NULL::jsonb AS kpi_keys,
                       p.created_at, p.updated_at,
                       v.id AS version_id, v.version_label, v.status AS version_status,
                       v.approved_by, v.approved_at, v.created_at AS version_created_at
                FROM policies p
                {join_sql}
                LEFT JOIN LATERAL (
                  SELECT id, version_label, status, approved_by, approved_at, created_at
                  FROM policy_versions
                  WHERE policy_id = p.id
                  ORDER BY created_at DESC
                  LIMIT 1
                ) v ON TRUE
                {where_sql}
                ORDER BY p.updated_at DESC NULLS LAST, p.created_at DESC
                LIMIT ${len(params)+1} OFFSET ${len(params)+2}
                """,
                *params,
                limit,
                offset,
            )

    items = []
    for r in rows:
        items.append(
            {
                "id": r["id"],
                "title": r["title"],
                "owner_role": r["owner_role"],
                "status": r["status"],
                "iso42001_requirement": r["iso42001_requirement"],
                "iso42001_status": r["iso42001_status"],
                "euaiact_requirements": r["euaiact_requirements"],
                "nistairmf_requirements": r["nistairmf_requirements"],
                "comment": r["comment"],
                "action": r["action"],
                "kpi_keys": r["kpi_keys"],
                "template": str(r["template"]) if r["template"] else None,
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
                "latest_version": {
                    "id": r["version_id"],
                    "version_label": r["version_label"],
                    "status": r["version_status"],
                    "approved_by": r["approved_by"],
                    "approved_at": r["approved_at"].isoformat() if r["approved_at"] else None,
                    "created_at": r["version_created_at"].isoformat() if r["version_created_at"] else None,
                }
                if r["version_id"]
                else None,
            }
        )
    return {"items": items, "total": int(total or 0)}


@router.post("/policies/finalize")
async def finalize_policy_register(
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        existing_rows = await conn.fetch(
            """
            SELECT policy_id, kpi_keys
            FROM entity_policy_register
            WHERE entity_id = $1
            """,
            entity_id,
        )
        existing_kpis = {row["policy_id"]: row["kpi_keys"] for row in existing_rows}
        rows = await conn.fetch(
            """
            SELECT p.id, p.title, p.status, p.updated_at,
                   p.iso42001_requirement, p.iso42001_status, p.comment, p.action,
                   p.euaiact_requirements, p.nistairmf_requirements,
                   v.id AS version_id, v.version_label, v.status AS version_status
            FROM policies p
            LEFT JOIN LATERAL (
              SELECT id, version_label, status
              FROM policy_versions
              WHERE policy_id = p.id
              ORDER BY created_at DESC
              LIMIT 1
            ) v ON TRUE
            WHERE p.status = 'active'
            ORDER BY p.title
            """
        )

        await conn.execute(
            "DELETE FROM entity_policy_register WHERE entity_id = $1",
            entity_id,
        )

        for r in rows:
            await conn.execute(
                """
                INSERT INTO entity_policy_register (
                  entity_id, policy_id, policy_title, policy_status,
                  iso42001_requirements, iso42001_status, comment, action,
                  euaiact_requirements, nistairmf_requirements, kpi_keys,
                  version_id, version_label, version_status, captured_at, updated_at
                )
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,NOW(),$15)
                """,
                entity_id,
                r["id"],
                r["title"],
                r["status"],
                r["iso42001_requirement"],
                r["iso42001_status"],
                r["comment"],
                r["action"],
                r["euaiact_requirements"],
                r["nistairmf_requirements"],
                existing_kpis.get(r["id"]),
                r["version_id"],
                r["version_label"],
                r["version_status"],
                r["updated_at"],
            )
            await conn.execute(
                """
                INSERT INTO policy_review_tasks (
                  entity_id, policy_id, policy_title, due_at
                )
                VALUES ($1,$2,$3, NOW() + INTERVAL '180 days')
                ON CONFLICT (entity_id, policy_id)
                DO UPDATE SET policy_title = EXCLUDED.policy_title,
                              updated_at = NOW()
                """,
                entity_id,
                r["id"],
                r["title"],
            )

        await conn.execute(
            """
            INSERT INTO entity_policy_register_status (entity_id, status, finalized_at, updated_at)
            VALUES ($1, 'finalised', NOW(), NOW())
            ON CONFLICT (entity_id)
            DO UPDATE SET status = EXCLUDED.status,
                          finalized_at = EXCLUDED.finalized_at,
                          updated_at = EXCLUDED.updated_at
            """,
            entity_id,
        )

    await _log_audit(
        event_type="policy_register_finalized",
        actor="system",
        source_service="core-svc",
        object_type="entity_policy_register",
        object_id=str(entity_id),
        details={"saved": len(rows)},
    )

    return {"ok": True, "saved": len(rows)}


@router.get("/policies/finalize-status")
async def get_policy_register_status(
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT status, finalized_at
            FROM entity_policy_register_status
            WHERE entity_id = $1
            """,
            entity_id,
        )
    if not row:
        return {"status": "pending", "finalized_at": None}
    return {
        "status": row["status"],
        "finalized_at": row["finalized_at"].isoformat() if row["finalized_at"] else None,
    }


@router.get("/policy-control-map")
async def list_policy_control_map(
    policy_id: Optional[str] = None,
    project_slug: Optional[str] = None,
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    clauses = ["entity_id = $1"]
    params: list = [entity_id]
    if policy_id:
        params.append(policy_id)
        clauses.append(f"policy_id = ${len(params)}")
    if project_slug:
        params.append(project_slug)
        clauses.append(f"project_slug = ${len(params)}")
    where_sql = f"WHERE {' AND '.join(clauses)}"

    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"""
            SELECT policy_id, project_slug, control_id, created_at, updated_at
            FROM policy_control_map
            {where_sql}
            ORDER BY policy_id, project_slug, control_id
            """,
            *params,
        )

    items = []
    for r in rows:
        items.append(
            {
                "policy_id": r["policy_id"],
                "project_slug": r["project_slug"],
                "control_id": r["control_id"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
            }
        )
    return {"items": items}


@router.put("/policy-control-map")
async def upsert_policy_control_map(
    body: PolicyControlMapUpdate,
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> dict:
    policy_id = (body.policy_id or "").strip()
    project_slug = (body.project_slug or "").strip()
    if not policy_id or not project_slug:
        raise HTTPException(status_code=400, detail="policy_id and project_slug are required")

    control_ids = [cid.strip() for cid in body.control_ids if cid and cid.strip()]
    unique_control_ids = sorted(set(control_ids))

    pool = await get_pool()
    async with pool.acquire() as conn:
        await _get_entity_id_from_project_slug(conn, project_slug, entity_id)
        exists = await conn.fetchval(
            "SELECT 1 FROM policies WHERE id = $1",
            policy_id,
        )
        if not exists:
            raise HTTPException(status_code=404, detail="Policy not found")

        await conn.execute(
            """
            DELETE FROM policy_control_map
            WHERE entity_id = $1 AND policy_id = $2 AND project_slug = $3
            """,
            entity_id,
            policy_id,
            project_slug,
        )

        if unique_control_ids:
            await conn.executemany(
                """
                INSERT INTO policy_control_map (
                  entity_id, policy_id, project_slug, control_id, created_at, updated_at
                )
                VALUES ($1,$2,$3,$4,NOW(),NOW())
                ON CONFLICT (entity_id, policy_id, project_slug, control_id)
                DO NOTHING
                """,
                [
                    (entity_id, policy_id, project_slug, control_id)
                    for control_id in unique_control_ids
                ],
            )

    await _log_audit(
        event_type="policy_control_map_updated",
        actor="system",
        source_service="core-svc",
        object_type="policy_control_map",
        object_id=policy_id,
        details={
            "entity_id": str(entity_id),
            "project_slug": project_slug,
            "controls": len(unique_control_ids),
        },
    )

    return {"ok": True, "saved": len(unique_control_ids)}


@router.get("/policy-review-tasks")
async def list_policy_review_tasks(
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, policy_id, policy_title, status, due_at, created_at, updated_at,
                   last_reminded_at, remind_count
            FROM policy_review_tasks
            WHERE entity_id = $1
            ORDER BY due_at ASC, policy_title
            """,
            entity_id,
        )

    items = []
    for r in rows:
        items.append(
            {
                "id": str(r["id"]),
                "policy_id": r["policy_id"],
                "policy_title": r["policy_title"],
                "status": r["status"],
                "due_at": r["due_at"].isoformat() if r["due_at"] else None,
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
                "last_reminded_at": r["last_reminded_at"].isoformat() if r["last_reminded_at"] else None,
                "remind_count": int(r["remind_count"] or 0),
            }
        )
    return {"items": items}


@router.post("/policy-review-tasks/reminders")
async def send_policy_review_task_reminders(
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
    days_ahead: int = Query(14, ge=1, le=365),
    force: bool = Query(False),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        meta, emails, phones = await _get_policy_review_recipients(conn, entity_id)

        reminder_clause = ""
        if not force:
            reminder_clause = "AND (last_reminded_at IS NULL OR last_reminded_at < NOW() - INTERVAL '24 hours')"

        rows = await conn.fetch(
            f"""
            SELECT id, policy_id, policy_title, due_at
            FROM policy_review_tasks
            WHERE entity_id = $1
              AND status != 'completed'
              AND due_at <= NOW() + ($2::text || ' days')::interval
              {reminder_clause}
            ORDER BY due_at ASC, policy_title
            """,
            entity_id,
            str(days_ahead),
        )

        if not rows:
            return {"ok": True, "reminded": 0, "emails": 0, "sms": 0}

        entity_name = meta.get("entity_name")
        entity_slug = meta.get("entity_slug")
        app_url = (
            os.getenv("APP_PUBLIC_URL")
            or os.getenv("AUTH_URL")
            or os.getenv("NEXTAUTH_URL")
            or "http://localhost:3000"
        ).rstrip("/")
        policy_url = (
            f"{app_url}/{entity_slug}/scorecard/admin/governance-execution/policy-execution"
            if entity_slug
            else None
        )

        email_sent = 0
        sms_sent = 0
        reminder_ids = []

        for row in rows:
            due_at = row["due_at"].isoformat() if row.get("due_at") else None
            for email in emails:
                try:
                    send_policy_review_email(
                        email,
                        policy_title=row["policy_title"],
                        entity_name=entity_name,
                        due_at=due_at,
                        policy_url=policy_url,
                    )
                    email_sent += 1
                except EmailConfigError:
                    pass
            for phone in phones:
                try:
                    send_policy_review_sms(
                        phone,
                        policy_title=row["policy_title"],
                        entity_name=entity_name,
                        due_at=due_at,
                        policy_url=policy_url,
                    )
                    sms_sent += 1
                except SMSConfigError:
                    pass
            reminder_ids.append(row["id"])

        if reminder_ids:
            await conn.executemany(
                """
                UPDATE policy_review_tasks
                SET last_reminded_at = NOW(),
                    remind_count = COALESCE(remind_count, 0) + 1,
                    updated_at = NOW()
                WHERE id = $1
                """,
                [(rid,) for rid in reminder_ids],
            )

    await _log_audit(
        event_type="policy_review_reminders_sent",
        actor="system",
        source_service="core-svc",
        object_type="policy_review_tasks",
        object_id=str(entity_id),
        details={
            "tasks": len(reminder_ids),
            "emails": email_sent,
            "sms": sms_sent,
            "days_ahead": days_ahead,
            "force": force,
        },
    )

    return {"ok": True, "reminded": len(reminder_ids), "emails": email_sent, "sms": sms_sent}


@router.post("/policy-review-tasks/{task_id}/complete")
async def complete_policy_review_task(
    task_id: UUID,
    entity_id: UUID = Depends(get_entity_id_with_auth_editor),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE policy_review_tasks
            SET status = 'completed', updated_at = NOW()
            WHERE id = $1 AND entity_id = $2
            RETURNING id, policy_id, policy_title
            """,
            task_id,
            entity_id,
        )
    if not row:
        raise HTTPException(status_code=404, detail="Review task not found")

    await _log_audit(
        event_type="policy_review_task_completed",
        actor="system",
        source_service="core-svc",
        object_type="policy_review_tasks",
        object_id=str(row["id"]),
        details={"policy_id": row["policy_id"], "policy_title": row["policy_title"]},
    )
    return {"ok": True, "id": str(row["id"])}


@router.post("/policies")
async def create_policy(body: PolicyIn) -> dict:
    policy_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO policies (
              id, title, owner_role, status,
              iso42001_requirement, iso42001_status, euaiact_requirements, nistairmf_requirements,
              comment, action, template,
              created_at, updated_at
            )
            VALUES ($1,$2,$3,COALESCE($4,'draft'),$5,$6,$7,$8,$9,$10,$11,NOW(),NOW())
            """,
            policy_id,
            body.title,
            body.owner_role,
            body.status,
            body.iso42001_requirement,
            body.iso42001_status,
            body.euaiact_requirements,
            body.nistairmf_requirements,
            body.comment,
            body.action,
            body.template,
        )
        version_id = None
        if body.version_label or body.content:
            version_id = str(uuid4())
            await conn.execute(
                """
                INSERT INTO policy_versions (
                  id, policy_id, version_label, content, status, created_at
                )
                VALUES ($1,$2,COALESCE($3,'v1'),$4,'draft',NOW())
                """,
                version_id,
                policy_id,
                body.version_label,
                body.content,
            )

    await _log_audit(
        event_type="policy_created",
        actor="system",
        source_service="core-svc",
        object_type="policy",
        object_id=policy_id,
        details={"title": body.title},
    )
    if body.version_label or body.content:
        await _log_audit(
            event_type="policy_version_created",
            actor="system",
            source_service="core-svc",
            object_type="policy_version",
            object_id=version_id,
            details={"policy_id": policy_id, "version": body.version_label or "v1"},
        )
    return {"ok": True, "id": policy_id}


@router.put("/policies/{policy_id}")
async def update_policy(policy_id: str, body: PolicyUpdate) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE policies
            SET title = COALESCE($1, title),
                owner_role = COALESCE($2, owner_role),
                status = COALESCE($3, status),
                iso42001_requirement = COALESCE($4, iso42001_requirement),
                iso42001_status = COALESCE($5, iso42001_status),
                euaiact_requirements = COALESCE($6, euaiact_requirements),
                nistairmf_requirements = COALESCE($7, nistairmf_requirements),
                comment = COALESCE($8, comment),
                action = COALESCE($9, action),
                template = COALESCE($10, template),
                updated_at = NOW()
            WHERE id = $11
            """,
            body.title,
            body.owner_role,
            body.status,
            body.iso42001_requirement,
            body.iso42001_status,
            body.euaiact_requirements,
            body.nistairmf_requirements,
            body.comment,
            body.action,
            body.template,
            policy_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Policy not found")
    await _log_audit(
        event_type="policy_updated",
        actor="system",
        source_service="core-svc",
        object_type="policy",
        object_id=policy_id,
        details={"fields": body.model_dump(exclude_none=True)},
    )
    return {"ok": True, "id": policy_id}


@router.get("/policies/{policy_id}/versions")
async def list_policy_versions(policy_id: str, locale: Optional[str] = None) -> dict:
    normalized_locale = _normalize_locale(locale)
    pool = await get_pool()
    async with pool.acquire() as conn:
        if normalized_locale:
            rows = await conn.fetch(
                """
                SELECT v.id, v.policy_id, v.version_label,
                       COALESCE(t.content, v.content) AS content,
                       v.status, v.approved_by, v.approved_at, v.created_at
                FROM policy_versions v
                LEFT JOIN policy_version_translations t
                  ON t.version_id = v.id AND t.locale = $2
                WHERE v.policy_id = $1
                ORDER BY v.created_at DESC
                """,
                policy_id,
                normalized_locale,
            )
        else:
            rows = await conn.fetch(
                """
                SELECT id, policy_id, version_label, content, status, approved_by, approved_at, created_at
                FROM policy_versions
                WHERE policy_id = $1
                ORDER BY created_at DESC
                """,
                policy_id,
            )
    items = []
    for r in rows:
        items.append(
            {
                "id": r["id"],
                "policy_id": r["policy_id"],
                "version_label": r["version_label"],
                "content": r["content"],
                "status": r["status"],
                "approved_by": r["approved_by"],
                "approved_at": r["approved_at"].isoformat() if r["approved_at"] else None,
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            }
        )
    return {"items": items}


@router.post("/policies/{policy_id}/versions")
async def create_policy_version(policy_id: str, body: PolicyVersionIn) -> dict:
    version_id = str(uuid4())
    pool = await get_pool()
    async with pool.acquire() as conn:
        exists = await conn.fetchval("SELECT 1 FROM policies WHERE id=$1", policy_id)
        if not exists:
            raise HTTPException(status_code=404, detail="Policy not found")
        await conn.execute(
            """
            INSERT INTO policy_versions (
              id, policy_id, version_label, content, status, created_at
            )
            VALUES ($1,$2,$3,$4,'draft',NOW())
            """,
            version_id,
            policy_id,
            body.version_label,
            body.content,
        )
    await _log_audit(
        event_type="policy_version_created",
        actor="system",
        source_service="core-svc",
        object_type="policy_version",
        object_id=version_id,
        details={"policy_id": policy_id, "version": body.version_label},
    )
    return {"ok": True, "id": version_id}


@router.put("/policy-versions/{version_id}")
async def update_policy_version(version_id: str, body: PolicyVersionUpdate) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE policy_versions
            SET status = COALESCE($1, status),
                approved_by = CASE
                    WHEN $1 = 'approved' THEN COALESCE($2, approved_by)
                    ELSE approved_by
                END,
                approved_at = CASE
                    WHEN $1 = 'approved' THEN NOW()
                    ELSE approved_at
                END,
                content = COALESCE($3, content)
            WHERE id = $4
            """,
            body.status,
            body.approved_by,
            body.content,
            version_id,
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Policy version not found")
        row = await conn.fetchrow(
            "SELECT policy_id, version_label FROM policy_versions WHERE id=$1", version_id
        )

    await _log_audit(
        event_type="policy_version_updated",
        actor=body.approved_by or "system",
        source_service="core-svc",
        object_type="policy_version",
        object_id=version_id,
        details={"status": body.status, "policy_id": row["policy_id"] if row else None},
    )
    return {"ok": True, "id": version_id}


@router.get("/evidence/{evidence_id}:download")
async def evidence_download_file(evidence_id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT uri, name, mime FROM evidence WHERE id=$1", evidence_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Evidence not found")
        uri = row["uri"] or ""
        filename = row["name"] or f"evidence-{evidence_id}"
        mime = row["mime"] or "application/octet-stream"
        if uri.startswith("file://"):
            local_path = _local_path_from_uri(uri)
            if not local_path or not local_path.exists():
                raise HTTPException(status_code=404, detail="Evidence file not found")
            return FileResponse(local_path, filename=filename, media_type=mime)
        if uri.startswith("s3://"):
            key = _s3_key_from_uri(uri)
            bucket = _s3_bucket_from_uri(uri)
            if not key:
                raise HTTPException(status_code=400, detail="Invalid evidence URI")
            try:
                obj = get_object(key, bucket=bucket or None)
            except Exception:
                raise HTTPException(status_code=404, detail="Evidence file not found")
            body = obj["Body"]
            headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
            if obj.get("ContentLength") is not None:
                headers["Content-Length"] = str(obj["ContentLength"])
            if obj.get("ContentType"):
                mime = obj["ContentType"]
            return StreamingResponse(body, media_type=mime, headers=headers)
        raise HTTPException(status_code=400, detail="Evidence URI is not supported")


# ---------- Evidence metadata-only update ----------
@router.post("/projects/{project_slug}/controls/{control_id}/evidence")
async def save_control_evidence(project_slug: str, control_id: str, body: EvidenceUpdate) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            """
            UPDATE control_values
            SET evidence_source = COALESCE($3, evidence_source),
                owner_role      = COALESCE($4, owner_role),
                notes           = COALESCE($5, notes),
                updated_at      = NOW()
            WHERE project_slug = $1 AND (control_id::text=$2 OR control_id=$2::uuid)
            """,
            project_slug, control_id, body.evidence_source, body.owner_role, body.notes
        )
        if res.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="Control not found")
        
        # Invalidate LLM report cache for this project since evidence data changed
        try:
            from app.services.llm_report_cache import invalidate_cache_async
            await invalidate_cache_async(project_slug)
        except Exception as exc:
            # Don't fail the request if cache invalidation fails
            print(f"Warning: Failed to invalidate LLM report cache for {project_slug}: {exc}")

    await _log_audit(
        event_type="control_values.evidence_update",
        actor="system",
        source_service="core-svc",
        object_type="control_values",
        object_id=str(control_id),
        project_slug=project_slug,
        details={
            "evidence_source": body.evidence_source,
            "owner_role": body.owner_role,
            "notes": body.notes,
        },
    )
        
    return {"ok": True}


# =====================================================================
#  Excel Import/Export (per-project)
# =====================================================================

def _naive_or_same(x: Optional[datetime]):
    if isinstance(x, datetime):
        return x.replace(tzinfo=None) if x.tzinfo is not None else x
    return x


@router.get("/projects/{project_slug}/pillar_overrides.xlsx")
async def export_pillar_overrides(project_slug: str):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT po.pillar_key AS pillar, po.score_pct, po.maturity
            FROM pillar_overrides po
            JOIN projects p ON p.id = po.project_id
            WHERE p.slug = $1
            ORDER BY po.pillar_key
            """,
            project_slug,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "pillar_overrides"
        ws.append(["pillar", "score_pct", "maturity"])
        if rows:
            for r in rows:
                ws.append([
                    r["pillar"],
                    float(r["score_pct"]) if r["score_pct"] is not None else None,
                    int(r["maturity"]) if r["maturity"] is not None else None,
                ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

    filename = f"pillar_overrides_{project_slug}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/projects/{project_slug}/pillar_overrides")
async def import_pillar_overrides(project_slug: str, file: UploadFile = File(...)) -> dict:
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    pool = await get_pool()
    async with pool.acquire() as conn:
        project_id = await get_project_id_by_slug(conn, project_slug)

        data = await file.read()
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Unable to read workbook: {e}")

        ws = wb.active
        headers = {str(ws.cell(row=1, column=c).value or "").strip().lower(): c
                   for c in range(1, ws.max_column + 1)}
        required = {"pillar"}
        if not required.issubset(headers):
            missing = sorted(required - set(headers))
            raise HTTPException(status_code=400, detail=f"Missing required column(s): {missing}")

        n_upserts = 0
        for r in range(2, ws.max_row + 1):
            pillar = ws.cell(row=r, column=headers["pillar"]).value
            if not pillar:
                continue
            score = None
            maturity = None
            if "score_pct" in headers:
                val = ws.cell(row=r, column=headers["score_pct"]).value
                if val is not None and str(val).strip() != "":
                    score = float(val)
            if "maturity" in headers:
                val = ws.cell(row=r, column=headers["maturity"]).value
                if val is not None and str(val).strip() != "":
                    maturity = int(val)

            await conn.execute(
                """
                INSERT INTO pillar_overrides (id, project_id, pillar_key, score_pct, maturity, updated_at)
                VALUES (gen_random_uuid()::text, $1, $2, $3, $4, NOW())
                ON CONFLICT (project_id, pillar_key)
                DO UPDATE SET
                  score_pct  = EXCLUDED.score_pct,
                  maturity   = COALESCE(EXCLUDED.maturity, pillar_overrides.maturity),
                  updated_at = NOW()
                """,
                project_id, str(pillar), score, maturity
            )
            n_upserts += 1

    return {"ok": True, "upserts": n_upserts}


@router.get("/projects/{project_slug}/control_values.xlsx")
async def export_control_values(project_slug: str):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        rows = await conn.fetch(
            """
            SELECT v.control_id,
                   COALESCE(c.name, v.control_id::text) AS name,
                   v.raw_value,
                   v.normalized_pct,
                   v.observed_at,
                   v.updated_at
            FROM control_values v
            LEFT JOIN controls c ON c.id = v.control_id
            WHERE v.project_slug = $1
            ORDER BY v.control_id
            """,
            project_slug,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "control_values"
        ws.append(["control_id", "name", "raw_value", "normalized_pct", "observed_at", "updated_at"])
        for r in rows:
            ws.append([
                r["control_id"],
                r["name"],
                float(r["raw_value"]) if r["raw_value"] is not None else None,
                float(r["normalized_pct"]) if r["normalized_pct"] is not None else None,
                _naive_or_same(r["observed_at"]),
                _naive_or_same(r["updated_at"]),
            ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

    filename = f"control_values_{project_slug}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/projects/{project_slug}/control_values")
async def import_control_values(project_slug: str, file: UploadFile = File(...)) -> dict:
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        data = await file.read()
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Unable to read workbook: {e}")

        ws = wb.active
        headers = {str(ws.cell(row=1, column=c).value or "").strip().lower(): c
                   for c in range(1, ws.max_column + 1)}
        required = {"control_id", "raw_value"}
        if not required.issubset(headers):
            missing = sorted(required - set(headers))
            raise HTTPException(status_code=400, detail=f"Missing required column(s): {missing}")

        n_upserts = 0
        for r in range(2, ws.max_row + 1):
            control_id = ws.cell(row=r, column=headers["control_id"]).value
            if not control_id:
                continue
            raw = ws.cell(row=r, column=headers["raw_value"]).value
            if raw is None or str(raw).strip() == "":
                continue
            try:
                raw_f = float(raw)
            except Exception:
                raise HTTPException(status_code=400, detail=f"Row {r}: raw_value must be numeric")

            # Minimal upsert: write raw/normalized via control bounds
            await conn.execute(
                """
                INSERT INTO control_values (
                  project_slug, control_id, kpi_key, raw_value, normalized_pct, updated_at,
                  unit, target_text, target_numeric, evidence_source, owner_role, frequency
                )
                SELECT $1, $2, c.kpi_key,
                       $3,
                       CASE
                         WHEN c.norm_min IS NULL OR c.norm_max IS NULL OR c.norm_min = c.norm_max
                           THEN LEAST(GREATEST($3, 0), 100)
                         WHEN c.higher_is_better
                           THEN LEAST(GREATEST( (($3 - c.norm_min) / NULLIF(c.norm_max - c.norm_min, 0)) * 100, 0), 100)
                         ELSE LEAST(GREATEST( ((c.norm_max - $3) / NULLIF(c.norm_max - c.norm_min, 0)) * 100, 0), 100)
                       END,
                       NOW(),
                       c.unit,
                       c.target_text,
                       c.target_numeric,
                       c.evidence_source,
                       c.owner_role,
                       c.frequency
                FROM controls c
                WHERE c.id = $2
                ON CONFLICT (project_slug, control_id)
                DO UPDATE SET
                  raw_value      = EXCLUDED.raw_value,
                  normalized_pct = EXCLUDED.normalized_pct,
                  updated_at     = NOW(),
                  unit = COALESCE(control_values.unit, EXCLUDED.unit),
                  target_text = COALESCE(control_values.target_text, EXCLUDED.target_text),
                  target_numeric = COALESCE(control_values.target_numeric, EXCLUDED.target_numeric),
                  evidence_source = COALESCE(control_values.evidence_source, EXCLUDED.evidence_source),
                  owner_role = COALESCE(control_values.owner_role, EXCLUDED.owner_role),
                  frequency = COALESCE(control_values.frequency, EXCLUDED.frequency)
                """,
                project_slug, str(control_id), float(raw_f)
            )
            n_upserts += 1

        # Invalidate LLM report cache for this project since KPI data changed
        try:
            from app.services.llm_report_cache import invalidate_cache_async
            await invalidate_cache_async(project_slug)
        except Exception as exc:
            # Don't fail the request if cache invalidation fails
            print(f"Warning: Failed to invalidate LLM report cache for {project_slug}: {exc}")

    return {"ok": True, "upserts": n_upserts}


@router.get("/projects/{project_slug}/control-values/kpis")
async def list_control_value_kpis(
    project_slug: str,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        params: list = [project_slug]
        where = "cv.project_slug = $1"
        if entity_id:
            params.append(entity_id)
            where += f" AND cv.entity_id = ${len(params)}"
        rows = await conn.fetch(
            f"""
            SELECT DISTINCT cv.kpi_key
            FROM control_values cv
            WHERE {where}
            ORDER BY cv.kpi_key
            """,
            *params,
        )
    return {"items": [r["kpi_key"] for r in rows]}


@router.post("/projects/{project_slug}/control-values/sync")
async def sync_control_values(
    project_slug: str,
    body: ControlValuesSyncIn,
    entity_id: Optional[UUID] = Depends(get_entity_id_optional),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        resolved_entity_id = await _get_entity_id_from_project_slug(
            conn, project_slug, entity_id
        )
        entity_slug = await conn.fetchval(
            "SELECT slug FROM entity WHERE id = $1",
            resolved_entity_id,
        )
        desired = [k for k in (body.kpi_keys or []) if isinstance(k, str) and k.strip()]
        if not desired:
            existing_rows = await conn.fetch(
                """
                SELECT DISTINCT kpi_key
                FROM control_values
                WHERE project_slug = $1 AND entity_id = $2
                """,
                project_slug,
                resolved_entity_id,
            )
            removed = 0
            if existing_rows:
                await conn.execute(
                    """
                    DELETE FROM control_values
                    WHERE project_slug = $1 AND entity_id = $2
                    """,
                    project_slug,
                    resolved_entity_id,
                )
                removed = len(existing_rows)
            return {
                "ok": True,
                "project_slug": project_slug,
                "entity_id": str(resolved_entity_id),
                "added": 0,
                "removed": removed,
                "requested": 0,
            }

        existing_rows = await conn.fetch(
            """
            SELECT DISTINCT kpi_key
            FROM control_values
            WHERE project_slug = $1 AND entity_id = $2
            """,
            project_slug,
            resolved_entity_id,
        )
        existing_keys = {r["kpi_key"] for r in existing_rows if r["kpi_key"]}
        desired_keys = {k.strip() for k in desired}

        to_add = sorted(desired_keys - existing_keys)
        to_remove = sorted(existing_keys - desired_keys)

        added = 0
        removed = 0

        if to_add:
            result = await conn.execute(
                """
                INSERT INTO control_values (
                  entity_id, entity_slug, project_slug, control_id, kpi_key,
                  raw_value, normalized_pct, updated_at,
                  unit, target_text, target_numeric, evidence_source, owner_role, frequency,
                  failure_action, maturity_anchor_l3, current_value, as_of, notes, kpi_score
                )
                SELECT $1,
                       $2,
                       $3,
                       c.id,
                       c.kpi_key,
                       0,
                       CASE
                         WHEN c.norm_min IS NULL OR c.norm_max IS NULL OR c.norm_min = c.norm_max
                           THEN LEAST(GREATEST(0, 0), 100)
                         WHEN c.higher_is_better
                           THEN LEAST(GREATEST( ((0 - c.norm_min) / NULLIF(c.norm_max - c.norm_min, 0)) * 100, 0), 100)
                         ELSE LEAST(GREATEST( ((c.norm_max - 0) / NULLIF(c.norm_max - c.norm_min, 0)) * 100, 0), 100)
                       END,
                       NOW(),
                       c.unit,
                       c.target_text,
                       c.target_numeric,
                       c.evidence_source,
                       c.owner_role,
                       c.frequency,
                       c.failure_action,
                       c.maturity_anchor_l3,
                       c.current_value,
                       c.as_of,
                       c.notes,
                       c.kpi_score
                FROM controls c
                WHERE c.kpi_key = ANY($4::text[])
                ON CONFLICT (project_slug, control_id)
                DO NOTHING
                """,
                resolved_entity_id,
                entity_slug,
                project_slug,
                to_add,
            )
            try:
                added = int(result.split(" ")[2])
            except Exception:
                added = len(to_add)

        await conn.execute(
            """
            UPDATE control_values v
            SET entity_id = COALESCE(v.entity_id, $1),
                entity_slug = COALESCE(v.entity_slug, $2),
                unit = COALESCE(v.unit, c.unit),
                target_text = COALESCE(v.target_text, c.target_text),
                target_numeric = COALESCE(v.target_numeric, c.target_numeric),
                evidence_source = COALESCE(v.evidence_source, c.evidence_source),
                owner_role = COALESCE(v.owner_role, c.owner_role),
                frequency = COALESCE(v.frequency, c.frequency),
                failure_action = COALESCE(v.failure_action, c.failure_action),
                maturity_anchor_l3 = COALESCE(v.maturity_anchor_l3, c.maturity_anchor_l3),
                current_value = COALESCE(v.current_value, c.current_value),
                as_of = COALESCE(v.as_of, c.as_of),
                notes = COALESCE(v.notes, c.notes),
                kpi_score = COALESCE(v.kpi_score, c.kpi_score)
            FROM controls c
            WHERE v.project_slug = $3
              AND v.control_id = c.id
              AND (v.entity_id IS NULL OR v.entity_slug IS NULL
                   OR v.unit IS NULL OR v.target_text IS NULL OR v.target_numeric IS NULL
                   OR v.evidence_source IS NULL OR v.owner_role IS NULL
                   OR v.frequency IS NULL OR v.failure_action IS NULL
                   OR v.maturity_anchor_l3 IS NULL OR v.current_value IS NULL
                   OR v.as_of IS NULL OR v.notes IS NULL OR v.kpi_score IS NULL)
            """,
            resolved_entity_id,
            entity_slug,
            project_slug,
        )

        if to_remove:
            result = await conn.execute(
                """
                DELETE FROM control_values
                WHERE project_slug = $1 AND entity_id = $2 AND kpi_key = ANY($3::text[])
                """,
                project_slug,
                resolved_entity_id,
                to_remove,
            )
            try:
                removed = int(result.split(" ")[1])
            except Exception:
                removed = len(to_remove)

        try:
            from app.services.llm_report_cache import invalidate_cache_async
            await invalidate_cache_async(project_slug)
        except Exception as exc:
            print(f"Warning: Failed to invalidate LLM report cache for {project_slug}: {exc}")

    await _log_audit(
        event_type="control_values.sync",
        actor="system",
        source_service="core-svc",
        object_type="control_values",
        object_id=project_slug,
        project_slug=project_slug,
        details={
            "entity_id": str(resolved_entity_id),
            "added": added,
            "removed": removed,
            "requested": len(desired_keys),
        },
    )

    return {
        "ok": True,
        "project_slug": project_slug,
        "entity_id": str(resolved_entity_id),
        "added": added,
        "removed": removed,
        "requested": len(desired_keys),
    }


# ---------- Control Values Exec (AI Control Register) ----------
@router.get("/projects/{project_slug}/control-values-exec")
async def list_control_values_exec(
    project_slug: str,
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        resolved_entity_id = await _get_entity_id_from_project_slug(
            conn, project_slug, entity_id
        )
        entity_slug = await conn.fetchval(
            "SELECT slug FROM entity WHERE id = $1",
            resolved_entity_id,
        )
        if not entity_slug:
            raise HTTPException(status_code=400, detail="Entity slug is missing")

        rows = await conn.fetch(
            """
            SELECT v.project_slug,
                   v.control_id::text AS control_id,
                   COALESCE(c.kpi_key, v.kpi_key) AS kpi_key,
                   k.name AS kpi_name,
                   c.name AS control_name,
                   v.owner_role,
                   COALESCE(v.unit, c.unit) AS unit,
                   v.target_text,
                   v.evidence_source,
                   e.designated_owner_name,
                   e.designated_owner_email,
                   e.due_date,
                   COALESCE(
                     e.frequency,
                     c.frequency,
                     (
                       SELECT c2.frequency
                       FROM controls c2
                       WHERE c2.kpi_key = COALESCE(c.kpi_key, v.kpi_key)
                         AND c2.frequency IS NOT NULL
                       LIMIT 1
                     )
                   ) AS frequency,
                   e.reminder_day,
                   e.reminder_count,
                   e.designated_owner_manager,
                   e.designated_owner_manager_email,
                   e.provide_url,
                   e.forward_request,
                   e.forward_email,
                   e.comment_text,
                   ev.status AS evidence_status,
                   ev.uri AS evidence_url
            FROM control_values v
            LEFT JOIN controls c ON c.id = v.control_id
            LEFT JOIN kpis k ON k.key = COALESCE(c.kpi_key, v.kpi_key)
            LEFT JOIN control_values_exec e
              ON e.entity_slug = $3
             AND e.project_slug = v.project_slug
             AND e.control_id = v.control_id
            LEFT JOIN LATERAL (
                SELECT status, uri
                FROM evidence ev
                WHERE ev.project_slug = v.project_slug
                  AND ev.control_id = v.control_id
                ORDER BY COALESCE(ev.updated_at, ev.created_at) DESC
                LIMIT 1
            ) ev ON true
            WHERE v.project_slug = $1
              AND v.entity_id = $2
            ORDER BY COALESCE(c.kpi_key, v.kpi_key), v.control_id
            """,
            project_slug,
            resolved_entity_id,
            entity_slug,
        )

    def _iso(val):
        return val.isoformat() if val else None

    items = []
    for r in rows:
        items.append(
            {
                "project_slug": r["project_slug"],
                "control_id": r["control_id"],
                "kpi_key": r["kpi_key"],
                "kpi_name": r["kpi_name"],
                "control_name": r["control_name"],
                "owner_role": r["owner_role"],
                "unit": r["unit"],
                "target_text": r["target_text"],
                "evidence_source": r["evidence_source"],
                "frequency": r["frequency"],
                "designated_owner_name": r["designated_owner_name"],
                "designated_owner_email": r["designated_owner_email"],
                "due_date": _iso(r["due_date"]),
                "reminder_day": r["reminder_day"],
                "reminder_count": r["reminder_count"],
                "designated_owner_manager": r["designated_owner_manager"],
                "designated_owner_manager_email": r["designated_owner_manager_email"],
                "provide_url": r["provide_url"],
                "forward_request": r["forward_request"],
                "forward_email": r["forward_email"],
                "comment_text": r["comment_text"],
                "evidence_status": r["evidence_status"],
                "evidence_url": r["evidence_url"],
            }
        )

    return {
        "entity_id": str(resolved_entity_id),
        "entity_slug": entity_slug,
        "project_slug": project_slug,
        "items": items,
    }


@router.put("/projects/{project_slug}/control-values-exec")
async def upsert_control_values_exec(
    project_slug: str,
    body: ControlValuesExecUpdateIn,
    background_tasks: BackgroundTasks,
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
    force_notify: bool = Query(False),
) -> dict:
    if not body.items:
        raise HTTPException(status_code=400, detail="items is required")

    def _normalize(val: Optional[str]) -> str:
        return (val or "").strip().lower()

    def _should_notify_owner(prev: dict | None, item: ControlValuesExecRowIn) -> bool:
        if not item.designated_owner_email or not item.designated_owner_name:
            return False
        new_email = _normalize(item.designated_owner_email)
        new_name = _normalize(item.designated_owner_name)
        if not new_email or not new_name:
            return False
        if not prev:
            return True
        prev_email = _normalize(prev.get("designated_owner_email"))
        prev_name = _normalize(prev.get("designated_owner_name"))
        prev_due = prev.get("due_date").isoformat() if prev.get("due_date") else ""
        next_due = item.due_date.isoformat() if item.due_date else ""
        return prev_email != new_email or prev_name != new_name or prev_due != next_due

    def _should_notify_manager(prev: dict | None, item: ControlValuesExecRowIn) -> bool:
        if not item.designated_owner_manager_email:
            return False
        new_email = _normalize(item.designated_owner_manager_email)
        if not new_email:
            return False
        if not prev:
            return True
        prev_email = _normalize(prev.get("designated_owner_manager_email"))
        prev_due = prev.get("due_date").isoformat() if prev.get("due_date") else ""
        next_due = item.due_date.isoformat() if item.due_date else ""
        if prev_email != new_email or prev_due != next_due:
            return True
        new_mgr = (item.designated_owner_manager or "").strip()
        prev_mgr = (prev.get("designated_owner_manager") or "").strip()
        return bool(new_mgr) and _normalize(prev_mgr) != _normalize(new_mgr)

    pool = await get_pool()
    emails_queued = 0
    async with pool.acquire() as conn:
        resolved_entity_id = await _get_entity_id_from_project_slug(
            conn, project_slug, entity_id
        )
        entity_slug = await conn.fetchval(
            "SELECT slug FROM entity WHERE id = $1",
            resolved_entity_id,
        )
        if not entity_slug:
            raise HTTPException(status_code=400, detail="Entity slug is missing")

        upserts = 0
        for item in body.items:
            prev = await conn.fetchrow(
                """
                SELECT designated_owner_name,
                       designated_owner_email,
                       designated_owner_manager,
                       designated_owner_manager_email,
                       due_date,
                       forward_request,
                       forward_email,
                       comment_text
                FROM control_values_exec
                WHERE entity_slug = $1
                  AND project_slug = $2
                  AND (control_id::text = $3 OR control_id = $3::uuid)
                """,
                entity_slug,
                project_slug,
                item.control_id,
            )
            should_notify_owner = _should_notify_owner(prev, item)
            should_notify_manager = _should_notify_manager(prev, item)
            should_forward = False
            if item.forward_request and (item.forward_email or "").strip():
                if not prev:
                    should_forward = True
                else:
                    prev_flag = bool(prev.get("forward_request"))
                    prev_email = _normalize(prev.get("forward_email"))
                    next_email = _normalize(item.forward_email)
                    should_forward = (not prev_flag) or (prev_email != next_email)
            if force_notify:
                should_notify_owner = bool(
                    (item.designated_owner_email or "").strip()
                    and (item.designated_owner_name or "").strip()
                )
                should_notify_manager = bool(
                    (item.designated_owner_manager_email or "").strip()
                )
                should_forward = bool(
                    item.forward_request and (item.forward_email or "").strip()
                )
            row = await conn.fetchrow(
                """
                INSERT INTO control_values_exec (
                  entity_slug,
                  project_slug,
                  control_id,
                  kpi_key,
                  owner_role,
                  designated_owner_name,
                  designated_owner_email,
                  due_date,
                  frequency,
                  reminder_day,
                  reminder_count,
                  designated_owner_manager,
                  designated_owner_manager_email,
                  provide_url,
                  forward_request,
                  forward_email,
                  comment_text,
                  updated_at
                )
                SELECT $1,
                       $2,
                       v.control_id,
                       COALESCE(c.kpi_key, v.kpi_key),
                       $4,
                       $5,
                       $6,
                       $7,
                       COALESCE(
                         $8,
                         c.frequency,
                         (
                           SELECT c2.frequency
                           FROM controls c2
                           WHERE c2.kpi_key = COALESCE(c.kpi_key, v.kpi_key)
                             AND c2.frequency IS NOT NULL
                           LIMIT 1
                         )
                       ),
                       $9,
                       $10,
                       $11,
                       $12,
                       $13,
                       $15,
                       $16,
                       $17,
                       NOW()
                FROM control_values v
                LEFT JOIN controls c ON c.id = v.control_id
                WHERE v.project_slug = $2
                  AND v.entity_id = $3
                  AND (v.control_id::text = $14 OR v.control_id = $14::uuid)
                ON CONFLICT (entity_slug, project_slug, control_id)
                DO UPDATE SET
                  kpi_key = EXCLUDED.kpi_key,
                  owner_role = EXCLUDED.owner_role,
                  designated_owner_name = EXCLUDED.designated_owner_name,
                  designated_owner_email = EXCLUDED.designated_owner_email,
                  due_date = EXCLUDED.due_date,
                  frequency = EXCLUDED.frequency,
                  reminder_day = EXCLUDED.reminder_day,
                  reminder_count = EXCLUDED.reminder_count,
                  designated_owner_manager = EXCLUDED.designated_owner_manager,
                  designated_owner_manager_email = EXCLUDED.designated_owner_manager_email,
                  provide_url = EXCLUDED.provide_url,
                  forward_request = EXCLUDED.forward_request,
                  forward_email = EXCLUDED.forward_email,
                  comment_text = EXCLUDED.comment_text,
                  updated_at = NOW()
                RETURNING 1
                """,
                entity_slug,
                project_slug,
                resolved_entity_id,
                item.owner_role,
                item.designated_owner_name,
                item.designated_owner_email,
                item.due_date,
                item.frequency,
                item.reminder_day,
                item.reminder_count,
                item.designated_owner_manager,
                item.designated_owner_manager_email,
                item.provide_url,
                item.control_id,
                item.forward_request,
                item.forward_email,
                item.comment_text,
            )
            if row:
                upserts += 1
                await _log_audit(
                    event_type="control_values_exec.upsert",
                    source_service="core-svc",
                    object_type="control_values_exec",
                    object_id=str(item.control_id),
                    project_slug=project_slug,
                    details={
                        "kpi_key": item.kpi_key,
                        "designated_owner_name": item.designated_owner_name,
                        "designated_owner_email": item.designated_owner_email,
                        "due_date": item.due_date.isoformat() if item.due_date else None,
                        "forward_request": item.forward_request,
                        "forward_email": item.forward_email,
                        "comment_text": item.comment_text,
                    },
                )
                if (should_notify_owner or should_notify_manager) and background_tasks is not None:
                    details = await conn.fetchrow(
                        """
                        SELECT COALESCE(c.kpi_key, v.kpi_key) AS kpi_key,
                               k.name AS kpi_name,
                               c.name AS control_name,
                               v.owner_role,
                               v.target_text,
                               v.evidence_source
                        FROM control_values v
                        LEFT JOIN controls c ON c.id = v.control_id
                        LEFT JOIN kpis k ON k.key = COALESCE(c.kpi_key, v.kpi_key)
                        WHERE v.project_slug = $1
                          AND v.entity_id = $2
                          AND (v.control_id::text = $3 OR v.control_id = $3::uuid)
                        """,
                        project_slug,
                        resolved_entity_id,
                        item.control_id,
                    )
                    if details and item.designated_owner_email and item.designated_owner_name:
                        base_url = (
                            os.getenv("APP_BASE_URL")
                            or os.getenv("AUTH_URL")
                            or "http://localhost:3000"
                        )
                        evidence_url = _build_evidence_url(base_url, entity_slug, project_slug)
                        if should_notify_owner:
                            background_tasks.add_task(
                                _send_control_assignment_email_safe,
                                to_email=item.designated_owner_email,
                                recipient_name=item.designated_owner_name,
                                designated_owner_name=item.designated_owner_name,
                                entity_slug=entity_slug,
                                project_slug=project_slug,
                                kpi_key=details["kpi_key"],
                                kpi_name=details["kpi_name"],
                                control_name=details["control_name"],
                                owner_role=details["owner_role"],
                                target_text=details["target_text"],
                                evidence_source=details["evidence_source"],
                                due_date=item.due_date,
                                comment_text=item.comment_text,
                                evidence_url=evidence_url,
                            )
                            emails_queued += 1
                        if should_notify_manager and item.designated_owner_manager_email:
                            manager_email = item.designated_owner_manager_email
                            if _normalize(manager_email) != _normalize(item.designated_owner_email):
                                background_tasks.add_task(
                                    _send_control_assignment_email_safe,
                                    to_email=manager_email,
                                    recipient_name=item.designated_owner_manager or "Manager",
                                    designated_owner_name=item.designated_owner_name,
                                    entity_slug=entity_slug,
                                    project_slug=project_slug,
                                    kpi_key=details["kpi_key"],
                                    kpi_name=details["kpi_name"],
                                    control_name=details["control_name"],
                                    owner_role=details["owner_role"],
                                    target_text=details["target_text"],
                                    evidence_source=details["evidence_source"],
                                    due_date=item.due_date,
                                    comment_text=item.comment_text,
                                    evidence_url=evidence_url,
                                )
                                emails_queued += 1
                        if should_forward and item.forward_email:
                            background_tasks.add_task(
                                _send_control_assignment_email_safe,
                                to_email=item.forward_email,
                                recipient_name="Forwarded recipient",
                                designated_owner_name=item.designated_owner_name,
                                entity_slug=entity_slug,
                                project_slug=project_slug,
                                kpi_key=details["kpi_key"],
                                kpi_name=details["kpi_name"],
                                control_name=details["control_name"],
                                owner_role=details["owner_role"],
                                target_text=details["target_text"],
                                evidence_source=details["evidence_source"],
                                due_date=item.due_date,
                                comment_text=item.comment_text,
                                evidence_url=evidence_url,
                            )
                            emails_queued += 1

    return {"ok": True, "upserts": upserts, "emails_queued": emails_queued}


@router.post("/projects/{project_slug}/control-values-exec/forward-email")
async def send_forward_control_email(
    project_slug: str,
    body: dict,
    entity_id: UUID = Depends(get_entity_id_with_auth_viewer),
) -> dict:
    control_id = body.get("control_id")
    if not control_id:
        raise HTTPException(status_code=400, detail="control_id is required")

    pool = await get_pool()
    async with pool.acquire() as conn:
        resolved_entity_id = await _get_entity_id_from_project_slug(
            conn, project_slug, entity_id
        )
        entity_slug = await conn.fetchval(
            "SELECT slug FROM entity WHERE id = $1",
            resolved_entity_id,
        )
        if not entity_slug:
            raise HTTPException(status_code=400, detail="Entity slug is missing")

        row = await conn.fetchrow(
            """
            SELECT e.forward_request,
                   e.forward_email,
                   e.comment_text,
                   e.due_date,
                   COALESCE(c.kpi_key, v.kpi_key) AS kpi_key,
                   k.name AS kpi_name,
                   c.name AS control_name,
                   v.owner_role,
                   v.target_text,
                   v.evidence_source,
                   e.designated_owner_name
            FROM control_values v
            LEFT JOIN controls c ON c.id = v.control_id
            LEFT JOIN kpis k ON k.key = COALESCE(c.kpi_key, v.kpi_key)
            LEFT JOIN control_values_exec e
              ON e.entity_slug = $3
             AND e.project_slug = v.project_slug
             AND e.control_id = v.control_id
            WHERE v.project_slug = $1
              AND v.entity_id = $2
              AND (v.control_id::text = $4 OR v.control_id = $4::uuid)
            LIMIT 1
            """,
            project_slug,
            resolved_entity_id,
            entity_slug,
            control_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Control not found")

        forward_email = (row["forward_email"] or "").strip()
        if not row["forward_request"] or not forward_email:
            raise HTTPException(status_code=400, detail="Forward request is not enabled")

        base_url = (
            os.getenv("APP_BASE_URL")
            or os.getenv("AUTH_URL")
            or "http://localhost:3000"
        )
        evidence_url = _build_evidence_url(base_url, entity_slug, project_slug)
        due_str = row["due_date"].isoformat() if row["due_date"] else None

    # Send outside connection
    _send_control_assignment_email_safe(
        to_email=forward_email,
        recipient_name="Forwarded recipient",
        designated_owner_name=row["designated_owner_name"],
        entity_slug=entity_slug,
        project_slug=project_slug,
        kpi_key=row["kpi_key"],
        kpi_name=row["kpi_name"],
        control_name=row["control_name"],
        owner_role=row["owner_role"],
        target_text=row["target_text"],
        evidence_source=row["evidence_source"],
        due_date=due_str,
        comment_text=row["comment_text"],
        evidence_url=evidence_url,
    )

    await _log_audit(
        event_type="control_values_exec.forward_email",
        source_service="core-svc",
        object_type="control_values_exec",
        object_id=str(control_id),
        project_slug=project_slug,
        details={
            "entity_id": str(resolved_entity_id),
            "forward_email": forward_email,
        },
    )

    return {"ok": True, "sent": True, "to": forward_email}


def _send_control_assignment_email_safe(**kwargs) -> None:
    try:
        send_control_assignment_email(**kwargs)
    except EmailConfigError:
        # Email not configured; skip without failing request
        email_logger.warning("Email not sent: EMAIL_SERVER not configured")
    except Exception:
        # Avoid surfacing email errors to API clients
        email_logger.exception("Email send failed")


# ---------- KPI Score Recompute (manual trigger + audit) ----------
@router.post("/kpi-recompute")
async def post_kpi_recompute(
    project_slug: Optional[str] = Query(default=None, description="Project slug to recompute; omit for all projects"),
) -> dict:
    """
    Manually trigger KPI score recompute (and pillar recompute) for all projects
    or a single project. Logs the action to the audit log.
    """
    from app.score_engine import recompute_all

    scope = project_slug or "ALL"
    try:
        result = await recompute_all(
            project_id_or_none=project_slug,
            pool=None,
            verbose=False,
        )
    except Exception as exc:
        await _log_audit(
            event_type="kpi_recompute.failed",
            source_service="core-svc",
            object_type="kpi_recompute",
            object_id=scope,
            project_slug=project_slug,
            details={"scope": scope, "error": str(exc)},
        )
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    await _log_audit(
        event_type="kpi_recompute.completed",
        source_service="core-svc",
        object_type="kpi_recompute",
        object_id=scope,
        project_slug=project_slug,
        details={
            "scope": scope,
            "status": result.get("status"),
            "kpis": result.get("kpis"),
            "pillars": result.get("pillars"),
        },
    )

    return result
