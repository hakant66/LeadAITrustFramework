# app/admin.py
from __future__ import annotations

import io
from datetime import date, datetime, timezone
from typing import List, Optional

import asyncpg
from fastapi import APIRouter, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from .scorecard import (
    get_pool,
    ensure_schema,              # ensures controls + control_values demo tables
    get_project_id_by_slug,     # helper to resolve project_id
    upsert_control_value,       # reuse normalization logic
)

router = APIRouter(prefix="/admin", tags=["admin"])


def _clean_optional_str(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    trimmed = value.strip()
    return trimmed if trimmed else None


# ---------- Models ----------
class ControlIn(BaseModel):
    control_id: str = Field(min_length=1)
    name: str = Field(min_length=1)
    pillar: Optional[str] = None
    unit: Optional[str] = None
    norm_min: Optional[float] = None
    norm_max: Optional[float] = None
    higher_is_better: bool = True
    weight: float = 1.0


class ControlOut(ControlIn):
    pass


class ProjectIn(BaseModel):
    slug: str = Field(min_length=1)
    name: str = Field(min_length=1)
    risk_level: Optional[str] = None
    target_threshold: float = 0.75  # 0..1
    priority: Optional[str] = None
    sponsor: Optional[str] = None
    owner: Optional[str] = None
    creation_date: Optional[date] = None
    update_date: Optional[datetime] = None


class ProjectOut(ProjectIn):
    id: str


# ---------- Controls CRUD (global) ----------
@router.get("/controls", response_model=List[ControlOut])
async def list_controls() -> List[ControlOut]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        rows = await conn.fetch(
            """
            SELECT control_id, name, pillar, unit, norm_min, norm_max, higher_is_better, weight
            FROM controls
            ORDER BY pillar NULLS LAST, control_id
            """
        )
        return [ControlOut(**dict(r)) for r in rows]


@router.post("/controls", response_model=ControlOut)
async def create_control(body: ControlIn) -> ControlOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        try:
            await conn.execute(
                """
                INSERT INTO controls
                  (control_id, name, pillar, unit, norm_min, norm_max, higher_is_better, weight)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
                """,
                body.control_id, body.name, body.pillar, body.unit,
                body.norm_min, body.norm_max, body.higher_is_better, body.weight,
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(status_code=409, detail="control_id already exists")
        return ControlOut(**body.model_dump())


@router.put("/controls/{control_id}", response_model=ControlOut)
async def update_control(control_id: str, body: ControlIn) -> ControlOut:
    if control_id != body.control_id:
        raise HTTPException(status_code=400, detail="control_id mismatch between path and body")

    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        result = await conn.execute(
            """
            UPDATE controls
            SET name=$2, pillar=$3, unit=$4, norm_min=$5, norm_max=$6, higher_is_better=$7, weight=$8
            WHERE control_id=$1
            """,
            body.control_id, body.name, body.pillar, body.unit,
            body.norm_min, body.norm_max, body.higher_is_better, body.weight,
        )
        if result.split()[-1] == "0":
            raise HTTPException(status_code=404, detail="control not found")
        return ControlOut(**body.model_dump())


@router.delete("/controls/{control_id}")
async def delete_control(control_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        # explicit delete for clarity; control_values has FK ON DELETE CASCADE in demo schema
        await conn.execute("DELETE FROM control_values WHERE control_id=$1", control_id)
        res = await conn.execute("DELETE FROM controls WHERE control_id=$1", control_id)
        if int(res.split()[-1]) == 0:
            raise HTTPException(status_code=404, detail="control not found")
        return {"deleted": control_id}


# ---------- Projects (create/upsert; used by global 'Capture AI Project' page) ----------
@router.post("/projects", response_model=ProjectOut)
async def create_project(body: ProjectIn) -> ProjectOut:
    pool = await get_pool()
    async with pool.acquire() as conn:
        existing = await conn.fetchrow(
            """
            SELECT id, risk_level, priority, sponsor, owner, creation_date
            FROM projects
            WHERE slug=$1
            """,
            body.slug,
        )

        base_risk = _clean_optional_str(body.risk_level) or (existing["risk_level"] if existing and existing["risk_level"] else "medium")
        priority = _clean_optional_str(body.priority) or base_risk
        sponsor = _clean_optional_str(body.sponsor)
        owner = _clean_optional_str(body.owner)

        if existing:
            if sponsor is None and existing["sponsor"] is not None and body.sponsor is None:
                sponsor = existing["sponsor"]
            if owner is None and existing["owner"] is not None and body.owner is None:
                owner = existing["owner"]
            creation_date = (
                body.creation_date if body.creation_date is not None else existing["creation_date"]
            )
        else:
            creation_date = body.creation_date or date.today()

        update_ts = body.update_date or datetime.now(timezone.utc)

        if existing:
            await conn.execute(
                """
                UPDATE projects
                SET name=$2,
                    risk_level=$3,
                    target_threshold=$4,
                    priority=$5,
                    sponsor=$6,
                    owner=$7,
                    creation_date=$8,
                    update_date=$9
                WHERE slug=$1
                """,
                body.slug,
                body.name,
                base_risk,
                body.target_threshold,
                priority,
                sponsor,
                owner,
                creation_date,
                update_ts,
            )
            row = await conn.fetchrow(
                """
                SELECT id, slug, name, risk_level, target_threshold, priority, sponsor, owner, creation_date, update_date
                FROM projects
                WHERE slug=$1
                """,
                body.slug,
            )
            return ProjectOut(**dict(row))

        try:
            row = await conn.fetchrow(
                """
                INSERT INTO projects (
                    id,
                    slug,
                    name,
                    risk_level,
                    target_threshold,
                    priority,
                    sponsor,
                    owner,
                    creation_date,
                    update_date
                )
                VALUES (gen_random_uuid()::text, $1, $2, $3, $4, $5, $6, $7, $8, $9)
                RETURNING id, slug, name, risk_level, target_threshold, priority, sponsor, owner, creation_date, update_date
                """,
                body.slug,
                body.name,
                base_risk,
                body.target_threshold,
                priority,
                sponsor,
                owner,
                creation_date,
                update_ts,
            )
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to create project: {e}")
        return ProjectOut(**dict(row))


# =====================================================================
#  Excel Import/Export (per-project)
#  A) Pillar Overrides (pillar_overrides.xlsx)
#  B) Control Values   (control_values.xlsx)
# =====================================================================

# ---------- A) Pillar Overrides ----------
@router.get("/projects/{project_slug}/pillar_overrides.xlsx")
async def export_pillar_overrides(project_slug: str):
    """
    Exports columns:
      pillar (pillar_key), score_pct (0..100), maturity (1..5)
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT po.pillar_key AS pillar, po.score_pct, po.maturity
            FROM pillar_overrides po
            JOIN projects p ON p.id = po.project_id
            WHERE p.slug = $1
            ORDER BY po.pillar_key
            """,
            project_slug,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "pillar_overrides"
        ws.append(["pillar", "score_pct", "maturity"])
        if rows:
            for r in rows:
                ws.append([
                    r["pillar"],
                    float(r["score_pct"]) if r["score_pct"] is not None else None,
                    int(r["maturity"]) if r["maturity"] is not None else None,
                ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

    filename = f"pillar_overrides_{project_slug}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/projects/{project_slug}/pillar_overrides")
async def import_pillar_overrides(project_slug: str, file: UploadFile = File(...)) -> dict:
    """
    Accepts an .xlsx with a sheet containing columns:
      pillar, score_pct, maturity
    Upserts into pillar_overrides for the given project.
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    pool = await get_pool()
    async with pool.acquire() as conn:
        project_id = await get_project_id_by_slug(conn, project_slug)

        data = await file.read()
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Unable to read workbook: {e}")

        ws = wb.active
        headers = {str(ws.cell(row=1, column=c).value or "").strip().lower(): c
                   for c in range(1, ws.max_column + 1)}
        required = {"pillar"}
        if not required.issubset(headers):
            missing = sorted(required - set(headers))
            raise HTTPException(status_code=400, detail=f"Missing required column(s): {missing}")

        n_upserts = 0
        for r in range(2, ws.max_row + 1):
            pillar = ws.cell(row=r, column=headers["pillar"]).value
            if not pillar:
                continue
            score = None
            maturity = None
            if "score_pct" in headers:
                val = ws.cell(row=r, column=headers["score_pct"]).value
                if val is not None and str(val).strip() != "":
                    score = float(val)
            if "maturity" in headers:
                val = ws.cell(row=r, column=headers["maturity"]).value
                if val is not None and str(val).strip() != "":
                    maturity = int(val)

            await conn.execute(
                """
                INSERT INTO pillar_overrides (id, project_id, pillar_key, score_pct, maturity, updated_at)
                VALUES (gen_random_uuid()::text, $1, $2, $3, $4, NOW())
                ON CONFLICT (project_id, pillar_key)
                DO UPDATE SET
                  score_pct  = EXCLUDED.score_pct,
                  maturity   = COALESCE(EXCLUDED.maturity, pillar_overrides.maturity),
                  updated_at = NOW()
                """,
                project_id, str(pillar), score, maturity
            )
            n_upserts += 1

    return {"ok": True, "upserts": n_upserts}


# ---------- B) Control Values ----------
def _naive_or_same(x: Optional[datetime]):
    """
    Excel cells cannot hold tz-aware datetimes.
    - If x is a datetime with tzinfo -> return x.replace(tzinfo=None)
    - If x is a naive datetime -> return as-is
    - Otherwise -> return as-is (e.g., None or string)
    """
    if isinstance(x, datetime):
        return x.replace(tzinfo=None) if x.tzinfo is not None else x
    return x


@router.get("/projects/{project_slug}/control_values.xlsx")
async def export_control_values(project_slug: str):
    """
    Exports columns:
      control_id, name, raw_value, normalized_pct, observed_at, updated_at
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)  # control_values table
        rows = await conn.fetch(
            """
            SELECT v.control_id,
                   COALESCE(c.name, v.control_id) AS name,
                   v.raw_value,
                   v.normalized_pct,
                   v.observed_at,
                   v.updated_at
            FROM control_values v
            LEFT JOIN controls c ON c.control_id = v.control_id
            WHERE v.project_slug = $1
            ORDER BY v.control_id
            """,
            project_slug,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "control_values"
        ws.append(["control_id", "name", "raw_value", "normalized_pct", "observed_at", "updated_at"])
        for r in rows:
            ws.append([
                r["control_id"],
                r["name"],
                float(r["raw_value"]) if r["raw_value"] is not None else None,
                float(r["normalized_pct"]) if r["normalized_pct"] is not None else None,
                _naive_or_same(r["observed_at"]),
                _naive_or_same(r["updated_at"]),
            ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

    filename = f"control_values_{project_slug}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/projects/{project_slug}/control_values")
async def import_control_values(project_slug: str, file: UploadFile = File(...)) -> dict:
    """
    Accepts an .xlsx with columns:
      control_id, raw_value  (normalized_pct optional; will be recomputed anyway)
    Each row is upserted via upsert_control_value() to apply normalization.
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    pool = await get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        data = await file.read()
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Unable to read workbook: {e}")

        ws = wb.active
        headers = {str(ws.cell(row=1, column=c).value or "").strip().lower(): c
                   for c in range(1, ws.max_column + 1)}
        required = {"control_id", "raw_value"}
        if not required.issubset(headers):
            missing = sorted(required - set(headers))
            raise HTTPException(status_code=400, detail=f"Missing required column(s): {missing}")

        n_upserts = 0
        for r in range(2, ws.max_row + 1):
            control_id = ws.cell(row=r, column=headers["control_id"]).value
            if not control_id:
                continue
            raw = ws.cell(row=r, column=headers["raw_value"]).value
            if raw is None or str(raw).strip() == "":
                continue
            try:
                raw_f = float(raw)
            except Exception:
                raise HTTPException(status_code=400, detail=f"Row {r}: raw_value must be numeric")

            await upsert_control_value(conn, project_slug, str(control_id), raw_f)
            n_upserts += 1

    return {"ok": True, "upserts": n_upserts}
